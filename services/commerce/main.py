"""
Commerce Service - Aarya Clothing
Product Management, Categories, Cart, Orders, Inventory

This service handles:
- Product catalog management
- Category hierarchy management
- Shopping cart operations
- Order processing
- Inventory management
- Image uploads to Cloudflare R2
- Reviews and ratings
- Address management
- Returns and refunds
"""
import logging
import asyncio
import os
from contextlib import asynccontextmanager
from datetime import datetime, timezone
from starlette.concurrency import run_in_threadpool
from fastapi import FastAPI, Depends, HTTPException, status, Request, UploadFile, File, Query
from fastapi.responses import StreamingResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import text, func
from typing import Optional, List
from decimal import Decimal
from pydantic import BaseModel
import json

logger = logging.getLogger(__name__)

from core.config import settings
from core.redis_client import redis_client
from core.advanced_cache import cache, cached
from database.database import get_db, init_db, SessionLocal, get_db_context
from search.meilisearch_client import (
    init_products_index, sync_all_products,
    search_products as meili_search_products,
    index_product as meili_index_product,
    delete_product as meili_delete_product,
)
# Models
from models.product import Product
from models.collection import Collection
from models.category import Category  # backward-compat alias
from models.product_image import ProductImage
from models.inventory import Inventory
from models.order import Order, OrderItem, OrderStatus
from models.address import Address, AddressType
from models.review import Review
from models.order_tracking import OrderTracking
from models.return_request import ReturnRequest, ReturnStatus, ReturnReason

# Schemas
from schemas.product import (
    ProductCreate, ProductResponse, ProductUpdate, ProductDetailResponse,
    BulkPriceUpdate, BulkStatusUpdate, BulkCollectionAssign, BulkInventoryUpdate, BulkDeleteProducts
)
from schemas.collection import (
    CollectionCreate, CollectionUpdate, CollectionResponse, CollectionWithProducts,
    BulkCollectionStatusUpdate, BulkCollectionReorder,
)
from schemas.category import (
    CategoryCreate, CategoryUpdate, CategoryResponse, CategoryWithChildren,  # backward-compat
)
from schemas.product_image import ProductImageCreate, ProductImageResponse
from schemas.inventory import InventoryCreate, InventoryUpdate, InventoryResponse, StockAdjustment, LowStockItem
from schemas.order import (
    OrderCreate, OrderResponse, CartItem, CartResponse,
    BulkOrderStatusUpdate, SetDeliveryState,
    GuestOrderTrackResponse, GuestOrderTrackItem,
)
from schemas.address import AddressCreate, AddressUpdate, AddressResponse
from schemas.review import ReviewCreate, ReviewResponse
from schemas.order_tracking import OrderTrackingCreate, OrderTrackingResponse
from schemas.return_request import ReturnRequestCreate, ReturnRequestUpdate, ReturnRequestResponse
from schemas.error import ErrorResponse, PaginatedResponse

# Services
from service.collection_service import CollectionService
from service.category_service import CategoryService  # backward-compat alias
from service.inventory_service import InventoryService
from service.r2_service import r2_service
from service.color_utils import _hex_to_color_name
from service.product_service import ProductService
from service.cart_service import CartService
from service.order_service import OrderService
from service.address_service import AddressService
from service.review_service import ReviewService
from service.order_tracking_service import OrderTrackingService
from service.return_service import ReturnService
from service.guest_tracking_token import parse_guest_tracking_token

# Route modules (for better code organization)
# These modularize the 2800+ line main.py into manageable route files
try:
    from routes import (
        addresses_router,
        chat_router,
        internal_router,
        landing_router,
        orders_router,
        products_router,
        size_guide_router,
    )
    ROUTES_AVAILABLE = True
except ImportError:
    ROUTES_AVAILABLE = False
    # Fallback to monolithic routes in main.py
    pass

# Concurrency control
from core.cart_lock import CartConcurrencyManager, cart_operation_lock


# R2 URL + product/collection enrichment now live in helpers.py.
# Aliased here with the legacy underscore names so existing call sites keep working.
from helpers import (  # noqa: E402
    enrich_collection as _enrich_collection,
    enrich_product as _enrich_product,
    is_hex_color as _is_hex_color,
    r2_url as _r2_url,
)

# Middleware
from shared.auth_middleware import (
    get_current_user,
    get_current_user_optional,
    require_admin,
    require_staff,
    initialize_auth_middleware
)
from shared.roles import is_staff
from shared.request_id_middleware import RequestIDMiddleware

def reconcile_cart_reservations(db: Session) -> int:
    """
    Reconcile inventory.reserved_quantity against active cart reservations in Redis.

    This prevents reservation leaks when cart reservation TTLs expire or Redis
    evicts keys without corresponding DB updates.
    """
    if not redis_client.ping():
        return 0

    try:
        raw_client = redis_client.client
    except Exception:
        return 0

    keys = list(raw_client.scan_iter("cache:cart:reservation:*", count=100))
    if not keys:
        # Clear any stale reservations in DB
        stale_rows = db.query(Inventory).filter(Inventory.reserved_quantity > 0).all()
        if not stale_rows:
            return 0
        for row in stale_rows:
            row.reserved_quantity = 0
        db.commit()
        return len(stale_rows)

    reserved_by_sku = {}
    for key in keys:
        raw = raw_client.get(key)
        if not raw:
            continue
        try:
            data = json.loads(raw)
        except Exception:
            continue
        sku = data.get("sku")
        qty = data.get("quantity")
        if not sku or qty is None:
            continue
        try:
            qty_int = int(qty)
        except Exception:
            continue
        if qty_int <= 0:
            continue
        reserved_by_sku[sku] = reserved_by_sku.get(sku, 0) + qty_int

    skus = list(reserved_by_sku.keys())
    rows = db.query(Inventory).filter(Inventory.reserved_quantity > 0).all()
    if skus:
        rows += db.query(Inventory).filter(Inventory.sku.in_(skus)).all()

    changed = 0
    seen = set()
    for row in rows:
        if row.sku in seen:
            continue
        seen.add(row.sku)
        target = reserved_by_sku.get(row.sku, 0)
        if row.reserved_quantity != target:
            row.reserved_quantity = max(0, target)
            changed += 1

    if changed:
        db.commit()
    return changed


async def _reservation_reconciler(stop_event: asyncio.Event, interval_seconds: int = 600):
    """Background task to reconcile cart reservations on a schedule."""
    while not stop_event.is_set():
        try:
            db = SessionLocal()
            try:
                changed = reconcile_cart_reservations(db)
                if changed:
                    logger.info("Reconciled %s inventory reservations", changed)
            finally:
                db.close()
        except Exception as e:
            logger.warning("Reservation reconciler failed: %s", e)

        try:
            await asyncio.wait_for(stop_event.wait(), timeout=interval_seconds)
        except asyncio.TimeoutError:
            continue



# ==================== Rate Limiting Helpers ====================

import ipaddress

LOCAL_TEST_IPS = {"127.0.0.1", "::1", "localhost", "testclient"}


def _get_client_ip(request: Request) -> str:
    """Resolve the best-effort client IP, honoring proxy headers when present."""
    forwarded_for = request.headers.get("x-forwarded-for", "")
    if forwarded_for:
        first_hop = forwarded_for.split(",")[0].strip()
        if first_hop:
            return first_hop
    return request.client.host if request.client else "unknown"


def _should_bypass_local_rate_limit(request: Request) -> bool:
    """
    Keep rate limiting for deployed traffic while allowing local dev/test
    automation from loopback addresses to exercise flows repeatedly.
    """
    if not settings.is_development:
        return False

    client_ip = _get_client_ip(request)
    if client_ip in LOCAL_TEST_IPS:
        return True

    try:
        parsed_ip = ipaddress.ip_address(client_ip)
    except ValueError:
        return False

    return parsed_ip.is_loopback or parsed_ip.is_private


def _check_rate_limit(request: Request, endpoint: str, limit: int, window: int = 60, user_identifier: str = None) -> bool:
    """
    Check rate limit for a specific endpoint.
    Uses user_identifier when provided (per-customer), falls back to IP only
    for unauthenticated endpoints (search, public pages).
    
    Args:
        request: FastAPI request object
        endpoint: Endpoint identifier (e.g., 'cart_add', 'order_create')
        limit: Maximum requests allowed in window
        window: Time window in seconds (default: 60)
        user_identifier: Customer identifier (email/user_id). If None, uses IP.
    
    Returns:
        True if within limit, False if exceeded
    """
    if _should_bypass_local_rate_limit(request):
        return True
    
    try:
        # Use user identifier when available — blocks abuse per customer,
        # not per IP. Shared IPs (office/mobile/NAT) would block innocent users.
        rate_id = user_identifier if user_identifier else _get_client_ip(request)
        limit_key = f'rate_limit:{endpoint}:{rate_id}'
        count = redis_client.get_cache(limit_key) or 0
        
        if int(count) >= limit:
            return False
        
        redis_client.set_cache(limit_key, int(count) + 1, ttl=window)
        return True
    except Exception as e:
        logger.warning(f'Rate limit check error (skipping): {e}')
        return True  # Allow on error (fail open)


# ==================== Lifespan ====================

# Global instances
event_bus = None

@asynccontextmanager
async def lifespan(app: FastAPI):
    """Application lifespan events."""
    global event_bus
    
    # Initialize DB (create tables)
    init_db()
    
    # Initialize auth middleware
    initialize_auth_middleware(
        secret_key=settings.SECRET_KEY,
        algorithm=settings.ALGORITHM,
        redis_client=redis_client
    )
    
    # Ensure Redis connects
    if redis_client.is_connected():
        logger.info("Redis connected for Commerce service")
    else:
        logger.warning("Redis ping failed")
        
    # Start background reservation reconciler
    stop_event = asyncio.Event()
    task = asyncio.create_task(_reservation_reconciler(stop_event))
    
    # Initialize Event Bus
    from shared.event_bus import EventBus
    event_bus = EventBus(redis_client=redis_client, service_name="commerce_service")
    app.state.event_bus = event_bus  # Store in app.state for route access
    
    # Initialize Meilisearch
    try:
        init_products_index()
        db = SessionLocal()
        try:
            count = sync_all_products(db)
            logger.info(f"✓ Commerce service: Meilisearch synced {count} products")
        finally:
            db.close()
    except Exception as e:
        logger.warning(f"⚠ Commerce service: Meilisearch init skipped ({e})")

    # Sync invoice sequence to prevent collisions after data migrations/restores
    try:
        db = SessionLocal()
        try:
            from service.order_service import sync_invoice_sequence
            sync_invoice_sequence(db)
        finally:
            db.close()
    except Exception as e:
        logger.warning(f"⚠ Could not sync invoice sequence on startup: {e}")

    yield
    
    # Shutdown
    logger.info("Commerce service shutting down")


# ==================== FastAPI App ====================

_env = os.environ.get("ENVIRONMENT", "production")
app = FastAPI(
    title="Aarya Clothing - Commerce Service",
    description="Product Management, Categories, Cart, Orders, Inventory",
    version="2.0.0",
    lifespan=lifespan,
    docs_url="/docs" if _env != "production" else None,
    redoc_url="/redoc" if _env != "production" else None,
)

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=settings.ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "PATCH", "DELETE", "OPTIONS"],
    allow_headers=["Content-Type", "Authorization", "X-Requested-With", "Accept", "Origin", "X-CSRF-Token"],
)

# Prometheus metrics — /metrics endpoint for scraping
try:
    from prometheus_fastapi_instrumentator import Instrumentator
    Instrumentator().instrument(app).expose(app, endpoint="/metrics")
except Exception:
    pass  # Graceful degradation if prometheus lib is missing

# Request ID
app.add_middleware(RequestIDMiddleware)

# Register route modules.
# Cart + orders endpoints remain inline in main.py because they need direct access
# to the CartConcurrencyManager / order service plumbing wired up below.
if ROUTES_AVAILABLE:
    app.include_router(products_router)
    # Collections routes are defined inline (with R2 URL enrichment).
    # orders_router intentionally NOT registered — see inline order routes below.
    app.include_router(addresses_router)
    app.include_router(size_guide_router)
    app.include_router(chat_router)
    app.include_router(landing_router)
    app.include_router(internal_router)
    logger.info("Route modules registered successfully")
else:
    logger.warning("Route modules not available - using monolithic routes in main.py")


# ==================== Health ====================

@app.get("/health", tags=["Health"])
async def health_check(db: Session = Depends(get_db)):
    """Health check endpoint."""
    redis_status = "healthy" if redis_client.ping() else "unhealthy"
    db_status = "healthy"
    try:
        db.execute(text("SELECT 1"))
    except Exception:
        db_status = "unhealthy"
        
    return {
        "status": "healthy" if db_status == "healthy" else "degraded",
        "service": "commerce",
        "version": "2.0.0",
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "dependencies": {
            "redis": redis_status,
            "database": db_status
        }
    }


@app.get("/api/v1/health", tags=["Health"])
async def health_api(db: Session = Depends(get_db)):
    """API health check endpoint."""
    redis_status = "healthy" if redis_client.ping() else "unhealthy"
    db_status = "healthy"
    try:
        db.execute(text("SELECT 1"))
    except Exception:
        db_status = "unhealthy"
    return {
        "status": "healthy" if db_status == "healthy" else "degraded",
        "service": "commerce",
        "version": "2.0.0",
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "dependencies": {
            "redis": redis_status,
            "database": db_status
        }
    }


# ==================== Collections / Categories (same thing) ====================

@app.get("/api/v1/collections", tags=["Collections"])
@app.get("/api/v1/categories", tags=["Collections"])  # backward compat
async def list_collections(
    featured_only: bool = False,
    active_only: bool = True,
    db: Session = Depends(get_db)
):
    """List all collections with full R2 image URLs."""
    cache_key = f"collections:list:{featured_only}:{active_only}"
    
    async def fetch_collections():
        query = db.query(Collection)
        if active_only:
            query = query.filter(Collection.is_active == True)
        if featured_only:
            query = query.filter(Collection.is_featured == True)
        collections = query.order_by(Collection.display_order, Collection.name).all()
        return [_enrich_collection(c) for c in collections]

    return await cache.get_or_set(cache_key, fetch_collections, ttl=300)


@app.get("/api/v1/collections/{collection_id}", tags=["Collections"])
@app.get("/api/v1/categories/{category_id}", tags=["Collections"])  # backward compat
async def get_collection(
    collection_id: Optional[int] = None,
    category_id: Optional[int] = None,
    db: Session = Depends(get_db)
):
    """Get collection by ID with full R2 image URL."""
    cid = collection_id or category_id
    col = db.query(Collection).filter(Collection.id == cid).first()
    if not col:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Collection not found")
    return _enrich_collection(col)


@app.get("/api/v1/collections/slug/{slug}", tags=["Collections"])
@app.get("/api/v1/categories/slug/{slug}", tags=["Collections"])  # backward compat
async def get_collection_by_slug(slug: str, db: Session = Depends(get_db)):
    """Get collection by slug with full R2 image URL."""
    col = db.query(Collection).filter(Collection.slug == slug).first()
    if not col:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Collection not found")
    return _enrich_collection(col)


# ==================== Admin Collection Routes ====================

@app.post("/api/v1/admin/collections", response_model=CategoryResponse,
          status_code=status.HTTP_201_CREATED,
          tags=["Admin - Collections"])
@app.post("/api/v1/admin/categories", response_model=CategoryResponse,
          status_code=status.HTTP_201_CREATED,
          tags=["Admin - Collections"])
async def create_category(
    category: CategoryCreate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_admin)
):
    """Create a new collection/category."""
    # Check if slug already exists
    existing = db.query(Collection).filter(Collection.slug == category.slug).first()
    if existing:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Slug already exists")

    db_collection = Collection(**{k: v for k, v in category.model_dump().items() if hasattr(Collection, k)})
    db.add(db_collection)
    db.commit()
    db.refresh(db_collection)
    cache.invalidate_pattern("collections:*")
    cache.invalidate_pattern("products:*")
    return _enrich_collection(db_collection)


@app.get("/api/v1/admin/collections", response_model=List[CategoryResponse],
         tags=["Admin - Collections"])
@app.get("/api/v1/admin/categories", response_model=List[CategoryResponse],
         tags=["Admin - Collections"])
async def list_admin_categories(
    skip: int = 0,
    limit: int = 100,
    active_only: bool = False,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_admin)
):
    """List all categories for admin."""
    query = db.query(Collection)
    if active_only:
        query = query.filter(Collection.is_active == True)

    collections = query.order_by(Collection.display_order.asc(), Collection.name.asc()).offset(skip).limit(limit).all()
    return [_enrich_collection(col) for col in collections]


@app.patch("/api/v1/admin/collections/{collection_id}", response_model=CategoryResponse,
           tags=["Admin - Collections"])
@app.patch("/api/v1/admin/categories/{category_id}", response_model=CategoryResponse,
           tags=["Admin - Collections"])
async def update_category(
    category_id: Optional[int] = None,
    collection_id: Optional[int] = None,
    category_update: CategoryUpdate = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_admin)
):
    """Update a collection."""
    cid = collection_id or category_id
    db_collection = db.query(Collection).filter(Collection.id == cid).first()
    if not db_collection:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Collection not found")

    update_data = category_update.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        if hasattr(db_collection, field):
            setattr(db_collection, field, value)

    db.commit()
    db.refresh(db_collection)
    cache.invalidate_pattern("collections:*")
    cache.invalidate_pattern("products:*")
    return _enrich_collection(db_collection)


@app.delete("/api/v1/admin/collections/{collection_id}",
            status_code=status.HTTP_204_NO_CONTENT,
            tags=["Admin - Collections"])
@app.delete("/api/v1/admin/categories/{category_id}",
            status_code=status.HTTP_204_NO_CONTENT,
            tags=["Admin - Collections"])
async def delete_category(
    category_id: Optional[int] = None,
    collection_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_admin)
):
    """Delete a collection."""
    cid = collection_id or category_id
    db_collection = db.query(Collection).filter(Collection.id == cid).first()
    if not db_collection:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Collection not found")

    db.delete(db_collection)
    db.commit()
    # Invalidate caches
    cache.invalidate_pattern("collections:*")
    cache.invalidate_pattern("products:*")
    return


@app.post("/api/v1/admin/collections/{collection_id}/image",
          response_model=CategoryResponse,
          tags=["Admin - Collections"])
@app.post("/api/v1/admin/categories/{category_id}/image",
          response_model=CategoryResponse,
          tags=["Admin - Collections"])
async def upload_category_image(
    category_id: Optional[int] = None,
    collection_id: Optional[int] = None,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_admin)
):
    """Upload an image for a collection."""
    from service.r2_service import r2_service
    cid = collection_id or category_id
    # Check if collection exists
    db_collection = db.query(Collection).filter(Collection.id == cid).first()
    if not db_collection:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Collection not found")

    try:
        image_url = await r2_service.upload_image(file, folder="collections")
        db_collection.image_url = image_url
        db.commit()
        db.refresh(db_collection)
        cache.invalidate_pattern("collections:*")
        return _enrich_collection(db_collection)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to upload category image: {str(e)}"
        )


@app.post("/api/v1/admin/collections/bulk/status",
          tags=["Admin - Collections"])
@app.post("/api/v1/admin/categories/bulk/status",
          tags=["Admin - Collections"])
async def bulk_update_collection_status(
    payload: BulkCollectionStatusUpdate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_admin)
):
    """Bulk activate/deactivate collections."""
    updated = db.query(Collection).filter(Collection.id.in_(payload.ids)).update(
        {"is_active": payload.is_active}, synchronize_session=False
    )
    db.commit()
    cache.invalidate_pattern("collections:*")
    return {"updated": updated}


@app.post("/api/v1/admin/collections/bulk/reorder",
          tags=["Admin - Collections"])
@app.post("/api/v1/admin/categories/bulk/reorder",
          tags=["Admin - Collections"])
async def bulk_reorder_collections(
    payload: BulkCollectionReorder,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_admin)
):
    """Bulk update collection display order."""
    for item in payload.items:
        db.query(Collection).filter(Collection.id == item["id"]).update(
            {"display_order": item["display_order"]}, synchronize_session=False
        )
    db.commit()
    cache.invalidate_pattern("collections:*")
    return {"reordered": len(payload.items)}


# ==================== Product Search (supplemental — main product routes in products_router) ==


@app.get("/api/v1/search/suggestions", tags=["Search"])
async def get_search_suggestions(
    request: Request,
    q: str = Query(..., min_length=1, description="Search query for suggestions"),
    limit: int = Query(5, ge=1, le=10, description="Max suggestions per category"),
    db: Session = Depends(get_db)
):
    """
    Get search suggestions for autocomplete.
    Returns products, categories, and trending searches.
    Optimized for fast response times.
    """
    # Rate limiting: 30 requests per minute per IP
    if not _check_rate_limit(request, "search_suggestions", limit=30, window=60):
        raise HTTPException(
            status_code=status.HTTP_429_TOO_MANY_REQUESTS,
            detail="Too many suggestion requests. Please try again later."
        )

    from search.meilisearch_client import get_search_suggestions as meili_suggestions
    
    result = meili_suggestions(query=q, limit=limit, db_session=db)
    return result










# ==================== Cart Routes (Token Based) ====================

@app.get("/api/v1/cart", response_model=CartResponse, tags=["Cart"])
async def get_my_cart(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Get current user's cart (user_id from token)."""
    user_id = current_user["user_id"]  # Fixed: use user_id from auth middleware
    cart_service = CartService(db)
    cart_data = cart_service.get_cart(user_id)
    return CartResponse(**cart_data)


@app.post("/api/v1/cart/items", response_model=CartResponse, tags=["Cart"])
async def add_to_my_cart(
    item: CartItem,
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Add item to current user's cart (user_id from token)."""
    # Rate limiting: 100 requests per minute per IP (generous for shared IPs)
    if not _check_rate_limit(request, "cart_add", limit=100, window=60):
        raise HTTPException(
            status_code=status.HTTP_429_TOO_MANY_REQUESTS,
            detail="Too many cart operations. Please try again later."
        )
    
    user_id = current_user["user_id"]  # Fixed: use user_id from auth middleware
    
    # Validate product exists
    product = db.query(Product).filter(Product.id == item.product_id).first()
    if not product:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Product not found"
        )
    
    # Check variant-specific stock when variant_id is given; otherwise check total product stock
    if item.variant_id:
        variant = db.query(Inventory).filter(
            Inventory.id == item.variant_id,
            Inventory.product_id == item.product_id
        ).first()
        if not variant:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Variant not found"
            )
        if variant.available_quantity < item.quantity:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Only {variant.available_quantity} items available"
            )
    else:
        if product.total_stock < item.quantity:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Insufficient inventory"
            )
    
    item_data = {
        "product_id": item.product_id,
        "variant_id": item.variant_id,
        "quantity": item.quantity
    }
    
    cart_data = CartConcurrencyManager.add_to_cart_locked(user_id, item_data, db)
    return CartResponse(**cart_data)


class CartItemUpdate(BaseModel):
    """Body for PUT /cart/items/{product_id}."""
    quantity: int
    variant_id: Optional[int] = None


@app.put("/api/v1/cart/items/{product_id}", response_model=CartResponse, tags=["Cart"])
async def update_my_cart_item(
    product_id: int,
    body: CartItemUpdate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Update item quantity in cart. Accepts JSON body: {quantity, variant_id?}."""
    if body.quantity < 1:
        raise HTTPException(status_code=400, detail="quantity must be >= 1")
    user_id = current_user["user_id"]
    cart_data = CartConcurrencyManager.update_cart_item_locked(user_id, product_id, body.quantity, db, body.variant_id)
    return CartResponse(**cart_data)


@app.delete("/api/v1/cart/items/{product_id}", response_model=CartResponse, tags=["Cart"])
async def remove_from_my_cart(
    product_id: int,
    variant_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Remove item from cart."""
    user_id = current_user["user_id"]  # Fixed: use user_id from auth middleware
    cart_data = CartConcurrencyManager.remove_from_cart_locked(user_id, product_id, db, variant_id)
    return CartResponse(**cart_data)


@app.delete("/api/v1/cart", response_model=CartResponse, tags=["Cart"])
async def clear_my_cart(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Clear current user's cart."""
    user_id = current_user["user_id"]  # Fixed: use user_id from auth middleware
    cart_data = CartConcurrencyManager.clear_cart_locked(user_id, db)
    return CartResponse(**cart_data)


@app.post("/api/v1/cart/delivery-state", response_model=CartResponse, tags=["Cart"])
async def set_cart_delivery_state(
    payload: SetDeliveryState,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Set delivery state on cart to trigger correct GST (CGST+SGST vs IGST) calculation."""
    user_id = current_user["user_id"]
    cart_service = CartService(db)
    cart = cart_service.get_cart(user_id)
    cart["delivery_state"] = payload.delivery_state
    if payload.customer_gstin:
        cart["customer_gstin"] = payload.customer_gstin
    cart_service._recalculate_cart(cart)
    cart_service.save_cart(user_id, cart)
    return CartResponse(**cart)


@app.post("/api/v1/cart/clear-expired", tags=["Cart"])
async def clear_expired_cart_items(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Clear expired cart reservations and clean up stale cart data."""
    user_id = current_user["user_id"]
    cart_service = CartService(db)
    cart = cart_service.get_cart(user_id)
    # Refresh reservation expiry
    cart["reservation_expires_at"] = cart_service._get_earliest_reservation_expiry(user_id)
    cart_service.save_cart(user_id, cart)
    return {"status": "ok", "cart": cart}


# ==================== Cart Stock Validation SSE ====================

@app.get("/api/v1/cart/stock-stream", tags=["Cart"])
async def cart_stock_stream(
    request: Request,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Server-Sent Events endpoint for real-time cart stock updates.
    Streams stock availability changes for items in user's cart.
    Automatically disconnects when client closes connection or after max runtime.
    """
    user_id = current_user["user_id"]
    cart_service = CartService(db)
    MAX_RUNTIME_SECONDS = 3600  # 1 hour max connection lifetime
    start_time = asyncio.get_event_loop().time()
    
    async def event_generator():
        """Generate SSE events for stock updates."""
        while True:
            # Check if client disconnected or max runtime exceeded
            if await request.is_disconnected():
                logger.debug(f"Client disconnected from stock stream for user {user_id}")
                break
            
            elapsed = asyncio.get_event_loop().time() - start_time
            if elapsed > MAX_RUNTIME_SECONDS:
                logger.debug(f"Max runtime exceeded for stock stream user {user_id}")
                yield f"data: {json.dumps({'type': 'timeout', 'message': 'Connection timeout'})}\n\n"
                break
            
            try:
                # Get current cart
                cart = cart_service.get_cart(user_id)
                
                if cart["items"]:
                    # Check stock for each item
                    stock_updates = []
                    for item in cart["items"]:
                        # Query current stock from database
                        result = db.execute(
                            text("SELECT quantity, reserved_quantity FROM inventory WHERE sku = :sku"),
                            {"sku": item.get("sku")}
                        ).fetchone()
                        
                        if result:
                            available = max(0, result[0] - result[1])
                            stock_updates.append({
                                "product_id": item["product_id"],
                                "variant_id": item.get("variant_id"),
                                "sku": item.get("sku"),
                                "available_quantity": available,
                                "requested_quantity": item["quantity"],
                                "in_stock": available >= item["quantity"]
                            })
                    
                    # Send stock status event
                    event_data = {
                        "type": "stock_update",
                        "timestamp": datetime.now(timezone.utc).isoformat(),
                        "items": stock_updates
                    }
                    yield f"data: {json.dumps(event_data)}\n\n"
                
                # Send keepalive every 30 seconds
                await asyncio.sleep(30)
                yield ":keepalive\n\n"
                
            except ConnectionResetError:
                logger.debug(f"Connection reset for stock stream user {user_id}")
                break
            except TimeoutError:
                logger.warning(f"Database timeout in stock stream for user {user_id}")
                error_data = {"type": "error", "message": "Database timeout"}
                yield f"data: {json.dumps(error_data)}\n\n"
                await asyncio.sleep(60)
            except Exception as e:
                logger.error(f"Unexpected error in stock stream for user {user_id}: {e}", exc_info=True)
                error_data = {"type": "error", "message": "Internal server error"}
                yield f"data: {json.dumps(error_data)}\n\n"
                await asyncio.sleep(60)  # Wait longer on error
    
    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no"  # Disable nginx buffering
        }
    )


# ==================== Cart Routes (Legacy/Admin) ====================

@app.get("/api/v1/cart/{user_id}", response_model=CartResponse,
         tags=["Cart"])
async def get_cart(
    user_id: int,
    current_user: dict = Depends(get_current_user)
):
    """Get user's shopping cart."""
    # Authorization check
    if current_user["user_id"] != user_id and current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not authorized to view this cart"
        )
    
    cart_key = f"cart:{user_id}"
    cart_data = redis_client.get_cache(cart_key)
    
    if not cart_data:
        return CartResponse(user_id=user_id, items=[], total=0)
    
    return CartResponse(**cart_data)


# ==================== Checkout Routes ====================

@app.post("/api/v1/checkout/validate", tags=["Checkout"])
async def validate_checkout(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """
    Validate cart stock before payment. Returns structured payload for the storefront:
    { valid, out_of_stock: [{ sku, name, requested, available }], message }
    """
    cart_service = CartService(db)
    preview = cart_service.checkout_stock_preview(current_user["user_id"])
    if preview.get("valid"):
        return preview
    # Still 200 so the client can read out_of_stock without treating it as transport error
    return preview



@app.post("/api/v1/cart/{user_id}/add",
          response_model=CartResponse,
          tags=["Cart"])
async def add_to_cart(
    user_id: int,
    item: CartItem,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    # Authorization check
    if current_user["user_id"] != user_id and current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not authorized to modify this cart"
        )
    
    # Validate product exists and has sufficient stock
    product = db.query(Product).filter(Product.id == item.product_id).first()
    if not product:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Product not found"
        )
    
    # Check inventory (simplified - check total stock)
    if product.total_stock < item.quantity:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Insufficient inventory"
        )
    
    # Use distributed locking for cart operations
    item_data = {
        "product_id": item.product_id,
        "variant_id": item.variant_id,
        "quantity": item.quantity
    }
    
    cart_data = CartConcurrencyManager.add_to_cart_locked(user_id, item_data, db)
    
    return CartResponse(**cart_data)


@app.delete("/api/v1/cart/{user_id}/remove/{product_id}",
            response_model=CartResponse,
            tags=["Cart"])
async def remove_from_cart(
    user_id: int,
    product_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Remove item from cart (legacy endpoint — delegates to CartService for proper reservation release)."""
    if current_user["user_id"] != user_id and current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not authorized to modify this cart"
        )
    cart_data = CartConcurrencyManager.remove_from_cart_locked(user_id, product_id, db)
    return CartResponse(**cart_data)


@app.delete("/api/v1/cart/{user_id}/clear", response_model=CartResponse,
           tags=["Cart"])
async def clear_cart(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Clear user's shopping cart (legacy endpoint — delegates to CartService for proper reservation release)."""
    if current_user["user_id"] != user_id and current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not authorized to modify this cart"
        )
    cart_data = CartConcurrencyManager.clear_cart_locked(user_id, db)
    return CartResponse(**cart_data)


# ==================== Order Routes ====================

@app.post("/api/v1/orders", response_model=OrderResponse,
          status_code=status.HTTP_201_CREATED,
          tags=["Orders"])
async def create_order(
    order_data: OrderCreate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Create a new order from cart."""
    # Rate limiting: 10 orders per minute per IP
    if not _check_rate_limit(request, "order_create", limit=10, window=60, user_identifier=str(current_user["user_id"])):
        raise HTTPException(
            status_code=status.HTTP_429_TOO_MANY_REQUESTS,
            detail="Too many order creation attempts. Please try again later."
        )
    
    order_service = OrderService(db)
    # Create order via service
    try:
        from shared.event_bus import Event, EventType

        order = order_service.create_order(
            user_id=current_user["user_id"],
            shipping_address=order_data.shipping_address,
            address_id=order_data.address_id,
            order_notes=order_data.notes or order_data.order_notes,
            transaction_id=order_data.transaction_id or order_data.payment_id,
            payment_method=order_data.payment_method,
            razorpay_order_id=order_data.razorpay_order_id,
            payment_signature=order_data.razorpay_signature,
            qr_code_id=order_data.qr_code_id,
        )
        
        # Publish event
        if event_bus and order:
            try:
                event_data = {
                    "order_id": order.id,
                    "order_number": order.invoice_number,
                    "user_id": current_user["user_id"], # Changed from current_user["id"] to current_user["user_id"]
                    "total_amount": float(order.total_amount),
                    "shipping_address": order_data.shipping_address, # Changed from shipping_address to order_data.shipping_address
                    "status": order.status.value
                }
                order_event = Event(
                    event_type=EventType.ORDER_CREATED,
                    aggregate_id=str(order.id),
                    aggregate_type="order",
                    data=event_data,
                    metadata={"source": "commerce_service"}
                )
                await event_bus.publish(order_event)
            except Exception as exc:
                logger.warning(
                    "Failed to publish order created event for order %s: %s",
                    order.id,
                    exc
                )
        
        return order # Added return statement for the created order
            
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )


@app.get("/api/v1/orders", response_model=List[OrderResponse],
         tags=["Orders"])
async def list_orders(
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """List current user's orders."""
    order_service = OrderService(db)
    return order_service.get_user_orders(
        user_id=current_user["user_id"],
        skip=skip,
        limit=limit
    )


@app.get(
    "/api/v1/orders/track/{token}",
    response_model=GuestOrderTrackResponse,
    tags=["Orders"],
)
async def get_guest_order_by_tracking_token(token: str, db: Session = Depends(get_db)):
    """
    Public order status for guests — token is HMAC-signed (no login).
    Must stay registered before /orders/{order_id} so "track" is not parsed as an integer id.
    """
    order_id = parse_guest_tracking_token(token)
    if order_id is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Invalid or expired tracking link",
        )
    order = (
        db.query(Order)
        .options(joinedload(Order.items))
        .filter(Order.id == order_id)
        .first()
    )
    if not order:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Order not found",
        )
    items_out = [
        GuestOrderTrackItem(
            product_name=it.product_name,
            size=it.size,
            color=it.color,
            quantity=int(it.quantity),
            price=float(it.price) if it.price is not None else float(it.unit_price or 0),
        )
        for it in (order.items or [])
    ]
    return GuestOrderTrackResponse(
        order_id=order.id,
        status=order.status.value if hasattr(order.status, "value") else str(order.status),
        tracking_number=order.tracking_number,
        total_amount=float(order.total_amount),
        created_at=order.created_at,
        items=items_out,
    )


@app.get("/api/v1/orders/{order_id}", response_model=OrderResponse,
         tags=["Orders"])
async def get_order(
    order_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Get order details."""
    order_service = OrderService(db)
    order = order_service.get_order_by_id(order_id, user_id=current_user["user_id"])
    
    if not order:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Order not found"
        )
    
    return order


@app.post("/api/v1/orders/{order_id}/cancel", response_model=OrderResponse,
          tags=["Orders"])
async def cancel_order(
    order_id: int,
    reason: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Cancel an order."""
    order_service = OrderService(db)
    return order_service.cancel_order(
        order_id=order_id,
        user_id=current_user["user_id"],
        reason=reason
    )


@app.get("/api/v1/orders/{order_id}/tracking", response_model=List[OrderTrackingResponse],
         tags=["Orders"])
async def get_order_tracking(
    order_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Get order tracking history."""
    # Verify order belongs to user
    order_service = OrderService(db)
    order = order_service.get_order_by_id(order_id, user_id=current_user["user_id"])
    
    if not order:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Order not found"
        )
    
    tracking_service = OrderTrackingService(db)
    return tracking_service.get_order_tracking(order_id)


# ==================== Admin Payment Recovery ====================

@app.get("/api/v1/orders/admin/payment-recovery", tags=["Admin - Orders"])
async def get_payment_recovery_report(
    from_timestamp: Optional[int] = Query(None, description="Unix timestamp to fetch payments from"),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_admin)
):
    """
    Fetch all captured Razorpay payments and cross-reference with orders in DB.
    Returns matched payments and missing_orders (paid but no order created).
    Admin use only.
    """
    import razorpay as _razorpay
    from datetime import timedelta

    key_id = os.getenv("RAZORPAY_KEY_ID", "")
    key_secret = os.getenv("RAZORPAY_KEY_SECRET", "")
    if not key_id or not key_secret:
        raise HTTPException(status_code=503, detail="Razorpay credentials not configured")

    try:
        client = _razorpay.Client(auth=(key_id, key_secret))
        if not from_timestamp:
            from_timestamp = int((datetime.now(timezone.utc) - timedelta(hours=48)).timestamp())
        payments = client.payment.all({"from": from_timestamp, "count": 100})
        items = payments.get("items", [])
    except Exception as e:
        logger.error(f"Razorpay payment fetch failed: {e}")
        raise HTTPException(status_code=502, detail=f"Failed to fetch Razorpay payments: {str(e)}")

    existing_transactions = set(
        row[0] for row in db.execute(
            text("SELECT transaction_id FROM orders WHERE transaction_id IS NOT NULL")
        ).fetchall()
    )

    matched = []
    missing_orders = []
    total_missing_amount = 0.0

    for payment in items:
        payment_id = payment.get("id", "")
        order_id = payment.get("order_id", "")
        amount_inr = payment.get("amount", 0) / 100
        pay_status = payment.get("status", "")
        if pay_status != "captured":
            continue
        if payment_id in existing_transactions or order_id in existing_transactions:
            matched.append({"payment_id": payment_id, "order_id": order_id, "amount": amount_inr,
                            "email": payment.get("email", ""), "contact": payment.get("contact", ""),
                            "created_at": payment.get("created_at", 0), "status": "matched"})
        else:
            total_missing_amount += amount_inr
            missing_orders.append({"payment_id": payment_id, "razorpay_order_id": order_id,
                                   "amount": amount_inr, "email": payment.get("email", ""),
                                   "contact": payment.get("contact", ""), "method": payment.get("method", ""),
                                   "created_at": payment.get("created_at", 0), "status": "missing_order"})

    return {
        "from_timestamp": from_timestamp,
        "total_payments_fetched": len(items),
        "total_captured": len(matched) + len(missing_orders),
        "matched_count": len(matched),
        "missing_order_count": len(missing_orders),
        "total_missing_amount_inr": total_missing_amount,
        "matched": matched,
        "missing_orders": missing_orders,
    }


@app.post("/api/v1/orders/admin/force-create", tags=["Admin - Orders"])
async def force_create_order_from_payment(
    payload: dict,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_admin)
):
    """
    Force-create a DB order from a captured Razorpay payment.
    Fetches payment details from Razorpay, finds the matching user by email,
    pulls their default shipping address, and inserts a confirmed order.
    Admin use only — for payment recovery when customer's order was not created.
    """
    import razorpay as _razorpay

    payment_id = payload.get("payment_id", "").strip()
    if not payment_id:
        raise HTTPException(status_code=400, detail="payment_id is required")

    # Check if order already exists for this payment
    existing = db.execute(
        text("SELECT id, invoice_number FROM orders WHERE transaction_id = :pid OR razorpay_payment_id = :pid"),
        {"pid": payment_id}
    ).fetchone()
    if existing:
        raise HTTPException(status_code=409, detail=f"Order already exists for this payment: #{existing[0]} ({existing[1]})")

    key_id = os.getenv("RAZORPAY_KEY_ID", "")
    key_secret = os.getenv("RAZORPAY_KEY_SECRET", "")
    if not key_id or not key_secret:
        raise HTTPException(status_code=503, detail="Razorpay credentials not configured")

    # Fetch payment from Razorpay
    try:
        client = _razorpay.Client(auth=(key_id, key_secret))
        payment = client.payment.fetch(payment_id)
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Failed to fetch payment from Razorpay: {str(e)}")

    if payment.get("status") != "captured":
        raise HTTPException(status_code=400, detail=f"Payment {payment_id} is not captured (status: {payment.get('status')})")

    amount_inr = payment.get("amount", 0) / 100
    razorpay_order_id = payment.get("order_id") or ""
    customer_email = payment.get("email") or ""
    customer_contact = payment.get("contact") or ""
    method = payment.get("method") or "razorpay"

    # Find user by email
    user_row = None
    if customer_email:
        user_row = db.execute(
            text("SELECT id FROM users WHERE email = :email"),
            {"email": customer_email}
        ).fetchone()

    user_id = user_row[0] if user_row else 1  # fallback to admin user

    # Get user's default (or first) shipping address
    addr_row = db.execute(
        text("""
            SELECT id, full_name, phone, address_line1, address_line2, city, state, postal_code
            FROM addresses
            WHERE user_id = :uid
            ORDER BY is_default DESC, id ASC
            LIMIT 1
        """),
        {"uid": user_id}
    ).fetchone()

    shipping_address_id = None
    shipping_address_text = ""
    if addr_row:
        shipping_address_id = addr_row[0]
        parts = [addr_row[1], addr_row[3]]
        if addr_row[4]: parts.append(addr_row[4])
        parts += [addr_row[5], f"{addr_row[6]} {addr_row[7]}"]
        if addr_row[2]: parts.append(f"Ph: {addr_row[2]}")
        shipping_address_text = ", ".join(p for p in parts if p)
    elif customer_contact:
        shipping_address_text = f"Contact: {customer_contact} — address to be collected"

    # Generate next invoice number
    max_inv = db.execute(text("SELECT MAX(id) FROM orders")).scalar() or 0
    next_id = max_inv + 1
    invoice_number = f"INV-2026-{str(next_id).zfill(6)}"

    notes = (
        f"RECOVERED by admin ({current_user.get('email','admin')}): "
        f"Force-created from Razorpay {payment_id} (Rs{amount_inr}). "
        f"Method: {method}. Contact: {customer_contact or 'N/A'}."
    )

    result = db.execute(
        text("""
            INSERT INTO orders (
                user_id, transaction_id, razorpay_order_id, razorpay_payment_id,
                status, total_amount, subtotal, payment_method,
                invoice_number, shipping_address_id, shipping_address, order_notes,
                created_at, updated_at
            ) VALUES (
                :user_id, :transaction_id, :razorpay_order_id, :payment_id,
                'confirmed', :amount, :amount, 'razorpay',
                :invoice_number, :addr_id, :shipping_address, :notes,
                NOW(), NOW()
            ) RETURNING id, invoice_number
        """),
        {
            "user_id": user_id,
            "transaction_id": payment_id,
            "razorpay_order_id": razorpay_order_id,
            "payment_id": payment_id,
            "amount": amount_inr,
            "invoice_number": invoice_number,
            "addr_id": shipping_address_id,
            "shipping_address": shipping_address_text,
            "notes": notes,
        }
    )
    db.commit()
    row = result.fetchone()

    logger.info(f"Admin force-created order #{row[0]} for payment {payment_id} by {current_user.get('email')}")

    # Send order confirmation email if customer email available
    email_sent = False
    if customer_email:
        try:
            import smtplib, ssl
            from email.mime.text import MIMEText
            from email.mime.multipart import MIMEMultipart

            smtp_host = os.getenv("SMTP_HOST", "")
            smtp_port = int(os.getenv("SMTP_PORT", "465"))
            smtp_user = os.getenv("SMTP_USER", "")
            smtp_pass = os.getenv("SMTP_PASSWORD", "")
            from_email = os.getenv("EMAIL_FROM", smtp_user)
            from_name = os.getenv("EMAIL_FROM_NAME", "Aarya Clothing")

            if smtp_host and smtp_user and smtp_pass:
                customer_name = addr_row[1] if addr_row else customer_email.split("@")[0].title()
                subject = f"Order Confirmed! {invoice_number} - Aarya Clothing"
                html_body = f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8"></head>
<body style="font-family:sans-serif;background:#0B0608;margin:0;padding:20px;">
<div style="max-width:600px;margin:0 auto;background:linear-gradient(135deg,#180F14,#0B0608);border-radius:20px;padding:40px;border:1px solid rgba(183,110,121,0.3);">
  <div style="text-align:center;margin-bottom:24px;">
    <h1 style="color:#F2C29A;font-size:28px;margin:0;letter-spacing:3px;">AARYA</h1>
    <p style="color:#B76E79;font-size:13px;margin:6px 0 0;letter-spacing:4px;">CLOTHING</p>
  </div>
  <h2 style="color:#F2C29A;font-size:22px;margin-bottom:8px;">Order Confirmed! ✓</h2>
  <p style="color:#EAE0D5;margin-bottom:20px;">Dear {customer_name}, thank you for your purchase! Your order has been confirmed.</p>
  <div style="background:rgba(122,47,87,0.15);border:1px solid rgba(183,110,121,0.2);border-radius:12px;padding:20px;margin-bottom:20px;">
    <p style="color:#EAE0D5;margin:0 0 8px;font-size:13px;opacity:0.6;">ORDER NUMBER</p>
    <p style="color:#F2C29A;font-size:18px;font-weight:bold;margin:0;font-family:monospace;">{invoice_number}</p>
    <hr style="border:none;border-top:1px solid rgba(183,110,121,0.2);margin:16px 0;">
    <p style="color:#EAE0D5;margin:0 0 6px;font-size:13px;opacity:0.6;">AMOUNT PAID</p>
    <p style="color:#4ade80;font-size:20px;font-weight:bold;margin:0;">₹{amount_inr:.0f}</p>
    <hr style="border:none;border-top:1px solid rgba(183,110,121,0.2);margin:16px 0;">
    <p style="color:#EAE0D5;margin:0 0 6px;font-size:13px;opacity:0.6;">PAYMENT METHOD</p>
    <p style="color:#EAE0D5;margin:0;">{method.upper()} via Razorpay</p>
    {'<hr style="border:none;border-top:1px solid rgba(183,110,121,0.2);margin:16px 0;"><p style="color:#EAE0D5;margin:0 0 6px;font-size:13px;opacity:0.6;">DELIVERY ADDRESS</p><p style="color:#EAE0D5;margin:0;">' + shipping_address_text + '</p>' if shipping_address_text else ''}
  </div>
  <div style="text-align:center;margin:24px 0;">
    <a href="https://aaryaclothing.in/profile/orders" style="display:inline-block;background:linear-gradient(135deg,#B76E79,#7A2F57);color:#fff;text-decoration:none;padding:14px 32px;border-radius:10px;font-weight:600;font-size:15px;">View My Orders</a>
  </div>
  <p style="color:#EAE0D5;opacity:0.5;font-size:12px;text-align:center;margin-top:24px;">Questions? Email us at <a href="mailto:support@aaryaclothing.in" style="color:#B76E79;">support@aaryaclothing.in</a></p>
  <p style="color:#EAE0D5;opacity:0.3;font-size:11px;text-align:center;">© 2026 Aarya Clothing. All rights reserved.</p>
</div>
</body></html>"""
                msg = MIMEMultipart("alternative")
                msg["Subject"] = subject
                msg["From"] = f"{from_name} <{from_email}>"
                msg["To"] = customer_email
                msg.attach(MIMEText(html_body, "html"))
                ctx = ssl.create_default_context()
                with smtplib.SMTP_SSL(smtp_host, smtp_port, context=ctx, timeout=20) as srv:
                    srv.login(smtp_user, smtp_pass)
                    srv.sendmail(from_email, customer_email, msg.as_string())
                email_sent = True
                logger.info(f"Order confirmation email sent to {customer_email} for {invoice_number}")
        except Exception as email_err:
            logger.warning(f"Failed to send confirmation email to {customer_email}: {email_err}")

    return {
        "success": True,
        "order_id": row[0],
        "invoice_number": row[1],
        "payment_id": payment_id,
        "amount": amount_inr,
        "customer_email": customer_email,
        "shipping_address": shipping_address_text,
        "email_sent": email_sent,
        "message": f"Order #{row[0]} created successfully" + (" — confirmation email sent" if email_sent else ""),
    }


# ==================== Admin Order Routes ====================

@app.get("/api/v1/admin/orders",
         tags=["Admin - Orders"])
async def list_all_orders(
    status_filter: Optional[str] = None,
    status: Optional[str] = None,          # alias accepted by frontend
    search: Optional[str] = None,          # search by order id or customer
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_staff)
):
    """List all orders (admin/staff). Returns {orders, total}."""
    from models.order import Order
    from sqlalchemy import or_, cast, String
    effective_status = status_filter or status
    order_service = OrderService(db)
    order_status = OrderStatus(effective_status) if effective_status else None
    orders = order_service.get_all_orders(status=order_status, skip=skip, limit=limit)
    # Count query for pagination
    count_q = db.query(Order)
    if order_status:
        count_q = count_q.filter(Order.status == order_status)
    if search:
        count_q = count_q.filter(
            or_(
                cast(Order.id, String).ilike(f"%{search}%"),
                Order.customer_name.ilike(f"%{search}%"),
                Order.customer_email.ilike(f"%{search}%"),
            )
        )
        # Also filter the already-fetched list by search
        sl = search.lower()
        orders = [
            o for o in orders
            if sl in str(o.id) or
               (o.customer_name and sl in o.customer_name.lower()) or
               (o.customer_email and sl in o.customer_email.lower())
        ]
    total = count_q.count()
    return {"orders": orders, "total": total}


class OrderStatusUpdate(BaseModel):
    """Body for PATCH /admin/orders/{id}/status."""
    status: str
    pod_number: Optional[str] = None
    tracking_number: Optional[str] = None
    notes: Optional[str] = None


@app.patch("/api/v1/admin/orders/{order_id}/status", response_model=OrderResponse,
           tags=["Admin - Orders"])
async def update_order_status(
    order_id: int,
    body: OrderStatusUpdate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_staff)
):
    """Update order status (admin/staff). Accepts JSON body: {status, pod_number, tracking_number, notes}."""
    new_status = body.status
    tracking_number = body.pod_number or body.tracking_number
    notes = body.notes
    order_service = OrderService(db)
    order = order_service.update_order_status(
        order_id=order_id,
        new_status=OrderStatus(new_status),
        tracking_number=tracking_number,
        admin_notes=notes
    )

    tracking_service = OrderTrackingService(db)
    tracking_service.add_tracking_entry(
        order_id=order_id,
        status=OrderStatus(new_status),
        notes=notes,
        updated_by=current_user["user_id"]
    )

    redis_client.publish("order:status", {
        "order_id": order_id,
        "status": new_status,
        "tracking_number": tracking_number,
        "notes": notes,
        "updated_by": current_user["user_id"],
        "timestamp": datetime.now(timezone.utc).isoformat(),
    })

    return order


@app.get("/api/v1/admin/orders/{order_id}/tracking", tags=["Admin - Orders"])
async def get_order_tracking(
    order_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_staff)
):
    """Get order tracking/history (admin/staff)."""
    from models.order_tracking import OrderTracking
    
    tracking = db.query(OrderTracking).filter(
        OrderTracking.order_id == order_id
    ).order_by(OrderTracking.created_at.desc()).all()
    
    return {
        "tracking": [
            {
                "id": t.id,
                "status": t.status.value,
                "notes": t.notes,
                "location": t.location,
                "created_at": t.created_at.isoformat() if t.created_at else None,
            }
            for t in tracking
        ]
    }


# ==================== SSE Order Events ====================

@app.get("/api/v1/orders/{order_id}/events", tags=["Orders"])
async def order_status_events(
    order_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """
    Server-Sent Events endpoint for real-time order status updates.
    
    Subscribes to Redis Pub/Sub channel `order_updates:{order_id}`.
    When staff updates the order status, the event is pushed instantly.
    
    Usage: const es = new EventSource('/api/v1/orders/{order_id}/events')
    
    Security: Verifies that the user owns the order or is a staff member.
    """
    from starlette.responses import StreamingResponse
    from core.redis_client import redis_client
    from models.order import Order
    
    # Verify order exists and user has access
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Order not found"
        )
    
    # Check if user owns this order OR is a staff/admin
    is_staff = current_user.get("is_staff", False) or current_user.get("is_admin", False)
    if order.user_id != current_user.get("user_id") and not is_staff:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not authorized to view this order's events"
        )
    
    # Check Redis connectivity before creating SSE stream
    if not redis_client.is_connected():
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="Real-time updates unavailable. Please try again later."
        )

    async def event_generator():
        """Subscribe to Redis Pub/Sub and yield SSE events."""
        # Create a dedicated pubsub subscriber for this connection
        pubsub = redis_client.client.pubsub()
        channel = f"order_updates:{order_id}"
        pubsub.subscribe(channel)

        try:
            # Send initial connection event
            yield f"event: connected\ndata: {{\"order_id\": {order_id}}}\n\n"

            heartbeat_counter = 0
            while True:
                # Non-blocking check for new messages
                message = pubsub.get_message(ignore_subscribe_messages=True, timeout=1.0)

                if message and message["type"] == "message":
                    data = message["data"]
                    if isinstance(data, bytes):
                        data = data.decode("utf-8")
                    yield f"event: status_update\ndata: {data}\n\n"
                
                heartbeat_counter += 1
                # Send heartbeat every ~30 seconds (30 iterations × 1s timeout)
                if heartbeat_counter >= 30:
                    yield ": heartbeat\n\n"
                    heartbeat_counter = 0

                await asyncio.sleep(0)  # Yield control to event loop

        except asyncio.CancelledError:
            pass
        except Exception as e:
            logger.error(f"SSE error for order {order_id}: {e}")
        finally:
            # Ensure proper cleanup of pubsub connection
            if pubsub:
                try:
                    pubsub.unsubscribe(channel)
                    pubsub.close()
                except Exception as cleanup_err:
                    logger.warning(f"SSE pubsub cleanup error for order {order_id}: {cleanup_err}")

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        }
    )


# ==================== Admin Bulk Order Status ====================

class BulkOrderStatusUpdateBody(BaseModel):
    """Body for bulk status update — extends BulkOrderStatusUpdate with pod_number."""
    order_ids: List[int]
    status: str
    pod_number: Optional[str] = None
    tracking_number: Optional[str] = None
    notes: Optional[str] = None
    dry_run: bool = False


@app.patch("/api/v1/admin/orders/bulk-status", tags=["Admin - Orders"])
async def bulk_update_order_status(
    data: BulkOrderStatusUpdateBody,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_staff)
):
    """
    Bulk update order status for multiple orders.
    Accepts: {order_ids, status, pod_number?, notes?, dry_run?}
    """
    MAX_BULK_SIZE = 100
    if len(data.order_ids) > MAX_BULK_SIZE:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Cannot process more than {MAX_BULK_SIZE} orders at once"
        )

    if data.dry_run:
        from models.order import Order as OrderModel
        count = db.query(OrderModel).filter(OrderModel.id.in_(data.order_ids)).count()
        return {"updated": count, "order_ids": data.order_ids, "new_status": data.status, "dry_run": True}

    tracking_number = data.pod_number or data.tracking_number
    order_service = OrderService(db)
    # Update status
    updated_count = order_service.bulk_update_order_status(
        order_ids=data.order_ids,
        new_status=OrderStatus(data.status)
    )
    # If tracking number provided (e.g. ship), write it to each order
    if tracking_number:
        from models.order import Order as OrderModel
        db.query(OrderModel).filter(OrderModel.id.in_(data.order_ids)).update(
            {"tracking_number": tracking_number},
            synchronize_session=False
        )
        db.commit()
    return {"updated": updated_count, "order_ids": data.order_ids, "new_status": data.status, "dry_run": False}


@app.get("/api/v1/admin/orders/pod-template", tags=["Admin - Orders"])
async def download_pod_template(
    current_user: dict = Depends(require_staff)
):
    """Download a blank Excel template for bulk POD number upload."""
    import io
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "POD Upload"
        ws.append(["order_id", "pod_number"])
        ws.append(["123", "TRACK123456"])  # example row
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        from fastapi.responses import StreamingResponse as _SR
        return _SR(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=pod_template.xlsx"}
        )
    except ImportError:
        import csv, io as _io
        output = _io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["order_id", "pod_number"])
        writer.writerow(["123", "TRACK123456"])
        from fastapi.responses import Response as _Resp
        return _Resp(
            content=output.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=pod_template.csv"}
        )


@app.post("/api/v1/admin/orders/upload-pod-excel", tags=["Admin - Orders"])
async def upload_pod_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_staff)
):
    """Upload an Excel/CSV file with order_id + pod_number columns to bulk-set tracking numbers."""
    import io
    content = await file.read()
    rows_updated = 0
    errors = []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        oid_idx = headers.index("order_id") if "order_id" in headers else 0
        pod_idx = headers.index("pod_number") if "pod_number" in headers else 1
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                oid = int(row[oid_idx])
                pod = str(row[pod_idx]).strip() if row[pod_idx] else ""
                if not pod:
                    continue
                from models.order import Order as OrderModel
                updated = db.query(OrderModel).filter(OrderModel.id == oid).update(
                    {"tracking_number": pod}, synchronize_session=False
                )
                if updated:
                    rows_updated += 1
                else:
                    errors.append(f"Order {oid} not found")
            except Exception as row_err:
                errors.append(str(row_err))
        db.commit()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {e}")
    return {"updated": rows_updated, "errors": errors}


@app.get("/api/v1/admin/orders/export/excel", tags=["Admin - Orders"])
async def export_orders_excel(
    status_filter: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_staff)
):
    """
    Export all orders as Excel (XLSX) download.
    Includes order-level summary sheet AND detailed order items sheet.
    Color is exported as TEXT (Red, Blue, etc.) not hex codes.
    """
    import io
    from models.order import Order as OrderModel, OrderItem as OrderItemModel
    from models.product import Product as ProductModel
    from models.inventory import Inventory as InventoryModel
    from models.user import User as UserModel, UserProfile as UserProfileModel
    from sqlalchemy import and_

    q = db.query(OrderModel).options(
        selectinload(OrderModel.items).options(
            joinedload(OrderItemModel.inventory),
            joinedload(OrderItemModel.product)
        )
    )
    if status_filter:
        q = q.filter(OrderModel.status == OrderStatus(status_filter))
    if from_date:
        q = q.filter(OrderModel.created_at >= from_date)
    if to_date:
        q = q.filter(OrderModel.created_at <= to_date + " 23:59:59")
    orders = q.order_by(OrderModel.created_at.desc()).all()

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()

        # ===== STYLES =====
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=11, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        def style_header_row(ws):
            for cell in ws[1]:
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', wrap_text=True)
                cell.border = thin_border

        def auto_width_columns(ws, max_width=40):
            for column_cells in ws.columns:
                max_length = 0
                column_letter = column_cells[0].column_letter
                for cell in column_cells:
                    if cell.value:
                        max_length = max(max_length, min(len(str(cell.value)), max_width))
                ws.column_dimensions[column_letter].width = max_length + 2

        # ===== SHEET 1: Order Summary =====
        ws_summary = wb.active
        ws_summary.title = "Orders Summary"
        ws_summary.append([
            "Order ID", "Invoice #", "Order Date", "Customer Name", "Customer Email",
            "Phone", "Total Items", "Total Amount (INR)", "Payment Method",
            "Status", "Tracking #", "Shipping Address"
        ])
        style_header_row(ws_summary)

        # Batch fetch user info
        user_ids = list(set(o.user_id for o in orders))
        users = db.query(UserModel).filter(UserModel.id.in_(user_ids)).all()
        user_map = {u.id: u for u in users}
        profiles = db.query(UserProfileModel).filter(UserProfileModel.user_id.in_(user_ids)).all()
        profile_map = {p.user_id: p for p in profiles}

        for o in orders:
            profile = profile_map.get(o.user_id)
            customer_name = profile.full_name if profile and profile.full_name else (user_map.get(o.user_id).username if user_map.get(o.user_id) else "")
            customer_email = user_map.get(o.user_id).email if user_map.get(o.user_id) else ""
            phone = profile.phone if profile else ""
            total_items = sum(item.quantity for item in o.items) if o.items else 0

            ws_summary.append([
                o.id,
                o.invoice_number or f"INV-{o.id}",
                o.created_at.strftime("%Y-%m-%d %H:%M") if o.created_at else "",
                customer_name,
                customer_email,
                phone,
                total_items,
                float(o.total_amount) if o.total_amount else 0,
                o.payment_method or "",
                o.status.value if hasattr(o.status, 'value') else str(o.status),
                o.tracking_number or "",
                o.shipping_address or "",
            ])
        auto_width_columns(ws_summary)

        # ===== SHEET 2: Order Items Detail =====
        ws_items = wb.create_sheet("Order Items Detail")
        ws_items.append([
            "Order ID", "Invoice #", "Order Date", "Customer Name", "Customer Email",
            "Item #", "Product Name", "SKU", "Size", "Color",
            "HSN Code", "GST Rate (%)", "Quantity", "Unit Price (INR)",
            "Item Total (INR)", "Payment Status", "Order Status"
        ])
        style_header_row(ws_items)

        for o in orders:
            profile = profile_map.get(o.user_id)
            customer_name = profile.full_name if profile and profile.full_name else (user_map.get(o.user_id).username if user_map.get(o.user_id) else "")
            customer_email = user_map.get(o.user_id).email if user_map.get(o.user_id) else ""
            order_status = o.status.value if hasattr(o.status, 'value') else str(o.status)

            # Determine payment status from transaction
            payment_status = "Unknown"
            if o.transaction_id or o.razorpay_payment_id:
                payment_status = "Paid" if o.status not in ['cancelled'] else "Refunded"
                if o.status == 'cancelled':
                    payment_status = "Cancelled/Refunded"

            for idx, item in enumerate(o.items, 1):
                # Get color as TEXT (not hex)
                color_text = item.color or ""
                # If color is a hex code (starts with #), convert to name
                if color_text.startswith('#') and len(color_text) == 7:
                    color_text = _hex_to_color_name(color_text)

                size_text = item.size or ""
                hsn = item.hsn_code or (item.product.hsn_code if item.product else "")
                gst_rate = item.gst_rate if item.gst_rate is not None else (item.product.gst_rate if item.product else None)

                ws_items.append([
                    o.id,
                    o.invoice_number or f"INV-{o.id}",
                    o.created_at.strftime("%Y-%m-%d %H:%M") if o.created_at else "",
                    customer_name,
                    customer_email,
                    idx,
                    item.product_name or (item.product.name if item.product else "Unknown"),
                    item.sku or "",
                    size_text,
                    color_text,
                    hsn or "",
                    float(gst_rate) if gst_rate is not None else "",
                    item.quantity,
                    float(item.unit_price) if item.unit_price else 0,
                    float(item.price) if item.price else 0,
                    payment_status,
                    order_status,
                ])
        auto_width_columns(ws_items)

        # Set column widths for items sheet
        ws_items.column_dimensions['G'].width = 35  # Product Name
        ws_items.column_dimensions['J'].width = 15  # Color
        ws_items.column_dimensions['K'].width = 12  # HSN Code

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        from fastapi.responses import StreamingResponse as _SR
        return _SR(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=orders_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"}
        )
    except ImportError:
        raise HTTPException(status_code=501, detail="openpyxl not installed — use client-side export")


# ==================== Admin - Customer Activity & Reconciliation ====================

@app.get("/api/v1/admin/customers/{user_id}/activity", tags=["Admin - Customer Activity"])
async def get_customer_activity(
    user_id: int,
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=200),
    activity_type: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_admin)
):
    """
    Get customer activity timeline for admin review.
    Shows all logged activities for a specific customer.
    """
    from sqlalchemy import text

    skip = (page - 1) * limit

    query = """
        SELECT id, user_id, activity_type, resource_type, resource_id,
               details, ip_address, user_agent, created_at
        FROM customer_activity_logs
        WHERE user_id = :user_id
    """
    params = {"user_id": user_id, "skip": skip, "limit": limit}

    if activity_type:
        query += " AND activity_type = :activity_type"
        params["activity_type"] = activity_type
    if from_date:
        query += " AND created_at >= :from_date"
        params["from_date"] = from_date
    if to_date:
        query += " AND created_at <= :to_date"
        params["to_date"] = to_date + " 23:59:59"

    query += " ORDER BY created_at DESC LIMIT :limit OFFSET :skip"

    result = db.execute(text(query), params)
    rows = result.fetchall()

    # Get total count
    count_query = "SELECT COUNT(*) FROM customer_activity_logs WHERE user_id = :user_id"
    count_params = {"user_id": user_id}
    if activity_type:
        count_query += " AND activity_type = :activity_type"
        count_params["activity_type"] = activity_type
    total = db.execute(text(count_query), count_params).scalar()

    activities = []
    for row in rows:
        activities.append({
            "id": row[0],
            "user_id": row[1],
            "activity_type": row[2],
            "resource_type": row[3],
            "resource_id": row[4],
            "details": row[5] if isinstance(row[5], dict) else {},
            "ip_address": row[6],
            "user_agent": row[7],
            "created_at": row[8].isoformat() if row[8] else None,
        })

    return {
        "activities": activities,
        "total": total,
        "page": page,
        "limit": limit,
        "has_more": skip + limit < total
    }


@app.get("/api/v1/admin/reconciliation/summary", tags=["Admin - Reconciliation"])
async def get_reconciliation_summary(
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_admin)
):
    """
    Get order/payment reconciliation summary.
    Shows orphaned payments, orders without payments, and amount mismatches.
    """
    from sqlalchemy import text

    # Run reconciliation function
    result = db.execute(text("SELECT * FROM run_order_reconciliation()"))
    issues = [{"issue_type": row[0], "issue_count": row[1], "details": row[2]} for row in result.fetchall()]

    # Get sample orphaned payments
    orphaned = db.execute(text("""
        SELECT * FROM v_orphaned_payments_enhanced LIMIT 10
    """)).fetchall()

    # Get sample orders without payment
    no_payment = db.execute(text("""
        SELECT * FROM v_orders_without_payment LIMIT 10
    """)).fetchall()

    # Get sample amount mismatches
    mismatches = db.execute(text("""
        SELECT * FROM v_payment_amount_mismatches LIMIT 10
    """)).fetchall()

    return {
        "summary": issues,
        "orphaned_payments_sample": [dict(r._mapping) for r in orphaned],
        "orders_without_payment_sample": [dict(r._mapping) for r in no_payment],
        "amount_mismatches_sample": [dict(r._mapping) for r in mismatches],
        "timestamp": datetime.now(timezone.utc).isoformat()
    }


@app.post("/api/v1/admin/reconciliation/fix-orphaned-payments", tags=["Admin - Reconciliation"])
async def fix_orphaned_payments(
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_admin)
):
    """
    Trigger recovery for orphaned payments.
    Calls the payment service recovery job.
    """
    import httpx
    import os

    payment_service_url = os.getenv("PAYMENT_SERVICE_URL", "http://payment:5003")
    internal_secret = os.getenv("INTERNAL_SERVICE_SECRET")

    if not internal_secret:
        raise HTTPException(status_code=500, detail="Internal service secret not configured")

    try:
        with httpx.Client(timeout=60.0) as client:
            response = client.post(
                f"{payment_service_url}/api/v1/admin/recovery/run",
                headers={"X-Internal-Secret": internal_secret}
            )
            if response.status_code == 200:
                return {"success": True, "result": response.json()}
            else:
                return {"success": False, "error": response.text[:500]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Recovery failed: {str(e)}")


# Internal service-to-service routes have been moved to routes/internal.py.


# ==================== Address Routes ====================

@app.post("/api/v1/addresses", response_model=AddressResponse,
          status_code=status.HTTP_201_CREATED,
          tags=["Addresses"])
async def create_address(
    address_data: AddressCreate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Create a new address."""
    address_service = AddressService(db)
    return address_service.create_address(current_user["user_id"], address_data)


@app.get("/api/v1/addresses", response_model=List[AddressResponse],
         tags=["Addresses"])
async def list_addresses(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """List current user's addresses."""
    address_service = AddressService(db)
    return address_service.get_user_addresses(current_user["user_id"])


@app.get("/api/v1/addresses/{address_id}", response_model=AddressResponse,
         tags=["Addresses"])
async def get_address(
    address_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Get address by ID."""
    address_service = AddressService(db)
    address = address_service.get_address_by_id(address_id, current_user["user_id"])
    
    if not address:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Address not found"
        )
    
    return address


@app.patch("/api/v1/addresses/{address_id}", response_model=AddressResponse,
           tags=["Addresses"])
async def update_address(
    address_id: int,
    address_data: AddressUpdate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Update an address."""
    address_service = AddressService(db)
    return address_service.update_address(
        address_id,
        current_user["user_id"],
        address_data
    )


@app.delete("/api/v1/addresses/{address_id}",
            status_code=status.HTTP_204_NO_CONTENT,
            tags=["Addresses"])
async def delete_address(
    address_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Delete an address."""
    address_service = AddressService(db)
    address_service.delete_address(address_id, current_user["user_id"])
    return


# ==================== Review Routes ====================

@app.post("/api/v1/reviews", response_model=ReviewResponse,
          status_code=status.HTTP_201_CREATED,
          tags=["Reviews"])
async def create_review(
    review_data: ReviewCreate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Create a product review."""
    review_service = ReviewService(db)
    return review_service.create_review(
        current_user["user_id"],
        product_id=review_data.product_id,
        rating=review_data.rating,
        title=review_data.title,
        comment=review_data.comment,
        order_id=review_data.order_id,
        image_urls=review_data.image_urls or []
    )


@app.post("/api/v1/reviews/upload-image",
          tags=["Reviews"])
async def upload_review_image(
    request: Request,
    file: UploadFile = File(..., description="Review image file"),
    current_user: dict = Depends(get_current_user)
):
    """
    Upload a single image for a review.
    Returns the public URL of the uploaded image.
    Supports JPG, PNG, WebP (max 5MB).
    Rate limited to 10 uploads per minute per user.
    """
    # Rate limiting: 10 uploads per minute per user
    if not _check_rate_limit(
        request,
        "review_upload",
        limit=10,
        window=60,
        user_identifier=str(current_user["user_id"])
    ):
        raise HTTPException(
            status_code=status.HTTP_429_TOO_MANY_REQUESTS,
            detail="Too many image uploads. Please wait before uploading more images."
        )

    from service.r2_service import r2_service
    import uuid

    # Validate file type
    allowed_types = ['image/jpeg', 'image/jpg', 'image/png', 'image/webp']
    if file.content_type not in allowed_types:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Invalid file type. Allowed types: JPG, PNG, WebP"
        )

    # Validate file size (5MB max)
    file.file.seek(0, 2)  # Seek to end
    file_size = file.file.tell()
    file.file.seek(0)  # Reset to beginning
    
    if file_size > 5 * 1024 * 1024:  # 5MB
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File size must be less than 5MB"
        )

    try:
        # Upload to R2 with unique filename
        ext = file.filename.split('.')[-1] if '.' in file.filename else 'jpg'
        unique_name = f"review_{current_user['user_id']}_{uuid.uuid4().hex[:8]}.{ext}"
        
        image_url = await r2_service.upload_image(
            file,
            folder="reviews",
            custom_filename=unique_name
        )
        
        return {
            "url": image_url,
            "filename": unique_name,
            "size": file_size
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Review image upload failed: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to upload image: {str(e)}"
        )


@app.get("/api/v1/products/{product_id}/reviews", response_model=List[ReviewResponse],
         tags=["Reviews"])
async def get_product_reviews_main(
    product_id: int,
    approved_only: bool = True,
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=100),
    db: Session = Depends(get_db)
):
    """Get reviews for a product."""
    review_service = ReviewService(db)
    return review_service.get_product_reviews(
        product_id,
        approved_only=approved_only,
        skip=skip,
        limit=limit
    )


@app.post("/api/v1/reviews/{review_id}/helpful",
          tags=["Reviews"])
async def mark_review_helpful(
    review_id: int,
    db: Session = Depends(get_db)
):
    """Mark a review as helpful."""
    review_service = ReviewService(db)
    review_service.mark_helpful(review_id)
    return {"message": "Review marked as helpful"}


@app.delete("/api/v1/reviews/{review_id}",
            status_code=status.HTTP_204_NO_CONTENT,
            tags=["Reviews"])
async def delete_review(
    review_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Delete own review."""
    review_service = ReviewService(db)
    review_service.delete_review(review_id, current_user["user_id"])
    return


@app.post("/api/v1/admin/reviews/{review_id}/approve",
          response_model=ReviewResponse,
          tags=["Admin - Reviews"])
async def approve_review(
    review_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_staff)
):
    """Approve a review (admin/staff)."""
    review_service = ReviewService(db)
    return review_service.approve_review(review_id)


# ==================== Return Routes ====================

@app.post("/api/v1/returns", response_model=ReturnRequestResponse,
          status_code=status.HTTP_201_CREATED,
          tags=["Returns"])
async def create_return_request(
    return_data: ReturnRequestCreate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Create a return request."""
    return_service = ReturnService(db)
    return return_service.create_return(current_user["user_id"], return_data)


@app.post("/api/v1/returns/upload-video",
          status_code=status.HTTP_201_CREATED,
          tags=["Returns"])
async def upload_return_video(
    file: UploadFile = File(..., description="Video file for return evidence"),
    current_user: dict = Depends(get_current_user)
):
    """
    Upload a video for return request evidence.
    
    - Accepts: mp4, mov, webm formats
    - Max file size: 50MB
    - Returns the video URL after successful upload
    """
    # Import the r2 service
    from service.r2_service import r2_service
    
    try:
        video_url = await r2_service.upload_video(file, folder="returns")
        return {
            "video_url": video_url,
            "message": "Video uploaded successfully"
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to upload video: {str(e)}"
        )


@app.get("/api/v1/returns", response_model=List[ReturnRequestResponse],
         tags=["Returns"])
async def list_returns(
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """List current user's return requests."""
    return_service = ReturnService(db)
    return return_service.get_user_returns(
        current_user["user_id"],
        skip=skip,
        limit=limit
    )


@app.get("/api/v1/returns/{return_id}", response_model=ReturnRequestResponse,
         tags=["Returns"])
async def get_return(
    return_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Get return request details."""
    return_service = ReturnService(db)
    return_request = return_service.get_return_by_id(
        return_id,
        user_id=current_user["user_id"]
    )
    
    if not return_request:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Return request not found"
        )
    
    return return_request


# ==================== Admin Return Routes ====================

@app.get("/api/v1/admin/returns", response_model=List[ReturnRequestResponse],
         tags=["Admin - Returns"])
async def list_all_returns(
    status_filter: Optional[str] = None,
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_staff)
):
    """List all return requests (admin/staff)."""
    return_service = ReturnService(db)
    return_status = ReturnStatus(status_filter) if status_filter else None
    return return_service.get_all_returns(status=return_status, skip=skip, limit=limit)


@app.post("/api/v1/admin/returns/{return_id}/approve",
          response_model=ReturnRequestResponse,
          tags=["Admin - Returns"])
async def approve_return(
    return_id: int,
    refund_amount: Optional[float] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_staff)
):
    """Approve a return request (admin/staff)."""
    return_service = ReturnService(db)
    return return_service.approve_return(
        return_id,
        approved_by=current_user["user_id"],
        refund_amount=Decimal(refund_amount) if refund_amount else None
    )


@app.post("/api/v1/admin/returns/{return_id}/reject",
          response_model=ReturnRequestResponse,
          tags=["Admin - Returns"])
async def reject_return(
    return_id: int,
    reason: str,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_staff)
):
    """Reject a return request (admin/staff)."""
    return_service = ReturnService(db)
    return return_service.reject_return(
        return_id,
        approved_by=current_user["user_id"],
        rejection_reason=reason
    )


@app.post("/api/v1/admin/returns/{return_id}/receive",
          response_model=ReturnRequestResponse,
          tags=["Admin - Returns"])
async def mark_return_received(
    return_id: int,
    tracking_number: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_staff)
):
    """Mark return item as received (admin/staff)."""
    return_service = ReturnService(db)
    return return_service.mark_item_received(
        return_id,
        tracking_number=tracking_number
    )


@app.post("/api/v1/admin/returns/{return_id}/refund",
          response_model=ReturnRequestResponse,
          tags=["Admin - Returns"])
async def process_return_refund(
    return_id: int,
    refund_transaction_id: str,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_staff)
):
    """Mark return as refunded with transaction ID (admin/staff)."""
    return_service = ReturnService(db)
    return return_service.mark_refunded(
        return_id,
        refund_transaction_id=refund_transaction_id
    )


@app.post("/api/v1/returns/{return_id}/cancel",
          response_model=ReturnRequestResponse,
          tags=["Returns"])
async def cancel_return_request(
    return_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Cancel a return request (customer). Only allowed for 'requested' status."""
    return_service = ReturnService(db)
    return_request = return_service.get_return_by_id(
        return_id,
        user_id=current_user["user_id"]
    )
    
    if not return_request:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Return request not found"
        )
    
    if return_request.status != ReturnStatus.REQUESTED:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Can only cancel return requests with 'requested' status"
        )
    
    # Delete the return request
    db.delete(return_request)
    db.commit()
    
    return {"message": "Return request cancelled successfully", "id": return_id}


# NOTE: /api/v1/products/browse is handled by routes/products.py (registered via router at line 560).
# The route module version is the SINGLE source of truth for product browsing.
# The duplicate endpoint below was removed to prevent confusion and stale cache issues.
# (Previously caused products to show as out-of-stock due to missing db parameter in _enrich_product)


@app.get("/api/v1/products/{product_id}/related", tags=["Products"])
async def get_related_products_main(
    product_id: int,
    limit: int = Query(8, ge=1, le=20),
    db: Session = Depends(get_db),
    current_user: Optional[dict] = Depends(get_current_user_optional)
):
    """Get related products based on same category."""
    user_role = current_user.get("role") if current_user else None
    
    product = db.query(Product).filter(Product.id == product_id, Product.is_active == True).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    related = db.query(Product).filter(
        Product.category_id == product.category_id,
        Product.id != product_id,
        Product.is_active == True
    ).order_by(func.random()).limit(limit).all()
    return {"products": [_enrich_product(p, user_role) for p in related]}


# ==================== Cart Enhancements ====================

@app.put("/api/v1/cart/{user_id}/update-quantity", response_model=CartResponse,
         tags=["Cart"])
async def update_cart_quantity(
    user_id: int,
    product_id: int,
    quantity: int = Query(ge=1),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Update quantity for an item already in the cart (legacy endpoint — delegates to CartService)."""
    if current_user["user_id"] != user_id and current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not authorized")
    cart_data = CartConcurrencyManager.update_cart_item_locked(user_id, product_id, quantity, db)
    return CartResponse(**cart_data)


@app.get("/api/v1/cart/{user_id}/summary", tags=["Cart"])
async def cart_summary(user_id: int):
    """Get cart summary with calculated totals, shipping, and any applied promo."""
    cart_key = f"cart:{user_id}"
    cart_data = redis_client.get_cache(cart_key)

    if not cart_data or not cart_data.get("items"):
        return {"subtotal": 0, "total_items": 0, "discount": 0, "shipping": 0, "total": 0, "items": []}

    subtotal = sum(i["price"] * i["quantity"] for i in cart_data["items"])
    total_items = sum(i["quantity"] for i in cart_data["items"])

    return {
        "subtotal": subtotal,
        "total_items": total_items,
        "shipping": 0,
        "total": subtotal,
        "items": cart_data["items"],
    }


# ==================== Customer Profile ====================

@app.get("/api/v1/me", tags=["Customer Profile"])
async def get_customer_me(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Backward-compatible alias for clients calling /api/v1/me."""
    return await get_customer_profile(db=db, current_user=current_user)


@app.get("/api/v1/me/profile", tags=["Customer Profile"])
async def get_customer_profile(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Get current customer's complete profile with stats."""
    user_id = current_user["user_id"]

    user = db.execute(text(
        "SELECT id, email, username, full_name, phone, role, is_active, created_at "
        "FROM users WHERE id = :id"
    ), {"id": user_id}).fetchone()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    order_stats = db.execute(text(
        "SELECT COUNT(*), COALESCE(SUM(total_amount), 0) FROM orders WHERE user_id = :uid AND status != 'cancelled'"
    ), {"uid": user_id}).fetchone()

    address_count = db.execute(text("SELECT COUNT(*) FROM addresses WHERE user_id = :uid"), {"uid": user_id}).scalar() or 0
    review_count = db.execute(text("SELECT COUNT(*) FROM reviews WHERE user_id = :uid"), {"uid": user_id}).scalar() or 0

    return {
        "user": {"id": user[0], "email": user[1], "username": user[2], "full_name": user[3], "phone": user[4], "role": str(user[5]), "is_active": user[6], "member_since": str(user[7])},
        "stats": {
            "total_orders": order_stats[0] if order_stats else 0,
            "total_spent": float(order_stats[1]) if order_stats else 0,
            "saved_addresses": address_count,
            "reviews_written": review_count,
        }
    }


@app.get("/api/v1/me/order-history", tags=["Customer Profile"])
async def get_order_history(
    status_filter: Optional[str] = None,
    skip: int = Query(0, ge=0),
    limit: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Get paginated order history with filtering."""
    user_id = current_user["user_id"]
    where = "WHERE o.user_id = :uid"
    params = {"uid": user_id, "lim": limit, "off": skip}
    if status_filter:
        where += " AND o.status = :status"
        params["status"] = status_filter

    rows = db.execute(text(f"""
        SELECT o.id, o.total_amount, o.status, o.created_at,
               (SELECT COUNT(*) FROM order_items WHERE order_id = o.id) as item_count
        FROM orders o {where} ORDER BY o.created_at DESC LIMIT :lim OFFSET :off
    """), params).fetchall()
    total = db.execute(text(f"SELECT COUNT(*) FROM orders o {where}"), params).scalar()

    return {
        "orders": [{"id": r[0], "total_amount": float(r[1]), "status": r[2], "created_at": str(r[3]), "item_count": r[4]} for r in rows],
        "total": total,
        "page": skip // limit + 1,
    }


# Customer chat endpoints have been moved to routes/chat.py.


# Landing-page endpoints have been moved to routes/landing.py.


# ==================== Exchange Requests ====================

@app.post("/api/v1/returns/{return_id}/exchange", tags=["Returns"])
async def request_exchange(
    return_id: int,
    exchange_product_id: int,
    exchange_variant_id: Optional[int] = None,
    notes: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Convert a return request into an exchange request."""
    user_id = current_user["user_id"]

    ret = db.execute(text(
        "SELECT id, status, user_id, order_id FROM return_requests WHERE id = :rid"
    ), {"rid": return_id}).fetchone()

    if not ret:
        raise HTTPException(status_code=404, detail="Return request not found")
    if ret[2] != user_id:
        raise HTTPException(status_code=403, detail="Not your return request")
    if ret[1] not in ("requested", "approved"):
        raise HTTPException(status_code=400, detail="Return cannot be converted to exchange")

    # Validate exchange product exists and is active
    product = db.query(Product).filter(Product.id == exchange_product_id, Product.is_active == True).first()
    if not product:
        raise HTTPException(status_code=404, detail="Exchange product not found")

    # Validate variant belongs to product if variant_id provided
    if exchange_variant_id:
        from models.inventory import Inventory as InventoryModel
        variant = db.query(InventoryModel).filter(
            InventoryModel.id == exchange_variant_id,
            InventoryModel.product_id == exchange_product_id,
            InventoryModel.is_active == True
        ).first()
        if not variant:
            raise HTTPException(status_code=404, detail="Exchange variant not found or doesn't belong to product")
        
        # Check variant stock availability
        available = (variant.quantity or 0) - (variant.reserved_quantity or 0)
        if available <= 0:
            raise HTTPException(status_code=400, detail="Exchange variant is out of stock")
    else:
        # Check product stock via inventory aggregation
        from models.inventory import Inventory as InventoryModel
        total = db.execute(
            text("SELECT COALESCE(SUM(quantity - reserved_quantity), 0) FROM inventory WHERE product_id = :pid AND is_active = true"),
            {"pid": exchange_product_id}
        ).scalar()
        if (total or 0) <= 0:
            raise HTTPException(status_code=400, detail="Exchange product is out of stock")

    # Validate exchange policy (same price or higher value)
    # Get original order item to compare prices - using order_id from return_requests
    order_id = ret[3]  # order_id from return_requests
    original_items = db.execute(text("""
        SELECT product_id, inventory_id, unit_price
        FROM order_items
        WHERE order_id = :oid
    """), {"oid": order_id}).fetchall()
    
    if original_items:
        # Get exchange item price
        if exchange_variant_id:
            exchange_price = db.execute(text("""
                SELECT COALESCE(variant_price, 0) FROM inventory
                WHERE id = :vid AND product_id = :pid
            """), {"vid": exchange_variant_id, "pid": exchange_product_id}).fetchone()
            exchange_price = float(exchange_price[0]) if exchange_price and exchange_price[0] else float(product.base_price or 0)
        else:
            exchange_price = float(product.base_price or 0)
        
        # Use the first item's price for comparison (or sum all items)
        original_price = sum(float(item[2]) for item in original_items)
        
        # Allow exchange only if price difference is reasonable (within 20% or higher value)
        if exchange_price < original_price * 0.8:
            raise HTTPException(
                status_code=400, 
                detail=f"Exchange item value too low. Original: ${original_price:.2f}, Exchange: ${exchange_price:.2f}"
            )

    # Update the return request with exchange info
    db.execute(text(
        "UPDATE return_requests SET description = CONCAT(COALESCE(description, ''), ' | Exchange to product #', :pid, '. ', :notes) WHERE id = :rid"
    ), {"pid": exchange_product_id, "notes": notes or '', "rid": return_id})
    db.commit()
    return {"message": "Exchange request submitted", "return_id": return_id, "exchange_product_id": exchange_product_id}


# About / Contact page endpoints have moved to routes/landing.py.


# ==================== Run Server ====================

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(
        "main:app",
        host="0.0.0.0",
        port=5002,
        reload=settings.DEBUG,
        log_level=settings.LOG_LEVEL.lower()
    )

