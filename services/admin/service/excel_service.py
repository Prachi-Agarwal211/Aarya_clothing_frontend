"""Excel import/export utilities for the admin service.

Uses openpyxl. Provides helpers for:
- Products: bulk export and import (with variants flattened to one row per variant)
- Orders: export with line items
- Inventory: export current stock levels
- Templates: downloadable empty templates so admins know expected columns
"""
from __future__ import annotations

import io
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


# --------------------------------------------------------------------------- #
# Generic helpers                                                              #
# --------------------------------------------------------------------------- #

def _autosize(ws) -> None:
    """Best-effort autosize; openpyxl can't measure exactly so we cap."""
    for col_idx, col in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = min(len(value), 60)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, max_len + 2)


def _write_sheet(ws, headers: list[str], rows: Iterable[Iterable[Any]]) -> None:
    ws.append(headers)
    for r in rows:
        ws.append(list(r))
    _autosize(ws)


def _wb_to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# --------------------------------------------------------------------------- #
# Products                                                                     #
# --------------------------------------------------------------------------- #

PRODUCT_HEADERS = [
    "product_id",
    "name",
    "slug",
    "description",
    "price",
    "collection_id",
    "primary_image",
    "is_active",
    "is_featured",
    "is_new_arrival",
    "variant_sku",
    "size",
    "color",
    "color_hex",
    "variant_image",
    "quantity",
    "low_stock_threshold",
]


def export_products(rows: list[dict]) -> bytes:
    """Flatten products + variants to one row per variant.

    `rows` is the raw join of products LEFT JOIN product_variants returned by
    the caller — each dict already has all PRODUCT_HEADERS keys (missing keys
    default to None).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "products"
    _write_sheet(
        ws,
        PRODUCT_HEADERS,
        ((row.get(h) for h in PRODUCT_HEADERS) for row in rows),
    )
    return _wb_to_bytes(wb)


def parse_products_import(file_bytes: bytes) -> list[dict]:
    """Parse an uploaded products workbook back into row dicts.

    Validates that all required columns exist; raises ValueError otherwise.
    Returns rows with original column names preserved so the caller can decide
    whether each row is a create or an update (presence of product_id).
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    missing = [h for h in PRODUCT_HEADERS if h not in headers]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")
    out: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        out.append(dict(zip(headers, row)))
    return out


def products_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "products"
    ws.append(PRODUCT_HEADERS)
    sample = [
        None, "Sample Saree", "sample-saree", "A short description",
        1499, 1, "products/sample.jpg", True, False, True,
        "SAMPLE-S-RED", "S", "Red", "#DC2626",
        "products/sample-red.jpg", 25, 5,
    ]
    ws.append(sample)
    _autosize(ws)
    return _wb_to_bytes(wb)


# --------------------------------------------------------------------------- #
# Orders                                                                       #
# --------------------------------------------------------------------------- #

ORDER_HEADERS = [
    "invoice_number",
    "order_id",
    "status",
    "payment_status",
    "customer_name",
    "customer_email",
    "total_amount",
    "subtotal",
    "tax_amount",
    "shipping_amount",
    "created_at",
    "product_name",
    "sku",
    "size",
    "color",
    "quantity",
    "unit_price",
    "line_total",
]


def export_orders(rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "orders"
    _write_sheet(
        ws,
        ORDER_HEADERS,
        ((row.get(h) for h in ORDER_HEADERS) for row in rows),
    )
    return _wb_to_bytes(wb)


# --------------------------------------------------------------------------- #
# Inventory                                                                    #
# --------------------------------------------------------------------------- #

INVENTORY_HEADERS = [
    "variant_id",
    "product_id",
    "product_name",
    "sku",
    "size",
    "color",
    "color_hex",
    "quantity",
    "reserved_quantity",
    "low_stock_threshold",
    "is_active",
    "updated_at",
]


def export_inventory(rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "inventory"
    _write_sheet(
        ws,
        INVENTORY_HEADERS,
        ((row.get(h) for h in INVENTORY_HEADERS) for row in rows),
    )
    return _wb_to_bytes(wb)


def inventory_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "inventory"
    ws.append(["sku", "quantity", "low_stock_threshold"])
    ws.append(["SAMPLE-S-RED", 25, 5])
    _autosize(ws)
    return _wb_to_bytes(wb)


# --------------------------------------------------------------------------- #
# Bulk POD Upload                                                              #
# --------------------------------------------------------------------------- #

POD_HEADERS = ["order_id", "invoice_number", "customer", "tracking_number", "courier_name"]

def pod_template(orders: list[dict]) -> bytes:
    """Generate a template for bulk POD upload from confirmed orders."""
    wb = Workbook()
    ws = wb.active
    ws.title = "shipments"
    ws.append(POD_HEADERS)
    for o in orders:
        ws.append([
            o.get("id"),
            o.get("invoice_number"),
            o.get("customer_name"),
            "", # empty tracking_number
            o.get("courier_name") or "Delhivery" # default courier
        ])
    _autosize(ws)
    return _wb_to_bytes(wb)

def parse_pod_import(file_bytes: bytes) -> list[dict]:
    """Parse bulk POD shipments; requires 'order_id' and 'tracking_number'."""
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    headers = [str(c.value).strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    
    if "order_id" not in headers or "tracking_number" not in headers:
        raise ValueError("Bulk POD upload requires 'order_id' and 'tracking_number' columns")
        
    out: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        data = dict(zip(headers, row))
        # Ensure values are clean
        if data.get("order_id") and data.get("tracking_number"):
            out.append(data)
    return out
