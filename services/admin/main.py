"""Aarya Clothing — Admin & Staff Service (Port 8004)."""

from contextlib import asynccontextmanager
from datetime import datetime, timedelta, timezone
import json
import os
import re
from typing import List, Optional
import asyncio
from fastapi import (
    FastAPI,
    Depends,
    HTTPException,
    Query,
    Request,
    WebSocket,
    WebSocketDisconnect,
    UploadFile,
    File,
    status,
)
import httpx
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, Response
from sqlalchemy.orm import Session
from sqlalchemy import func, desc, text

from core.config import settings
from shared.response_schemas import BaseResponse
from core.redis_client import redis_client
from shared.event_bus import EventBus
from service.event_handlers import OrderCreatedHandler
from core.validation import (
    validate_order_status,
    validate_user_role,
    validate_search_query,
    validate_pagination,
    sanitize_order_updates,
)
from core.exception_handler import (
    setup_exception_handlers,
    log_security_event,
    log_business_event,
    AdminServiceException,
    ValidationException,
    DatabaseException,
)
from database.database import get_db, init_db, Base, engine
from shared.auth_middleware import (
    get_current_user,
    require_admin,
    require_staff,
    require_super_admin,
    initialize_auth_middleware,
)
from models.chat import ChatRoom, ChatMessage, ChatStatus, SenderType
from models.landing_config import LandingConfig, LandingImage, LandingProduct
from models.analytics import (
    AnalyticsCache,
    InventoryMovement,
    StaffTask,
    StaffNotification,
)
from service.r2_service import r2_service
from shared.request_id_middleware import RequestIDMiddleware
from shared.error_responses import register_error_handlers
from schemas.admin import (
    DashboardOverview,
    InventoryAlert,
    RevenueAnalytics,
    RevenueData,
    CustomerAnalytics,
    TopProduct,
    TopProductsAnalytics,
    OrderStatusUpdate,
    BulkOrderUpdate,
    OrderResponse,
    AddStockRequest,
    AdjustStockRequest,
    BulkInventoryUpdate,
    InventoryAdjustRequest,
    VariantCreate,
    VariantUpdate,
    InventoryMovementResponse,
    StaffDashboard,
    OrderProcessRequest,
    OrderShipRequest,
    ReservationRelease,
    TaskComplete,
    UserListItem,
    UserStatusUpdate,
    BulkUserStatusUpdate,
    ChatRoomCreate,
    ChatMessageCreate,
    ChatRoomResponse,
    ChatMessageResponse,
    LandingConfigUpdate,
    LandingConfigResponse,
    LandingImageCreate,
    LandingImageResponse,
    LandingProductCreate,
    LandingProductUpdate,
    ProductCreate,
    ProductUpdate,
    CategoryCreate,
    CategoryUpdate,
    BulkPriceUpdate,
    BulkStatusUpdate,
    BulkCollectionAssign,
    BulkCollectionStatusUpdate,
    BulkCollectionReorder,
)


import logging

logger = logging.getLogger(__name__)

# Global instances
event_bus = None


@asynccontextmanager
async def lifespan(app: FastAPI):
    """Startup and shutdown events."""
    global event_bus
    logger.info(f"Starting {settings.SERVICE_NAME}...")
    init_db()

    # Initialize auth middleware
    initialize_auth_middleware(
        secret_key=settings.SECRET_KEY,
        algorithm=settings.ALGORITHM,
        redis_client=redis_client,
    )

    # Ensure Redis connects and pub/sub is active
    if redis_client.is_connected():
        logger.info("Redis connected for Admin service")
    else:
        logger.warning("Redis connection failed for Admin service")

    # Initialize Event Bus
    event_bus = EventBus(redis_client=redis_client, service_name="admin_service")

    # Register event handlers
    order_handler = OrderCreatedHandler()
    event_bus.register_handler(order_handler)

    # Validate AI provider config at startup (no network call — avoids blocking startup)
    from service.ai_service import _get_active_provider
    try:
        provider = _get_active_provider()
        if not provider.get("key") or not provider.get("base_url"):
            raise ValueError("Provider key or base_url missing")
        logger.info(f"AI service ready (Provider: {provider['name']}, Model: {provider['model']})")
    except ValueError as e:
        logger.warning(f"AI service disabled: {e}")
    except Exception as e:
        logger.warning(f"AI provider config check failed: {e}. AI features may not work.")

    yield

    # Cleanup event bus history on shutdown
    if event_bus:
        event_bus.clear_history()

    logger.info(f"Shutting down {settings.SERVICE_NAME}...")


_env = os.environ.get("ENVIRONMENT", "production")
app = FastAPI(
    title="Aarya Clothing - Admin & Staff Service",
    description="Dashboard, Analytics, Order/Inventory Management, Chat, Landing Config",
    version="1.0.0",
    lifespan=lifespan,
    docs_url="/docs" if _env != "production" else None,
    redoc_url="/redoc" if _env != "production" else None,
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=settings.ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "PATCH", "DELETE", "OPTIONS"],
    allow_headers=[
        "Content-Type",
        "Authorization",
        "X-Requested-With",
        "Accept",
        "Origin",
        "X-CSRF-Token",
    ],
)

# Prometheus metrics — /metrics endpoint for scraping
try:
    from prometheus_fastapi_instrumentator import Instrumentator
    Instrumentator().instrument(app).expose(app, endpoint="/metrics")
except Exception:
    pass  # Graceful degradation if prometheus lib is missing

# Request ID
app.add_middleware(RequestIDMiddleware)

# Standardized error handlers (complements admin's own exception handler)
register_error_handlers(app)

# Mount routers for AI dashboard and staff management endpoints
from routes.ai_dashboard_staff import router as ai_staff_router
from routes.ai_dashboard_staff_part2 import router as ai_staff_router2
app.include_router(ai_staff_router)
app.include_router(ai_staff_router2)


# ==================== Health ====================


@app.get("/health", tags=["Health"])
async def health_root(db: Session = Depends(get_db)):
    db_status = "healthy"
    try:
        db.execute(text("SELECT 1"))
    except Exception:
        db_status = "unhealthy"
    return {
        "status": "healthy" if db_status == "healthy" else "degraded",
        "service": "admin",
        "db": db_status,
    }


# Setup global exception handlers
setup_exception_handlers(app)


# ==================== Admin Dashboard ====================


@app.get(
    "/api/v1/admin/dashboard/overview",
    response_model=DashboardOverview,
    tags=["Admin Dashboard"],
)
async def get_dashboard_overview(
    db: Session = Depends(get_db), user: dict = Depends(require_admin)
):
    """Aggregate dashboard stats: revenue, orders, customers, inventory alerts."""
    cache_key = "admin:dashboard:overview"
    cached = redis_client.get_cache(cache_key)
    if cached:
        return cached

    today = datetime.now(timezone.utc).date()

    # Single consolidated query for all scalar stats (was 8 separate queries)
    stats = db.execute(
        text("""
        SELECT
            (SELECT COALESCE(SUM(total_amount), 0) FROM orders WHERE status != 'cancelled') AS total_revenue,
            (SELECT COUNT(*) FROM orders) AS total_orders,
            (SELECT COUNT(*) FROM users WHERE role = 'customer') AS total_customers,
            (SELECT COUNT(*) FROM products WHERE is_active = true) AS total_products,
            (SELECT COUNT(*) FROM orders WHERE status = 'confirmed') AS pending_orders,
            (SELECT COALESCE(SUM(total_amount), 0) FROM orders WHERE DATE(created_at) = :today AND status != 'cancelled') AS today_revenue,
            (SELECT COUNT(*) FROM orders WHERE DATE(created_at) = :today) AS today_orders,
            (SELECT COUNT(*) FROM inventory WHERE quantity <= low_stock_threshold AND quantity > 0) AS low_stock,
            (SELECT COUNT(*) FROM inventory WHERE quantity = 0) AS out_of_stock
    """),
        {"today": today},
    ).fetchone()

    recent = db.execute(
        text(
            "SELECT id, user_id, total_amount, status, created_at FROM orders ORDER BY created_at DESC LIMIT 5"
        )
    ).fetchall()
    recent_orders = [
        {
            "id": r[0],
            "user_id": r[1],
            "total_amount": float(r[2]),
            "status": r[3],
            "created_at": str(r[4]),
        }
        for r in recent
    ]

    result = DashboardOverview(
        total_revenue=float(stats[0]),
        total_orders=stats[1],
        total_customers=stats[2],
        total_products=stats[3],
        pending_orders=stats[4],
        today_revenue=float(stats[5]),
        today_orders=stats[6],
        inventory_alerts=InventoryAlert(low_stock=stats[7], out_of_stock=stats[8]),
        recent_orders=recent_orders,
    )
    redis_client.set_cache(cache_key, result.model_dump(), ttl=30)
    return result


# ==================== Analytics ====================


@app.get(
    "/api/v1/admin/analytics/revenue",
    response_model=RevenueAnalytics,
    tags=["Analytics"],
)
async def get_revenue_analytics(
    period: str = Query("30d", regex="^(7d|30d|90d|1y)$"),
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    days = {"7d": 7, "30d": 30, "90d": 90, "1y": 365}[period]
    since = datetime.now(timezone.utc) - timedelta(days=days)
    rows = db.execute(
        text(
            "SELECT DATE(created_at) as day, COALESCE(SUM(total_amount),0), COUNT(*) "
            "FROM orders WHERE created_at >= :since AND status != 'cancelled' "
            "GROUP BY DATE(created_at) ORDER BY day"
        ),
        {"since": since},
    ).fetchall()
    period_data = [
        RevenueData(period=str(r[0]), revenue=float(r[1]), orders=r[2]) for r in rows
    ]
    total = sum(d.revenue for d in period_data)
    return RevenueAnalytics(total_revenue=total, period_data=period_data)


@app.get(
    "/api/v1/admin/analytics/customers",
    response_model=CustomerAnalytics,
    tags=["Analytics"],
)
async def get_customer_analytics(
    db: Session = Depends(get_db), user: dict = Depends(require_admin)
):
    now = datetime.now(timezone.utc)
    total = (
        db.execute(text("SELECT COUNT(*) FROM users WHERE role = 'customer'")).scalar()
        or 0
    )
    today = (
        db.execute(
            text(
                "SELECT COUNT(*) FROM users WHERE role='customer' AND DATE(created_at)=:d"
            ),
            {"d": now.date()},
        ).scalar()
        or 0
    )
    week = (
        db.execute(
            text(
                "SELECT COUNT(*) FROM users WHERE role='customer' AND created_at >= :d"
            ),
            {"d": now - timedelta(days=7)},
        ).scalar()
        or 0
    )
    month = (
        db.execute(
            text(
                "SELECT COUNT(*) FROM users WHERE role='customer' AND created_at >= :d"
            ),
            {"d": now - timedelta(days=30)},
        ).scalar()
        or 0
    )
    returning = (
        db.execute(
            text(
                "SELECT COUNT(DISTINCT user_id) FROM orders GROUP BY user_id HAVING COUNT(*) > 1"
            )
        ).scalar()
        or 0
    )
    return CustomerAnalytics(
        total_customers=total,
        new_customers_today=today,
        new_customers_this_week=week,
        new_customers_this_month=month,
        returning_customers=returning,
    )


@app.get(
    "/api/v1/admin/analytics/products/top-selling",
    response_model=TopProductsAnalytics,
    tags=["Analytics"],
)
async def get_top_products(
    period: str = Query("30d"),
    limit: int = Query(10, le=50),
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    days = {"7d": 7, "30d": 30, "90d": 90, "1y": 365}.get(period, 30)
    since = datetime.now(timezone.utc) - timedelta(days=days)
    rows = db.execute(
        text(
            "SELECT i.product_id, COALESCE(p.name,'Unknown'), SUM(oi.quantity), SUM(oi.unit_price * oi.quantity) "
            "FROM order_items oi "
            "LEFT JOIN inventory i ON i.id = oi.inventory_id "
            "LEFT JOIN products p ON p.id = i.product_id "
            "JOIN orders o ON o.id = oi.order_id "
            "WHERE o.created_at >= :since AND o.status != 'cancelled' "
            "GROUP BY i.product_id, p.name ORDER BY SUM(oi.quantity) DESC LIMIT :lim"
        ),
        {"since": since, "lim": limit},
    ).fetchall()
    products = [
        TopProduct(
            product_id=r[0],
            product_name=r[1],
            total_sold=r[2],
            total_revenue=float(r[3]),
        )
        for r in rows
    ]
    return TopProductsAnalytics(top_products=products, period=period)


# ==================== Inventory (Admin) ====================


@app.get("/api/v1/admin/inventory", tags=["Admin Inventory"])
async def admin_list_inventory(
    limit: int = Query(500, le=1000),
    skip: int = 0,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    """List all inventory items with product names.
    
    🔥 CRITICAL FIX: Uses LEFT JOIN to include products even without inventory records.
    This ensures ALL products are visible in inventory management.
    """
    rows = db.execute(
        text(
            "SELECT i.*, p.name as product_name "
            "FROM products p "
            "LEFT JOIN inventory i ON p.id = i.product_id "
            "ORDER BY p.name, i.size, i.color "
            "LIMIT :limit OFFSET :skip"
        ),
        {"limit": limit, "skip": skip},
    ).fetchall()
    total = db.execute(text("SELECT COUNT(*) FROM products")).scalar() or 0
    return {"items": [dict(r._mapping) for r in rows], "total": total}


@app.get("/api/v1/admin/inventory/low-stock", tags=["Admin Inventory"])
async def admin_low_stock(
    db: Session = Depends(get_db), user: dict = Depends(require_admin)
):
    """Get low stock items. Uses INNER JOIN since we only want items WITH inventory."""
    rows = db.execute(
        text(
            "SELECT i.*, p.name as product_name FROM inventory i JOIN products p ON p.id = i.product_id "
            "WHERE i.quantity <= i.low_stock_threshold ORDER BY i.quantity ASC"
        )
    ).fetchall()
    return {"items": [dict(r._mapping) for r in rows]}


@app.get("/api/v1/admin/inventory/out-of-stock", tags=["Admin Inventory"])
async def admin_out_of_stock(
    db: Session = Depends(get_db), user: dict = Depends(require_admin)
):
    """Get out of stock items. Uses INNER JOIN since we only want items WITH inventory."""
    rows = db.execute(
        text(
            "SELECT i.*, p.name as product_name FROM inventory i JOIN products p ON p.id = i.product_id "
            "WHERE i.quantity = 0 ORDER BY p.name"
        )
    ).fetchall()
    return {"items": [dict(r._mapping) for r in rows]}


@app.get("/api/v1/admin/inventory/movements", tags=["Admin Inventory"])
async def admin_inventory_movements(
    product_id: Optional[int] = None,
    page: int = Query(1, ge=1),
    limit: int = Query(20, le=100),
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    offset = (page - 1) * limit
    if product_id:
        rows = db.execute(
            text("""
            SELECT im.*, p.name as product_name FROM inventory_movements im
            JOIN products p ON p.id = im.product_id
            WHERE im.product_id = :pid
            ORDER BY im.created_at DESC LIMIT :lim OFFSET :off
        """),
            {"pid": product_id, "lim": limit, "off": offset},
        ).fetchall()
        total = (
            db.execute(
                text(
                    "SELECT COUNT(*) FROM inventory_movements WHERE product_id = :pid"
                ),
                {"pid": product_id},
            ).scalar()
            or 0
        )
    else:
        rows = db.execute(
            text("""
            SELECT im.*, p.name as product_name FROM inventory_movements im
            JOIN products p ON p.id = im.product_id
            ORDER BY im.created_at DESC LIMIT :lim OFFSET :off
        """),
            {"lim": limit, "off": offset},
        ).fetchall()
        total = (
            db.execute(text("SELECT COUNT(*) FROM inventory_movements")).scalar() or 0
        )
    return {"movements": [dict(r._mapping) for r in rows], "total": total, "page": page}


@app.post("/api/v1/admin/inventory", tags=["Admin Inventory"])
async def admin_create_inventory(
    data: dict,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    """Create a new inventory record for a product variant."""
    product_id = data.get("product_id")
    sku = data.get("sku")
    if not product_id or not sku:
        raise HTTPException(status_code=400, detail="product_id and sku are required")

    # Check product exists
    product = db.execute(
        text("SELECT id FROM products WHERE id = :pid"), {"pid": product_id}
    ).fetchone()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    # Check for duplicate SKU
    existing = db.execute(
        text("SELECT id FROM inventory WHERE sku = :sku"), {"sku": sku}
    ).fetchone()
    if existing:
        raise HTTPException(
            status_code=409, detail=f"Inventory with SKU '{sku}' already exists"
        )

    result = db.execute(
        text("""
        INSERT INTO inventory (product_id, sku, size, color, quantity, reserved_quantity, low_stock_threshold, updated_at)
        VALUES (:pid, :sku, :size, :color, :qty, 0, :threshold, :now)
        RETURNING id
    """),
        {
            "pid": product_id,
            "sku": sku,
            "size": data.get("size"),
            "color": data.get("color"),
            "qty": data.get("quantity", 0),
            "threshold": data.get("low_stock_threshold", 5),
            "now": datetime.now(timezone.utc),
        },
    )
    inv_id = result.scalar()
    db.commit()
    redis_client.invalidate_pattern("products:*")
    return {"message": "Inventory record created", "inventory_id": inv_id}


@app.patch("/api/v1/admin/inventory/{inventory_id}", tags=["Admin Inventory"])
async def admin_update_inventory(
    inventory_id: int,
    data: dict,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    """Update inventory record fields (quantity, threshold, size, color)."""
    inv = db.execute(
        text("SELECT id FROM inventory WHERE id = :id"), {"id": inventory_id}
    ).fetchone()
    if not inv:
        raise HTTPException(status_code=404, detail="Inventory record not found")

    ALLOWED = {"quantity", "low_stock_threshold", "size", "color"}
    sets, params = (
        ["updated_at = :now"],
        {"id": inventory_id, "now": datetime.now(timezone.utc)},
    )
    for key, val in data.items():
        if key in ALLOWED:
            sets.append(f"{key} = :{key}")
            params[key] = val

    if len(sets) == 1:
        raise HTTPException(status_code=400, detail="No valid fields to update")

    db.execute(text(f"UPDATE inventory SET {', '.join(sets)} WHERE id = :id"), params)
    db.commit()
    redis_client.invalidate_pattern("products:*")
    return {"message": "Inventory updated", "inventory_id": inventory_id}


@app.post("/api/v1/admin/inventory/adjust", tags=["Admin Inventory"])
async def admin_adjust_inventory(
    data: InventoryAdjustRequest,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    """Adjust inventory by SKU. adjustment can be positive (add) or negative (subtract)."""
    sku = data.sku
    adjustment = data.adjustment
    reason = data.reason
    notes = data.notes or ""

    if not sku:
        raise HTTPException(status_code=400, detail="sku is required")

    inv = db.execute(
        text("SELECT id, product_id, quantity FROM inventory WHERE sku = :sku"),
        {"sku": sku},
    ).fetchone()
    if not inv:
        raise HTTPException(status_code=404, detail=f"Inventory SKU '{sku}' not found")

    new_qty = max(0, inv[2] + int(adjustment))
    db.execute(
        text("UPDATE inventory SET quantity = :q, updated_at = :now WHERE id = :id"),
        {"q": new_qty, "now": datetime.now(timezone.utc), "id": inv[0]},
    )
    # Log movement if table exists
    try:
        db.execute(
            text(
                "INSERT INTO inventory_movements (inventory_id, product_id, adjustment, reason, notes, created_at) "
                "VALUES (:iid, :pid, :adj, :reason, :notes, :now)"
            ),
            {
                "iid": inv[0],
                "pid": inv[1],
                "adj": adjustment,
                "reason": reason,
                "notes": notes,
                "now": datetime.now(timezone.utc),
            },
        )
    except Exception:
        pass  # movements table may not exist yet
    db.commit()
    redis_client.invalidate_pattern("products:*")
    return {
        "sku": sku,
        "previous_quantity": inv[2],
        "new_quantity": new_qty,
        "adjustment": adjustment,
    }


# ==================== Staff Dashboard ====================


@app.get(
    "/api/v1/staff/dashboard", response_model=StaffDashboard, tags=["Staff Dashboard"]
)
async def get_staff_dashboard(
    db: Session = Depends(get_db), user: dict = Depends(require_staff)
):
    low = (
        db.execute(
            text(
                "SELECT COUNT(*) FROM inventory WHERE quantity <= low_stock_threshold AND quantity > 0"
            )
        ).scalar()
        or 0
    )
    oos = (
        db.execute(text("SELECT COUNT(*) FROM inventory WHERE quantity = 0")).scalar()
        or 0
    )
    pending = (
        db.execute(
            text("SELECT COUNT(*) FROM orders WHERE status = 'confirmed'")
        ).scalar()
        or 0
    )
    stock_tasks = (
        db.execute(
            text(
                "SELECT COUNT(*) FROM staff_tasks WHERE task_type = 'stock_update' AND status = 'pending'"
            )
        ).scalar()
        or 0
    )
    order_tasks = (
        db.execute(
            text(
                "SELECT COUNT(*) FROM staff_tasks WHERE task_type = 'order_processing' AND status = 'pending'"
            )
        ).scalar()
        or 0
    )
    return StaffDashboard(
        inventory_alerts=InventoryAlert(low_stock=low, out_of_stock=oos),
        pending_orders=pending,
        today_tasks={"stock_updates": stock_tasks, "order_processing": order_tasks},
        quick_actions=[
            "add_stock",
            "process_orders",
            "update_inventory",
            "view_low_stock",
        ],
    )


# ==================== Staff Inventory ====================


@app.get("/api/v1/staff/inventory/low-stock", tags=["Staff Inventory"])
async def staff_low_stock(
    db: Session = Depends(get_db), user: dict = Depends(require_staff)
):
    rows = db.execute(
        text(
            "SELECT i.*, p.name as product_name FROM inventory i JOIN products p ON p.id = i.product_id "
            "WHERE i.quantity <= i.low_stock_threshold AND i.quantity > 0 ORDER BY i.quantity ASC"
        )
    ).fetchall()
    return {"items": [dict(r._mapping) for r in rows]}


@app.get("/api/v1/staff/inventory/out-of-stock", tags=["Staff Inventory"])
async def staff_out_of_stock(
    db: Session = Depends(get_db), user: dict = Depends(require_staff)
):
    rows = db.execute(
        text(
            "SELECT i.*, p.name as product_name FROM inventory i JOIN products p ON p.id = i.product_id WHERE i.quantity = 0"
        )
    ).fetchall()
    return {"items": [dict(r._mapping) for r in rows]}


@app.post("/api/v1/staff/inventory/add-stock", tags=["Staff Inventory"])
async def add_stock(
    data: AddStockRequest,
    db: Session = Depends(get_db),
    user: dict = Depends(require_staff),
):
    existing = db.execute(
        text(
            "SELECT id, quantity FROM inventory WHERE product_id = :pid AND sku = :sku"
        ),
        {"pid": data.product_id, "sku": data.sku},
    ).fetchone()
    if existing:
        new_qty = existing[1] + data.quantity
        db.execute(
            text(
                "UPDATE inventory SET quantity = :q, updated_at = :now WHERE id = :id"
            ),
            {"q": new_qty, "now": datetime.now(timezone.utc), "id": existing[0]},
        )
        inv_id = existing[0]
    else:
        result = db.execute(
            text(
                "INSERT INTO inventory (product_id, sku, quantity, cost_price, created_at, updated_at) "
                "VALUES (:pid, :sku, :qty, :cp, :now, :now) RETURNING id"
            ),
            {
                "pid": data.product_id,
                "sku": data.sku,
                "qty": data.quantity,
                "cp": data.cost_price,
                "now": datetime.now(timezone.utc),
            },
        )
        inv_id = result.scalar()
    # Record movement
    db.execute(
        text(
            "INSERT INTO inventory_movements (inventory_id, product_id, adjustment, reason, notes, supplier, cost_price, performed_by) "
            "VALUES (:iid, :pid, :adj, 'restock', :notes, :supplier, :cp, :by)"
        ),
        {
            "iid": inv_id,
            "pid": data.product_id,
            "adj": data.quantity,
            "notes": data.notes,
            "supplier": data.supplier,
            "cp": data.cost_price,
            "by": user.get("user_id"),
        },
    )
    db.commit()
    redis_client.invalidate_pattern("admin:dashboard:*")
    redis_client.invalidate_pattern("products:*")
    return {
        "message": "Stock added",
        "inventory_id": inv_id,
        "quantity_added": data.quantity,
    }


@app.post("/api/v1/staff/inventory/adjust-stock", tags=["Staff Inventory"])
async def adjust_stock(
    data: AdjustStockRequest,
    db: Session = Depends(get_db),
    user: dict = Depends(require_staff),
):
    inv = db.execute(
        text("SELECT id, product_id, quantity FROM inventory WHERE id = :id"),
        {"id": data.inventory_id},
    ).fetchone()
    if not inv:
        raise HTTPException(status_code=404, detail="Inventory item not found")
    new_qty = max(0, inv[2] + data.adjustment)
    db.execute(
        text("UPDATE inventory SET quantity = :q, updated_at = :now WHERE id = :id"),
        {"q": new_qty, "now": datetime.now(timezone.utc), "id": data.inventory_id},
    )
    db.execute(
        text(
            "INSERT INTO inventory_movements (inventory_id, product_id, adjustment, reason, notes, performed_by) "
            "VALUES (:iid, :pid, :adj, :reason, :notes, :by)"
        ),
        {
            "iid": data.inventory_id,
            "pid": inv[1],
            "adj": data.adjustment,
            "reason": data.reason,
            "notes": data.notes,
            "by": user.get("user_id"),
        },
    )
    db.commit()
    redis_client.invalidate_pattern("admin:dashboard:*")
    redis_client.invalidate_pattern("products:*")
    return {
        "message": "Stock adjusted",
        "inventory_id": data.inventory_id,
        "new_quantity": new_qty,
    }


@app.get("/api/v1/staff/inventory/movements", tags=["Staff Inventory"])
async def get_movements(
    product_id: Optional[int] = None,
    page: int = Query(1, ge=1),
    limit: int = Query(20, le=100),
    db: Session = Depends(get_db),
    user: dict = Depends(require_staff),
):
    offset = (page - 1) * limit
    if product_id:
        rows = db.execute(
            text("""
            SELECT im.*, p.name as product_name FROM inventory_movements im
            JOIN products p ON p.id = im.product_id
            WHERE im.product_id = :pid
            ORDER BY im.created_at DESC LIMIT :lim OFFSET :off
        """),
            {"pid": product_id, "lim": limit, "off": offset},
        ).fetchall()
    else:
        rows = db.execute(
            text("""
            SELECT im.*, p.name as product_name FROM inventory_movements im
            JOIN products p ON p.id = im.product_id
            ORDER BY im.created_at DESC LIMIT :lim OFFSET :off
        """),
            {"lim": limit, "off": offset},
        ).fetchall()
    return {"movements": [dict(r._mapping) for r in rows]}


@app.post("/api/v1/staff/inventory/bulk-update", tags=["Staff Inventory"])
async def bulk_update_inventory(
    data: BulkInventoryUpdate,
    db: Session = Depends(get_db),
    user: dict = Depends(require_staff),
):
    """Bulk update inventory items. Only 'quantity' and 'low_stock_threshold' columns are allowed."""
    MAX_BATCH = 100
    if len(data.updates) > MAX_BATCH:
        raise HTTPException(
            status_code=400,
            detail=f"Batch size {len(data.updates)} exceeds maximum of {MAX_BATCH} items per request.",
        )

    updated = 0
    ALLOWED_COLUMNS = {"quantity", "low_stock_threshold"}

    for item in data.updates:
        inv_id = item.get("inventory_id")
        if not inv_id:
            continue

        # Validate no extra columns
        item_keys = set(item.keys()) - {"inventory_id"}
        invalid_keys = item_keys - ALLOWED_COLUMNS
        if invalid_keys:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid columns: {invalid_keys}. Only {ALLOWED_COLUMNS} are allowed.",
            )

        # Validate value ranges
        if "quantity" in item:
            qty = item["quantity"]
            if not isinstance(qty, int) or qty < 0 or qty > 100_000:
                raise HTTPException(
                    status_code=400,
                    detail=f"quantity must be an integer 0-100000, got {qty!r}",
                )
        if "low_stock_threshold" in item:
            thresh = item["low_stock_threshold"]
            if not isinstance(thresh, int) or thresh < 0 or thresh > 10_000:
                raise HTTPException(
                    status_code=400,
                    detail=f"low_stock_threshold must be an integer 0-10000, got {thresh!r}",
                )

        # Build parameterized update (whitelisted column names only)
        if "quantity" in item and "low_stock_threshold" in item:
            db.execute(
                text(
                    "UPDATE inventory SET quantity = :qty, low_stock_threshold = :thresh, updated_at = :now WHERE id = :id"
                ),
                {
                    "qty": item["quantity"],
                    "thresh": item["low_stock_threshold"],
                    "now": datetime.now(timezone.utc),
                    "id": inv_id,
                },
            )
            updated += 1
        elif "quantity" in item:
            db.execute(
                text(
                    "UPDATE inventory SET quantity = :qty, updated_at = :now WHERE id = :id"
                ),
                {
                    "qty": item["quantity"],
                    "now": datetime.now(timezone.utc),
                    "id": inv_id,
                },
            )
            updated += 1
        elif "low_stock_threshold" in item:
            db.execute(
                text(
                    "UPDATE inventory SET low_stock_threshold = :thresh, updated_at = :now WHERE id = :id"
                ),
                {
                    "thresh": item["low_stock_threshold"],
                    "now": datetime.now(timezone.utc),
                    "id": inv_id,
                },
            )
            updated += 1

    db.commit()
    redis_client.invalidate_pattern("admin:dashboard:*")
    redis_client.invalidate_pattern("products:*")
    return {"message": f"Updated {updated} inventory items"}


# ==================== Product Variants (Staff) ====================


@app.get("/api/v1/staff/products/{product_id}/variants", tags=["Staff Variants"])
async def get_variants(
    product_id: str, db: Session = Depends(get_db), user: dict = Depends(require_staff)
):
    # Resolve slug or ID to actual product ID
    resolved = db.execute(
        text("SELECT id FROM products WHERE id::text = :pid OR slug = :pid"),
        {"pid": product_id},
    ).fetchone()
    if not resolved:
        raise HTTPException(status_code=404, detail="Product not found")
    pid = resolved[0]
    rows = db.execute(
        text(
            "SELECT * FROM inventory WHERE product_id = :pid ORDER BY size, color, sku"
        ),
        {"pid": pid},
    ).fetchall()
    return {"variants": [dict(r._mapping) for r in rows]}


@app.post("/api/v1/staff/products/{product_id}/variants", tags=["Staff Variants"])
async def create_variant(
    product_id: str,
    data: VariantCreate,
    db: Session = Depends(get_db),
    user: dict = Depends(require_staff),
):
    result = db.execute(
        text(
            "INSERT INTO inventory (product_id, sku, size, color, color_hex, quantity, reserved_quantity, created_at, updated_at) "
            "VALUES (:pid, :sku, :size, :color, :color_hex, :qty, 0, :now, :now) RETURNING id"
        ),
        {
            "pid": product_id,
            "sku": data.sku,
            "size": data.size,
            "color": data.color,
            "color_hex": data.color_hex,
            "qty": data.quantity,
            "now": datetime.now(timezone.utc),
        },
    )
    db.commit()
    return {"message": "Variant created", "variant_id": result.scalar()}


@app.put("/api/v1/staff/variants/{variant_id}", tags=["Staff Variants"])
async def update_variant(
    variant_id: int,
    data: VariantUpdate,
    db: Session = Depends(get_db),
    user: dict = Depends(require_staff),
):
    sets, params = (
        ["updated_at = :now"],
        {"id": variant_id, "now": datetime.now(timezone.utc)},
    )
    if data.size is not None:
        sets.append("size = :size")
        params["size"] = data.size
    if data.color is not None:
        sets.append("color = :color")
        params["color"] = data.color
    if data.color_hex is not None:
        sets.append("color_hex = :color_hex")
        params["color_hex"] = data.color_hex
    if data.quantity is not None:
        sets.append("quantity = :qty")
        params["qty"] = data.quantity
    db.execute(text(f"UPDATE inventory SET {', '.join(sets)} WHERE id = :id"), params)
    db.commit()
    redis_client.invalidate_pattern("products:*")
    return {"message": "Variant updated"}


@app.delete("/api/v1/staff/variants/{variant_id}", tags=["Staff Variants"])
async def delete_variant(
    variant_id: int, db: Session = Depends(get_db), user: dict = Depends(require_staff)
):
    db.execute(text("DELETE FROM inventory WHERE id = :id"), {"id": variant_id})
    db.commit()
    redis_client.invalidate_pattern("products:*")
    return {"message": "Variant deleted"}


# ==================== Staff Order Processing ====================


@app.get("/api/v1/staff/orders/confirmed", tags=["Staff Orders"])
async def staff_confirmed_orders(
    db: Session = Depends(get_db), user: dict = Depends(require_staff)
):
    """Get all confirmed orders awaiting shipment."""
    rows = db.execute(
        text(
            "SELECT o.id, o.user_id, o.total_amount, o.status, o.shipping_address, o.created_at, "
            "u.email as customer_email, u.username as customer_name "
            "FROM orders o LEFT JOIN users u ON u.id = o.user_id "
            "WHERE o.status = 'confirmed' ORDER BY o.created_at ASC"
        )
    ).fetchall()
    return {"orders": [dict(r._mapping) for r in rows]}


@app.put("/api/v1/staff/orders/{order_id}/ship", tags=["Staff Orders"])
async def ship_order(
    order_id: int,
    data: OrderShipRequest,
    db: Session = Depends(get_db),
    user: dict = Depends(require_staff),
):
    """Ship a confirmed order. POD number is required."""
    if not data.tracking_number or not data.tracking_number.strip():
        raise HTTPException(
            status_code=400, detail="POD number is required when shipping an order"
        )
    order = db.execute(
        text("SELECT id, status FROM orders WHERE id = :id"), {"id": order_id}
    ).fetchone()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if order[1] != "confirmed":
        raise HTTPException(
            status_code=400,
            detail=f"Only confirmed orders can be shipped (current status: {order[1]})",
        )
    now = datetime.now(timezone.utc)
    db.execute(
        text(
            "UPDATE orders SET status = 'shipped', shipped_at = :now, updated_at = :now, tracking_number = :tn WHERE id = :id"
        ),
        {"now": now, "id": order_id, "tn": data.tracking_number.strip()},
    )
    notes = data.notes or f"Order shipped — POD number: {data.tracking_number.strip()}"
    db.execute(
        text(
            "INSERT INTO order_tracking (order_id, status, notes, updated_by, created_at) VALUES (:oid, 'shipped', :notes, :by, :now)"
        ),
        {"oid": order_id, "notes": notes, "by": user.get("user_id"), "now": now},
    )
    db.commit()
    redis_client.delete_cache(f"order:{order_id}")
    return {
        "message": "Order shipped successfully",
        "order_id": order_id,
        "pod_number": data.tracking_number.strip(),
    }


@app.put("/api/v1/staff/orders/{order_id}/deliver", tags=["Staff Orders"])
async def deliver_order(
    order_id: int, db: Session = Depends(get_db), user: dict = Depends(require_staff)
):
    """Mark a shipped order as delivered."""
    order = db.execute(
        text("SELECT id, status FROM orders WHERE id = :id"), {"id": order_id}
    ).fetchone()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if order[1] != "shipped":
        raise HTTPException(
            status_code=400,
            detail=f"Only shipped orders can be marked delivered (current status: {order[1]})",
        )
    now = datetime.now(timezone.utc)
    db.execute(
        text(
            "UPDATE orders SET status = 'delivered', delivered_at = :now, updated_at = :now WHERE id = :id"
        ),
        {"now": now, "id": order_id},
    )
    db.execute(
        text(
            "INSERT INTO order_tracking (order_id, status, notes, updated_by, created_at) VALUES (:oid, 'delivered', 'Order delivered to customer', :by, :now)"
        ),
        {"oid": order_id, "by": user.get("user_id"), "now": now},
    )
    db.commit()
    redis_client.delete_cache(f"order:{order_id}")
    return {"message": "Order marked as delivered", "order_id": order_id}


# ==================== Staff Reservations ====================


@app.get("/api/v1/staff/reservations/pending", tags=["Staff Reservations"])
async def pending_reservations(
    db: Session = Depends(get_db), user: dict = Depends(require_staff)
):
    rows = db.execute(
        text(
            "SELECT i.id, i.product_id, p.name, i.sku, i.reserved_quantity FROM inventory i "
            "JOIN products p ON p.id = i.product_id WHERE i.reserved_quantity > 0"
        )
    ).fetchall()
    return {"reservations": [dict(r._mapping) for r in rows]}


@app.post("/api/v1/staff/reservations/release", tags=["Staff Reservations"])
async def release_reservation(
    data: ReservationRelease,
    db: Session = Depends(get_db),
    user: dict = Depends(require_staff),
):
    for item in data.items:
        inv_id = item.get("inventory_id")
        qty = item.get("quantity", 0)
        db.execute(
            text(
                "UPDATE inventory SET reserved_quantity = GREATEST(0, reserved_quantity - :qty), quantity = quantity + :qty WHERE id = :id"
            ),
            {"qty": qty, "id": inv_id},
        )
    db.commit()
    return {"message": "Reservations released"}


@app.post("/api/v1/staff/reservations/confirm", tags=["Staff Reservations"])
async def confirm_reservation(
    data: ReservationRelease,
    db: Session = Depends(get_db),
    user: dict = Depends(require_staff),
):
    for item in data.items:
        inv_id = item.get("inventory_id")
        qty = item.get("quantity", 0)
        db.execute(
            text(
                "UPDATE inventory SET reserved_quantity = GREATEST(0, reserved_quantity - :qty) WHERE id = :id"
            ),
            {"qty": qty, "id": inv_id},
        )
    db.commit()
    return {"message": "Reservations confirmed (stock deducted)"}


# ==================== Staff Reports ====================


@app.get("/api/v1/staff/reports/inventory/summary", tags=["Staff Reports"])
async def inventory_summary(
    db: Session = Depends(get_db), user: dict = Depends(require_staff)
):
    total_products = (
        db.execute(
            text("SELECT COUNT(*) FROM products WHERE is_active = true")
        ).scalar()
        or 0
    )
    total_variants = (
        db.execute(text("SELECT COUNT(*) FROM product_variants")).scalar() or 0
    )
    total_stock = (
        db.execute(text("SELECT COALESCE(SUM(quantity), 0) FROM inventory")).scalar()
        or 0
    )
    low = (
        db.execute(
            text(
                "SELECT COUNT(*) FROM inventory WHERE quantity <= low_stock_threshold AND quantity > 0"
            )
        ).scalar()
        or 0
    )
    oos = (
        db.execute(text("SELECT COUNT(*) FROM inventory WHERE quantity = 0")).scalar()
        or 0
    )
    top = db.execute(
        text(
            "SELECT p.name, COALESCE(SUM(ABS(im.adjustment)), 0) as movements FROM inventory_movements im "
            "JOIN products p ON p.id = im.product_id GROUP BY p.name ORDER BY movements DESC LIMIT 5"
        )
    ).fetchall()
    return {
        "total_products": total_products,
        "total_variants": total_variants,
        "total_stock": total_stock,
        "low_stock_items": low,
        "out_of_stock_items": oos,
        "top_moving_products": [{"product_name": r[0], "movements": r[1]} for r in top],
    }


@app.get("/api/v1/staff/reports/orders/processed", tags=["Staff Reports"])
async def processed_orders_report(
    db: Session = Depends(get_db), user: dict = Depends(require_staff)
):
    today = datetime.now(timezone.utc).date()
    shipped_today = (
        db.execute(
            text(
                "SELECT COUNT(*) FROM orders WHERE status = 'shipped' AND DATE(shipped_at) = :d"
            ),
            {"d": today},
        ).scalar()
        or 0
    )
    delivered_today = (
        db.execute(
            text(
                "SELECT COUNT(*) FROM orders WHERE status = 'delivered' AND DATE(delivered_at) = :d"
            ),
            {"d": today},
        ).scalar()
        or 0
    )
    confirmed_pending = (
        db.execute(
            text("SELECT COUNT(*) FROM orders WHERE status = 'confirmed'")
        ).scalar()
        or 0
    )
    return {
        "today_shipped": shipped_today,
        "today_delivered": delivered_today,
        "confirmed_awaiting_shipment": confirmed_pending,
    }


# ==================== Staff Tasks ====================


@app.get("/api/v1/staff/tasks", tags=["Staff Tasks"])
async def get_tasks(db: Session = Depends(get_db), user: dict = Depends(require_staff)):
    rows = db.execute(
        text(
            "SELECT * FROM staff_tasks WHERE status != 'completed' ORDER BY CASE priority WHEN 'high' THEN 1 WHEN 'medium' THEN 2 ELSE 3 END, due_time ASC NULLS LAST"
        )
    ).fetchall()
    return {"pending_tasks": [dict(r._mapping) for r in rows]}


@app.post("/api/v1/staff/tasks/{task_id}/complete", tags=["Staff Tasks"])
async def complete_task(
    task_id: int,
    data: TaskComplete,
    db: Session = Depends(get_db),
    user: dict = Depends(require_staff),
):
    db.execute(
        text(
            "UPDATE staff_tasks SET status = 'completed', completed_at = :now, updated_at = :now WHERE id = :id"
        ),
        {"now": datetime.now(timezone.utc), "id": task_id},
    )
    db.commit()
    return {"message": "Task completed"}


# ==================== Staff Notifications ====================


@app.get("/api/v1/staff/notifications", tags=["Staff Notifications"])
async def get_notifications(
    db: Session = Depends(get_db), user: dict = Depends(require_staff)
):
    uid = user.get("user_id")
    rows = db.execute(
        text(
            "SELECT * FROM staff_notifications WHERE (user_id = :uid OR user_id IS NULL) AND is_read = false ORDER BY created_at DESC LIMIT 50"
        ),
        {"uid": uid},
    ).fetchall()
    return {"alerts": [dict(r._mapping) for r in rows]}


@app.put("/api/v1/staff/notifications/{notif_id}/read", tags=["Staff Notifications"])
async def mark_notification_read(
    notif_id: int, db: Session = Depends(get_db), user: dict = Depends(require_staff)
):
    db.execute(
        text("UPDATE staff_notifications SET is_read = true WHERE id = :id"),
        {"id": notif_id},
    )
    db.commit()
    return {"message": "Notification marked as read"}


# ==================== Chat ====================


@app.get("/api/v1/admin/chat/rooms", tags=["Chat"])
async def list_chat_rooms(
    status: Optional[str] = None,
    db: Session = Depends(get_db),
    user: dict = Depends(require_staff),
):
    where, params = "", {}
    if status:
        where = "WHERE status = :status"
        params["status"] = status
    
    # Add unread count subquery
    rows = db.execute(
        text(f"""
            SELECT cr.*, 
                   COALESCE(
                       (SELECT COUNT(*) FROM chat_messages cm 
                        WHERE cm.room_id = cr.id 
                        AND cm.is_read = false 
                        AND cm.sender_type IN ('customer', 'user')),
                       0
                   ) as unread
            FROM chat_rooms cr 
            {where} 
            ORDER BY updated_at DESC 
            LIMIT 50
        """),
        params,
    ).fetchall()
    return {"rooms": [dict(r._mapping) for r in rows]}


@app.get("/api/v1/admin/chat/rooms/{room_id}/messages", tags=["Chat"])
async def get_chat_messages(
    room_id: int, db: Session = Depends(get_db), user: dict = Depends(require_staff)
):
    msgs = db.execute(
        text(
            "SELECT * FROM chat_messages WHERE room_id = :rid ORDER BY created_at ASC"
        ),
        {"rid": room_id},
    ).fetchall()
    # Mark as read
    db.execute(
        text(
            "UPDATE chat_messages SET is_read = true WHERE room_id = :rid AND sender_type != 'staff' AND sender_type != 'admin'"
        ),
        {"rid": room_id},
    )
    db.commit()
    return {"messages": [dict(m._mapping) for m in msgs]}


@app.post("/api/v1/admin/chat/rooms/{room_id}/messages", tags=["Chat"])
async def send_chat_message(
    room_id: int,
    data: ChatMessageCreate,
    db: Session = Depends(get_db),
    user: dict = Depends(require_staff),
):
    db.execute(
        text(
            "INSERT INTO chat_messages (room_id, sender_id, sender_type, message) VALUES (:rid, :sid, :st, :msg)"
        ),
        {
            "rid": room_id,
            "sid": user.get("user_id"),
            "st": data.sender_type,
            "msg": data.message,
        },
    )
    db.execute(
        text("UPDATE chat_rooms SET updated_at = :now WHERE id = :rid"),
        {"now": datetime.now(timezone.utc), "rid": room_id},
    )
    db.commit()
    redis_client.publish(
        "chat:notifications", {"room_id": room_id, "message": data.message}
    )
    return {"message": "Message sent"}


@app.put("/api/v1/admin/chat/rooms/{room_id}/assign", tags=["Chat"])
async def assign_chat_room(
    room_id: int, db: Session = Depends(get_db), user: dict = Depends(require_staff)
):
    db.execute(
        text(
            "UPDATE chat_rooms SET assigned_to = :uid, status = 'assigned', updated_at = :now WHERE id = :rid"
        ),
        {"uid": user.get("user_id"), "rid": room_id, "now": datetime.now(timezone.utc)},
    )
    db.commit()
    return {"message": "Chat room assigned to you"}


@app.put("/api/v1/admin/chat/rooms/{room_id}/close", tags=["Chat"])
async def close_chat_room(
    room_id: int, db: Session = Depends(get_db), user: dict = Depends(require_staff)
):
    db.execute(
        text(
            "UPDATE chat_rooms SET status = 'closed', closed_at = :now, updated_at = :now WHERE id = :rid"
        ),
        {"rid": room_id, "now": datetime.now(timezone.utc)},
    )
    db.commit()
    return {"message": "Chat room closed"}


# ==================== WebSocket Chat ====================

import asyncio
import uuid

# Unique worker ID to prevent re-broadcasting our own messages
_WORKER_ID = str(uuid.uuid4())[:8]


class ChatConnectionManager:
    """
    Manages WebSocket connections per chat room with Redis Pub/Sub
    for cross-worker broadcast.

    Architecture:
    - Each worker tracks local WebSocket connections per room
    - When a message arrives via WebSocket, it is:
        1. Saved to Postgres asynchronously
        2. Published to Redis channel `chat:room:{room_id}`
        3. Broadcast to local connections immediately
    - A background subscriber task listens for messages from OTHER workers
      and broadcasts them to local connections
    """

    def __init__(self):
        self.active: dict[int, list[WebSocket]] = {}
        self._subscribers: dict[int, asyncio.Task] = {}

    async def connect(self, room_id: int, ws: WebSocket):
        await ws.accept()
        self.active.setdefault(room_id, []).append(ws)

        # Start Redis subscriber for this room if not already running
        if room_id not in self._subscribers:
            self._subscribers[room_id] = asyncio.create_task(
                self._redis_subscriber(room_id)
            )

    def disconnect(self, room_id: int, ws: WebSocket):
        conns = self.active.get(room_id, [])
        if ws in conns:
            conns.remove(ws)

        # If no more connections for this room, cancel subscriber
        if not conns and room_id in self._subscribers:
            self._subscribers[room_id].cancel()
            del self._subscribers[room_id]

    async def broadcast_local(self, room_id: int, message: dict):
        """Broadcast to all local WebSocket connections in this room."""
        dead = []
        for ws in self.active.get(room_id, []):
            try:
                await ws.send_json(message)
            except Exception:
                dead.append(ws)
        # Cleanup dead connections
        for ws in dead:
            self.disconnect(room_id, ws)

    async def publish_to_redis(self, room_id: int, message: dict):
        """Publish message to Redis for cross-worker broadcast."""
        try:
            payload = json.dumps({**message, "_origin": _WORKER_ID})
            redis_client.client.publish(f"chat:room:{room_id}", payload)
        except Exception as e:
            logger.warning(f"Redis publish error for room {room_id}: {e}")

    async def _redis_subscriber(self, room_id: int):
        """Background task: subscribe to Redis and relay messages from other workers."""
        pubsub = redis_client.client.pubsub()
        channel = f"chat:room:{room_id}"
        pubsub.subscribe(channel)

        try:
            while True:
                message = pubsub.get_message(
                    ignore_subscribe_messages=True, timeout=1.0
                )
                if message and message["type"] == "message":
                    data = message["data"]
                    if isinstance(data, bytes):
                        data = data.decode("utf-8")
                    parsed = json.loads(data)

                    # Only broadcast if the message came from a DIFFERENT worker
                    if parsed.get("_origin") != _WORKER_ID:
                        parsed.pop("_origin", None)
                        await self.broadcast_local(room_id, parsed)

                await asyncio.sleep(0)  # Yield control
        except asyncio.CancelledError:
            pass
        except Exception as e:
            logger.error(f"Redis subscriber error for room {room_id}: {e}")
        finally:
            pubsub.unsubscribe(channel)
            pubsub.close()


chat_manager = ChatConnectionManager()


async def _save_chat_message_async(
    room_id: int, user_id: int, sender_type: str, msg_text: str
):
    """Save chat message to DB in a background task."""
    from database.database import SessionLocal

    db_session = SessionLocal()
    try:
        db_session.execute(
            text(
                "INSERT INTO chat_messages (room_id, sender_id, sender_type, message) "
                "VALUES (:rid, :sid, :st, :msg)"
            ),
            {"rid": room_id, "sid": user_id, "st": sender_type, "msg": msg_text},
        )
        db_session.execute(
            text("UPDATE chat_rooms SET updated_at = :now WHERE id = :rid"),
            {"now": datetime.now(timezone.utc), "rid": room_id},
        )
        db_session.commit()
    except Exception as e:
        db_session.rollback()
        logger.error(f"WS chat DB error: {e}")
    finally:
        db_session.close()


async def _websocket_staff_chat(
    ws: WebSocket, room_id: int, token: str | None, db: Session
):
    """
    Staff dashboard WebSocket (nginx: /api/v1/admin/chat/ws/* → admin service).

    Also exposed at /api/v1/chat/ws/{room_id} for direct service access / tests.
    Customer storefront uses Commerce's /api/v1/chat/ws/* (different upstream).
    """
    import jwt as pyjwt

    cookie_token = ws.cookies.get("access_token")
    resolved_token = cookie_token or token

    if not resolved_token:
        await ws.close(code=4001, reason="Missing token")
        return
    try:
        payload = pyjwt.decode(
            resolved_token, settings.SECRET_KEY, algorithms=[settings.ALGORITHM]
        )
        user_id = payload.get("user_id")
        role = payload.get("role", "customer")
    except Exception:
        await ws.close(code=4003, reason="Invalid token")
        return

    sender_type = "staff" if role in ("admin", "staff", "superadmin") else "customer"

    room = db.execute(
        text("SELECT id, customer_id, status FROM chat_rooms WHERE id = :rid"),
        {"rid": room_id},
    ).fetchone()
    if not room:
        await ws.close(code=4004, reason="Room not found")
        return

    if sender_type == "customer" and room[1] != user_id:
        await ws.close(code=4003, reason="Forbidden")
        return

    await chat_manager.connect(room_id, ws)
    try:
        while True:
            data = await ws.receive_json()
            msg_text = data.get("message", "").strip()
            if not msg_text:
                continue

            now = datetime.now(timezone.utc)

            message_payload = {
                "room_id": room_id,
                "sender_id": user_id,
                "sender_type": sender_type,
                "message": msg_text,
                "created_at": now.isoformat(),
            }

            await chat_manager.broadcast_local(room_id, message_payload)
            await chat_manager.publish_to_redis(room_id, message_payload)

            asyncio.create_task(
                _save_chat_message_async(room_id, user_id, sender_type, msg_text)
            )

    except WebSocketDisconnect:
        chat_manager.disconnect(room_id, ws)
    except Exception as e:
        logger.error(f"WebSocket error room {room_id}: {e}")
        chat_manager.disconnect(room_id, ws)


@app.websocket("/api/v1/chat/ws/{room_id}")
async def websocket_chat(ws: WebSocket, room_id: int, token: str = Query(None), db: Session = Depends(get_db)):
    """See _websocket_staff_chat."""
    await _websocket_staff_chat(ws, room_id, token, db)


@app.websocket("/api/v1/admin/chat/ws/{room_id}")
async def websocket_chat_admin_dashboard(
    ws: WebSocket, room_id: int, token: str = Query(None), db: Session = Depends(get_db)
):
    """Path expected by frontend (admin/chat) and nginx location ^/api/v1/admin/chat/ws/."""
    await _websocket_staff_chat(ws, room_id, token, db)


# ==================== Site Config ====================


@app.get("/api/v1/admin/site/config", tags=["Site Config"])
async def get_site_config(
    db: Session = Depends(get_db), user: dict = Depends(require_admin)
):
    rows = db.execute(
        text("SELECT key, value, description, updated_at FROM site_config")
    ).fetchall()
    return {
        "config": {
            r[0]: {"value": r[1], "description": r[2], "updated_at": str(r[3])}
            for r in rows
        }
    }


@app.put("/api/v1/admin/site/config", tags=["Site Config"])
async def update_site_config(
    data: dict, db: Session = Depends(get_db), user: dict = Depends(require_admin)
):
    for key, value in data.items():
        # Handle boolean/numeric to string conversion for TEXT column
        val_str = str(value).lower() if isinstance(value, bool) else str(value)
        db.execute(
            text(
                "INSERT INTO site_config (key, value) VALUES (:key, :value) "
                "ON CONFLICT (key) DO UPDATE SET value = :value, updated_at = :now"
            ),
            {"key": key, "value": val_str, "now": datetime.now(timezone.utc)},
        )
    db.commit()
    redis_client.invalidate_pattern("public:site:*")
    redis_client.invalidate_pattern("public:landing:*")
    return {"message": "Site config updated"}


# ==================== Admin Orders (Direct DB) ====================


@app.get("/api/v1/admin/orders", tags=["Admin"])
async def list_all_orders(
    status: Optional[str] = None,
    search: Optional[str] = None,
    user_id: Optional[int] = None,
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_staff),
):
    """List all orders (admin/staff) - direct DB query instead of httpx proxy."""
    where_clauses = []
    params = {"lim": limit, "off": skip}
    if status:
        where_clauses.append("o.status = :status")
        params["status"] = status
    if user_id is not None:
        where_clauses.append("o.user_id = :user_id")
        params["user_id"] = user_id
    if search:
        where_clauses.append(
            "(CAST(o.id AS TEXT) ILIKE :search OR u.email ILIKE :search OR u.username ILIKE :search OR COALESCE(o.tracking_number, '') ILIKE :search)"
        )
        params["search"] = f"%{search.strip()}%"

    where_clause = f"WHERE {' AND '.join(where_clauses)}" if where_clauses else ""

    rows = db.execute(
        text(f"""
        SELECT o.id, o.user_id, o.subtotal, o.discount_applied, o.shipping_cost,
               o.total_amount, o.payment_method, o.status, o.tracking_number,
               o.order_notes, o.created_at, o.updated_at,
               u.email as customer_email, COALESCE(up.full_name, u.username) as customer_name,
               COALESCE(up.phone, '') as customer_phone,
               o.invoice_number, o.shipping_address
        FROM orders o
        LEFT JOIN users u ON u.id = o.user_id
        LEFT JOIN user_profiles up ON up.user_id = o.user_id
        {where_clause}
        ORDER BY o.created_at DESC
        LIMIT :lim OFFSET :off
    """),
        params,
    ).fetchall()

    count_params = {k: v for k, v in params.items() if k not in ("lim", "off")}
    total = (
        db.execute(
            text(f"SELECT COUNT(*) FROM orders o LEFT JOIN users u ON u.id = o.user_id {where_clause}"),
            count_params,
        ).scalar()
        or 0
    )

    # Collect order IDs for batch fetching order items
    order_ids = [r[0] for r in rows]

    # Batch fetch order items for all orders in a single query (avoids N+1)
    items_by_order = {}
    if order_ids:
        placeholders = ", ".join(f":oid_{i}" for i in range(len(order_ids)))
        items_params = {f"oid_{i}": oid for i, oid in enumerate(order_ids)}
        items_rows = db.execute(
            text(f"""
                SELECT oi.order_id, oi.id as item_id, oi.product_id, oi.product_name,
                       oi.size, oi.color, oi.quantity, oi.unit_price, oi.price,
                       p.name as product_name_from_catalog,
                       pi.image_url as image_url
                FROM order_items oi
                LEFT JOIN products p ON p.id = oi.product_id
                LEFT JOIN product_images pi ON p.id = pi.product_id AND pi.is_primary = true
                WHERE oi.order_id IN ({placeholders})
                ORDER BY oi.order_id, oi.id
            """),
            items_params,
        ).fetchall()
        for item_row in items_rows:
            oid = item_row[0]
            if oid not in items_by_order:
                items_by_order[oid] = []
            items_by_order[oid].append({
                "id": item_row[1],
                "product_id": item_row[2],
                "product_name": item_row[3] or item_row[9] or "Unknown Product",
                "size": item_row[4],
                "color": item_row[5],
                "quantity": item_row[6],
                "unit_price": float(item_row[7] or 0),
                "total_price": float(item_row[8] or 0),
                "image_url": item_row[10],
            })

    orders = [
        {
            "id": r[0],
            "user_id": r[1],
            "subtotal": float(r[2] or 0),
            "discount_applied": float(r[3] or 0),
            "shipping_cost": float(r[4] or 0),
            "total_amount": float(r[5] or 0),
            "payment_method": r[6],
            "status": r[7],
            "tracking_number": r[8],
            "order_notes": r[9],
            "created_at": str(r[10]),
            "updated_at": str(r[11]),
            "customer_email": r[12],
            "customer_name": r[13],
            "customer_phone": r[14],
            "invoice_number": r[15],
            "shipping_address": r[16],
            "order_number": f"ORD-{r[0]:06d}",
            "items": items_by_order.get(r[0], []),
        }
        for r in rows
    ]

    return {"orders": orders, "total": total, "skip": skip, "limit": limit}


@app.get("/api/v1/admin/orders/{order_id}", tags=["Admin"])
async def get_order(
    order_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_staff),
):
    """Get order details - direct DB query instead of httpx proxy."""
    order = db.execute(
        text("""
        SELECT o.*, u.email as customer_email, u.username as customer_name, up.phone as customer_phone
        FROM orders o
        LEFT JOIN users u ON u.id = o.user_id
        LEFT JOIN user_profiles up ON up.user_id = u.id
        WHERE o.id = :id
    """),
        {"id": order_id},
    ).fetchone()

    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    order_dict = dict(order._mapping)

    # Fetch order items
    items = db.execute(
        text("""
        SELECT oi.id, oi.order_id, oi.inventory_id, oi.product_id, oi.product_name,
               oi.sku, oi.size, oi.color, oi.hsn_code, oi.gst_rate,
               oi.quantity, oi.unit_price, oi.price, oi.created_at,
               p.name as product_name,
               COALESCE(i.image_url, pi.image_url) as image_url
        FROM order_items oi
        LEFT JOIN products p ON p.id = oi.product_id
        LEFT JOIN inventory i ON i.id = oi.inventory_id
        LEFT JOIN LATERAL (
            SELECT image_url FROM product_images
            WHERE product_id = oi.product_id AND is_primary = true
            LIMIT 1
        ) pi ON true
        WHERE oi.order_id = :oid
    """),
        {"oid": order_id},
    ).fetchall()
    order_dict["items"] = [dict(i._mapping) for i in items]

    # Fetch tracking
    tracking = db.execute(
        text(
            "SELECT * FROM order_tracking WHERE order_id = :oid ORDER BY created_at DESC"
        ),
        {"oid": order_id},
    ).fetchall()
    order_dict["tracking"] = [dict(t._mapping) for t in tracking]

    return {
        "order": order_dict,
        "items": order_dict["items"],
        "tracking": order_dict["tracking"],
        "customer": {
            "id": order_dict.get("user_id"),
            "full_name": order_dict.get("customer_name") or "Customer",
            "email": order_dict.get("customer_email"),
            "phone": order_dict.get("customer_phone"),
        },
    }


@app.patch("/api/v1/admin/orders/{order_id}/status", tags=["Admin"])
async def update_order_status(
    order_id: int,
    data: OrderStatusUpdate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_staff),
):
    """Update order status (admin/staff only)."""
    new_status = data.status

    # Validate status
    valid_statuses = ["confirmed", "shipped", "delivered", "cancelled"]
    if new_status not in valid_statuses:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid status. Must be one of: {', '.join(valid_statuses)}",
        )

    # POD number is required when shipping
    pod_number = data.get_pod()
    if new_status == "shipped" and not pod_number:
        raise HTTPException(
            status_code=400,
            detail="POD number is required when marking order as shipped",
        )

    now = datetime.now(timezone.utc)
    extra_sets = ""
    params = {"status": new_status, "now": now, "id": order_id}

    if new_status == "shipped" and pod_number:
        # Extract courier_name from the request if provided
        courier_name = getattr(data, 'courier', None) or getattr(data, 'courier_name', None)
        extra_sets = ", tracking_number = :pod, courier_name = :courier_name, shipped_at = :now"
        params["pod"] = pod_number
        params["courier_name"] = courier_name
    elif new_status == "delivered":
        extra_sets = ", delivered_at = :now"
    elif new_status == "cancelled":
        extra_sets = ", cancelled_at = :now"

    result = db.execute(
        text(
            f"UPDATE orders SET status = :status, updated_at = :now{extra_sets} WHERE id = :id RETURNING id"
        ),
        params,
    )
    db.commit()

    if not result.fetchone():
        raise HTTPException(status_code=404, detail="Order not found")

    default_notes = {
        "shipped": f"Order shipped — POD number: {pod_number}"
        if pod_number
        else "Order shipped",
        "delivered": "Order delivered to customer",
        "cancelled": "Order cancelled by admin",
        "confirmed": "Order confirmed",
    }
    
    # Extract courier_name for tracking entry
    courier_name = getattr(data, 'courier', None) or getattr(data, 'courier_name', None)
    
    db.execute(
        text("""
        INSERT INTO order_tracking (order_id, status, notes, updated_by, created_at, courier_name)
        VALUES (:order_id, :status, :notes, :updated_by, :created_at, :courier_name)
    """),
        {
            "order_id": order_id,
            "status": new_status,
            "notes": data.notes or default_notes.get(new_status),
            "updated_by": current_user.get("user_id"),
            "created_at": now,
            "courier_name": courier_name,
        },
    )
    db.commit()

    # Invalidate cache
    redis_client.delete_cache(f"order:{order_id}")
    redis_client.delete_cache("admin:orders:*")

    log_business_event(
        "order_status_updated",
        {
            "order_id": order_id,
            "new_status": new_status,
            "updated_by": current_user.get("user_id"),
        },
    )

    return {"success": True, "order_id": order_id, "status": new_status}


@app.patch("/api/v1/admin/orders/bulk-status", tags=["Admin"])
async def bulk_update_order_status(
    data: BulkOrderUpdate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_staff),
):
    """Bulk update order statuses (admin/staff only)."""
    if not data.order_ids:
        raise HTTPException(status_code=400, detail="order_ids is required")

    valid_statuses = ["confirmed", "shipped", "delivered", "cancelled"]
    if data.status not in valid_statuses:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid status. Must be one of: {', '.join(valid_statuses)}",
        )

    now = datetime.now(timezone.utc)
    updated = db.execute(
        text("""
        UPDATE orders
        SET status = :status, updated_at = :now
            """ + (""", tracking_number = :pod, courier_name = :courier, shipped_at = :now""" if data.status == "shipped" and data.pod_number else "") + """
        WHERE id = ANY(:order_ids)
        RETURNING id
    """),
        {
            "status": data.status,
            "now": now,
            "order_ids": data.order_ids,
            "pod": data.pod_number if data.status == "shipped" else None,
            "courier": getattr(data, 'courier_name', None) if data.status == "shipped" else None,
        },
    ).fetchall()

    for row in updated:
        db.execute(
            text("""
            INSERT INTO order_tracking (order_id, status, notes, updated_by, created_at""" + (", tracking_number, courier_name" if data.status == "shipped" and data.pod_number else "") + """)
            VALUES (:order_id, :status, :notes, :updated_by, :created_at""" + (", :pod, :courier" if data.status == "shipped" and data.pod_number else "") + """)
        """),
            {
                "order_id": row[0],
                "status": data.status,
                "notes": data.notes,
                "updated_by": current_user.get("user_id"),
                "created_at": now,
                "pod": data.pod_number if data.status == "shipped" else None,
                "courier": getattr(data, 'courier_name', None) if data.status == "shipped" else None,
            },
        )

    db.commit()
    redis_client.delete_cache("admin:orders:*")
    return {"updated": len(updated), "status": data.status}


@app.get("/api/v1/admin/orders/{order_id}/tracking", tags=["Admin"])
async def get_order_tracking(
    order_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_staff),
):
    """Get order tracking history (admin/staff only)."""
    tracking = db.execute(
        text(
            "SELECT * FROM order_tracking WHERE order_id = :oid ORDER BY created_at DESC"
        ),
        {"oid": order_id},
    ).fetchall()
    return {"tracking": [dict(t._mapping) for t in tracking]}


@app.get("/api/v1/admin/orders/export/excel", tags=["Admin"])
async def export_orders_excel(
    status: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_staff),
):
    """Export orders as Excel file with separate sheets per day. Filters: status, from_date (YYYY-MM-DD), to_date (YYYY-MM-DD)."""
    try:
        import io
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from fastapi.responses import StreamingResponse
    except ImportError:
        raise HTTPException(
            status_code=500, detail="openpyxl not installed. Run: pip install openpyxl"
        )

    where_clauses = []
    params: dict = {}
    if status and status in ["confirmed", "shipped", "delivered", "cancelled"]:
        where_clauses.append("o.status = :status")
        params["status"] = status
    if from_date:
        where_clauses.append("DATE(o.created_at) >= :from_date")
        params["from_date"] = from_date
    if to_date:
        where_clauses.append("DATE(o.created_at) <= :to_date")
        params["to_date"] = to_date

    where_clause = f"WHERE {' AND '.join(where_clauses)}" if where_clauses else ""

    rows = db.execute(
        text(f"""
        SELECT o.id, u.email, COALESCE(up.full_name, u.username) as customer_name,
               up.phone as customer_phone,
               o.total_amount, o.payment_method, o.status, o.tracking_number, o.courier_name,
               o.shipping_address, o.created_at
        FROM orders o
        LEFT JOIN users u ON u.id = o.user_id
        LEFT JOIN user_profiles up ON up.user_id = o.user_id
        {where_clause}
        ORDER BY o.created_at DESC
        LIMIT 5000
    """),
        params,
    ).fetchall()

    # Group orders by date
    from collections import defaultdict
    orders_by_date = defaultdict(list)
    for r in rows:
        order_date = r[9]
        if order_date:
            date_key = order_date.strftime('%Y-%m-%d') if hasattr(order_date, 'strftime') else str(order_date)[:10]
            orders_by_date[date_key].append(r)

    wb = openpyxl.Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    headers = [
        "Order ID",
        "Order #",
        "Customer Email",
        "Customer Name",
        "Phone",
        "Total (₹)",
        "Payment Method",
        "POD / Tracking No.",
        "Courier Service",
        "Shipping Address",
        "Order Date",
    ]

    header_fill = PatternFill(
        start_color="180F14", end_color="180F14", fill_type="solid"
    )
    header_font = Font(bold=True, color="F2C29A")

    # Create a sheet for each date, sorted by date
    for date_key in sorted(orders_by_date.keys()):
        date_orders = orders_by_date[date_key]
        
        # Create sheet with date as name (truncate if too long)
        sheet_name = date_key[:10]  # YYYY-MM-DD
        # Handle sheet name length limit (31 chars max)
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]
        ws = wb.create_sheet(title=sheet_name)
        
        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Write data
        # SQL column indices: 0=id, 1=email, 2=customer_name, 3=customer_phone,
        # 4=total_amount, 5=payment_method, 6=status, 7=tracking_number, 8=courier_name,
        # 9=shipping_address, 10=created_at
        for row_idx, r in enumerate(date_orders, 2):
            order_id = r[0]
            ws.cell(row=row_idx, column=1, value=order_id)
            ws.cell(row=row_idx, column=2, value=f"ORD-{order_id:06d}")
            ws.cell(row=row_idx, column=3, value=r[1])   # email
            ws.cell(row=row_idx, column=4, value=r[2])   # customer_name
            ws.cell(row=row_idx, column=5, value=r[3] or "")   # customer_phone
            ws.cell(row=row_idx, column=6, value=float(r[4] or 0))  # total_amount
            ws.cell(row=row_idx, column=7, value=r[5])   # payment_method
            ws.cell(row=row_idx, column=8, value=r[7])   # tracking_number / POD
            ws.cell(row=row_idx, column=9, value=r[8] or "")   # courier_name
            ws.cell(row=row_idx, column=10, value=r[9])   # shipping_address
            ws.cell(row=row_idx, column=11, value=str(r[10])[:19] if r[10] else "")  # created_at
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    # If no orders, create empty sheet
    if not orders_by_date:
        ws = wb.create_sheet(title="No Orders")
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"orders_{status or 'all'}_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ==================== POD Excel Upload ====================


@app.post("/api/v1/admin/orders/upload-pod-excel", tags=["Admin"])
async def upload_pod_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_staff),
):
    """
    Upload Excel file with columns: Order ID, POD / Tracking No.
    Batch-updates tracking_number + status=shipped for confirmed orders.
    Returns: updated_count, skipped, errors list.
    """
    try:
        import io, openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx files are accepted")

    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    ws = wb.active

    # Detect header row — look for "Order ID", "POD", and optional "Courier" columns
    header_map = {}
    for col_idx, cell in enumerate(ws[1], 1):
        val = str(cell.value or "").strip().lower()
        if "order id" in val or val == "order #" or val == "id":
            header_map["order_id_col"] = col_idx
        if "pod" in val or "tracking" in val:
            header_map["pod_col"] = col_idx
        if "courier" in val:
            header_map["courier_col"] = col_idx

    if "order_id_col" not in header_map or "pod_col" not in header_map:
        raise HTTPException(
            status_code=400,
            detail="Excel must have columns: 'Order ID' and 'POD / Tracking No.'",
        )

    updated, skipped, errors = 0, 0, []
    staff_id = current_user.get("user_id")

    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            raw_id = row[header_map["order_id_col"] - 1]
            raw_pod = row[header_map["pod_col"] - 1]
            # Extract courier_name if column exists
            courier_name = None
            if "courier_col" in header_map:
                courier_name = str(row[header_map["courier_col"] - 1] or "").strip() or None
            
            if raw_id is None:
                continue
            # Handle "ORD-000001" format
            order_id_str = str(raw_id).strip().replace("ORD-", "").lstrip("0") or "0"
            order_id = int(order_id_str)
            pod = str(raw_pod or "").strip()
            if not pod:
                skipped += 1
                continue

            order = db.execute(
                text("SELECT id, status FROM orders WHERE id = :oid"), {"oid": order_id}
            ).fetchone()

            if not order:
                errors.append(f"Order #{order_id}: not found")
                continue
            if str(order[1]) not in ("confirmed",):
                errors.append(
                    f"Order #{order_id}: status is '{order[1]}', can only ship confirmed orders"
                )
                skipped += 1
                continue

            db.execute(
                text("""
                UPDATE orders
                SET status = 'shipped', tracking_number = :pod, courier_name = :courier_name,
                    shipped_at = :now, updated_at = :now
                WHERE id = :oid
            """),
                {"pod": pod, "courier_name": courier_name, "now": datetime.now(timezone.utc), "oid": order_id},
            )

            db.execute(
                text("""
                INSERT INTO order_tracking (order_id, status, notes, updated_by, created_at, courier_name)
                VALUES (:oid, 'shipped', :notes, :by, :now, :courier_name)
            """),
                {
                    "oid": order_id,
                    "notes": f"Shipped via Excel upload. POD: {pod}" + (f", Courier: {courier_name}" if courier_name else ""),
                    "by": staff_id,
                    "now": datetime.now(timezone.utc),
                    "courier_name": courier_name,
                },
            )
            updated += 1
        except (ValueError, IndexError) as e:
            errors.append(f"Row error: {e}")

    db.commit()
    return {
        "updated": updated,
        "skipped": skipped,
        "errors": errors[:20],
        "message": f"Successfully shipped {updated} order(s). {skipped} skipped.",
    }


@app.get("/api/v1/admin/orders/pod-template", tags=["Admin"])
async def download_pod_template(
    db: Session = Depends(get_db), current_user: dict = Depends(require_staff)
):
    """Download an Excel template pre-filled with confirmed orders for POD entry."""
    try:
        import io, openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from fastapi.responses import StreamingResponse
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    rows = db.execute(
        text("""
        SELECT o.id, u.email, u.username, o.total_amount, o.created_at,
               o.shipping_address
        FROM orders o
        LEFT JOIN users u ON u.id = o.user_id
        WHERE o.status = 'confirmed'
        ORDER BY o.created_at DESC
        LIMIT 500
    """)
    ).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "POD Upload"

    headers = [
        "Order ID",
        "Order #",
        "Customer Email",
        "Customer Name",
        "Total (₹)",
        "Order Date",
        "Shipping Address",
        "POD / Tracking No.",
        "Courier Name",
    ]
    hfill = PatternFill(start_color="180F14", end_color="180F14", fill_type="solid")
    hfont = Font(bold=True, color="F2C29A")

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center")

    # POD and Courier column highlights
    pod_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")

    for ri, r in enumerate(rows, 2):
        ws.cell(ri, 1, value=r[0])
        ws.cell(ri, 2, value=f"ORD-{r[0]:06d}")
        ws.cell(ri, 3, value=r[1])
        ws.cell(ri, 4, value=r[2])
        ws.cell(ri, 5, value=float(r[3] or 0))
        ws.cell(ri, 6, value=str(r[4])[:10] if r[4] else "")
        ws.cell(ri, 7, value=str(r[5] or "")[:80])
        pod_cell = ws.cell(ri, 8, value="")
        pod_cell.fill = pod_fill
        courier_cell = ws.cell(ri, 9, value="")
        courier_cell.fill = pod_fill  # Same highlight for both input fields

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    fname = f"pod_template_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


# ==================== AI Endpoints ====================


@app.post("/api/v1/ai/customer/chat", tags=["AI"])
async def ai_customer_chat(request: Request, db: Session = Depends(get_db)):
    """
    Customer AI Salesman — public endpoint (no auth required).
    Accepts: { message, session_id?, language?, cart_context? }
    Returns: { session_id, reply, tool_results, tokens_used, cost_usd }
    """
    from service.ai_service import customer_chat

    body = await request.json()
    message = str(body.get("message", "")).strip()
    session_id = body.get("session_id")
    language = body.get("language", "auto")
    cart_context = body.get("cart_context")  # optional: {items, total, item_count}
    user_id = None

    if not message:
        raise HTTPException(status_code=400, detail="message is required")
    if len(message) > 1000:
        raise HTTPException(status_code=400, detail="Message too long (max 1000 chars)")

    # Extract user_id from HttpOnly cookie (preferred) or Authorization header (fallback)
    try:
        cookie_token = request.cookies.get("access_token")
        auth_header = request.headers.get("Authorization", "")
        token = cookie_token or (auth_header.split(" ", 1)[1] if auth_header.startswith("Bearer ") else None)
        if token:
            from shared.auth_middleware import auth_middleware
            payload = auth_middleware.decode_token(token)
            user_info = auth_middleware.extract_user_info(payload)
            user_id = user_info.get("user_id")
    except Exception:
        pass

    return customer_chat(db, message, session_id, user_id, language=language, cart_context=cart_context)


@app.post("/api/v1/ai/customer/chat/stream", tags=["AI"])
async def ai_customer_chat_stream(request: Request, db: Session = Depends(get_db)):
    """
    SSE streaming variant of customer AI chat.
    Accepts: { message, session_id?, language?, cart_context? }
    Returns: text/event-stream with chunks: data: {"chunk": "..."} and final: data: {"done": true, ...}
    """
    from service.ai_service import customer_chat
    from fastapi.responses import StreamingResponse as _StreamingResponse
    import asyncio, json as _json

    body = await request.json()
    message = str(body.get("message", "")).strip()
    session_id = body.get("session_id")
    language = body.get("language", "auto")
    cart_context = body.get("cart_context")
    user_id = None

    if not message:
        raise HTTPException(status_code=400, detail="message is required")
    if len(message) > 1000:
        raise HTTPException(status_code=400, detail="Message too long (max 1000 chars)")

    try:
        cookie_token = request.cookies.get("access_token")
        auth_header = request.headers.get("Authorization", "")
        token = cookie_token or (auth_header.split(" ", 1)[1] if auth_header.startswith("Bearer ") else None)
        if token:
            from shared.auth_middleware import auth_middleware
            payload = auth_middleware.decode_token(token)
            user_info = auth_middleware.extract_user_info(payload)
            user_id = user_info.get("user_id")
    except Exception:
        pass

    async def event_generator():
        try:
            # Run the blocking customer_chat in a thread pool
            loop = asyncio.get_event_loop()
            result = await loop.run_in_executor(
                None,
                lambda: customer_chat(db, message, session_id, user_id, language=language, cart_context=cart_context)
            )
            reply = result.get("reply", "")
            # Stream the reply character-by-character in chunks for typing effect
            CHUNK_SIZE = 3
            for i in range(0, len(reply), CHUNK_SIZE):
                chunk = reply[i:i + CHUNK_SIZE]
                yield f"data: {_json.dumps({'chunk': chunk})}\n\n"
                await asyncio.sleep(0.008)
            # Final event with full metadata
            yield f"data: {_json.dumps({'done': True, 'session_id': result.get('session_id'), 'tool_results': result.get('tool_results'), 'tokens_used': result.get('tokens_used', 0)})}\n\n"
        except Exception as e:
            logger.error(f"AI stream error: {e}")
            yield f"data: {_json.dumps({'error': True, 'chunk': 'I ran into a small issue. Please try again! 🌸'})}\n\n"
            yield f"data: {_json.dumps({'done': True})}\n\n"

    return _StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.post("/api/v1/ai/admin/chat", tags=["AI"])
async def ai_admin_chat(
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_staff),
):
    """
    Admin AI Assistant — staff/admin only, full agentic tool suite.
    Accepts: { message, session_id?, images?: [{mime_type, data (base64)}] }
    Returns: { session_id, reply, tool_calls, tokens_used, cost_usd }
    """
    from service.ai_service import admin_chat

    body = await request.json()
    message = str(body.get("message", "")).strip()
    session_id = body.get("session_id")
    images = body.get("images")  # optional multimodal

    if not message:
        raise HTTPException(status_code=400, detail="message is required")
    if len(message) > 4000:
        raise HTTPException(status_code=400, detail="Message too long (max 4000 chars)")

    user_id = current_user["user_id"]
    return admin_chat(db, message, session_id, user_id, image_data=images)


@app.get("/api/v1/ai/admin/sessions", tags=["AI"])
async def get_ai_sessions(
    role: Optional[str] = None,
    days: int = 30,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_admin),
):
    """Get AI session list for analytics (admin only)."""
    where = f"WHERE s.created_at >= NOW() - INTERVAL '{int(days)} days'"
    params: dict = {}
    if role in ("customer", "admin"):
        where += " AND s.role = :role"
        params["role"] = role

    sessions = db.execute(
        text(f"""
        SELECT s.session_id, s.role, s.user_id, u.email,
               s.message_count, s.total_tokens_in, s.total_tokens_out,
               s.total_cost, s.created_at, s.last_activity
        FROM ai_sessions s
        LEFT JOIN users u ON u.id = s.user_id
        {where}
        ORDER BY s.last_activity DESC LIMIT 100
    """),
        params,
    ).fetchall()

    return {
        "sessions": [
            {
                "session_id": r[0],
                "role": r[1],
                "user_id": r[2],
                "email": r[3],
                "messages": r[4],
                "tokens_in": r[5],
                "tokens_out": r[6],
                "cost_usd": float(r[7] or 0),
                "created_at": str(r[8]),
                "last_activity": str(r[9]),
            }
            for r in sessions
        ]
    }


@app.post("/api/v1/ai/admin/execute-action", tags=["AI"])
async def execute_ai_action(
    data: dict,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_admin),
):
    """
    Execute a previously confirmed pending AI action.
    Called by admin after reviewing the pending action in the confirmation modal.
    DELETE operations are never supported here.
    """
    from service.ai_service import execute_confirmed_action

    action_type = data.get("action_type", "")
    params = data.get("params", {})
    admin_user_id = current_user.get("user_id")

    allowed_actions = {
        "ship_order",
        "update_product_price",
        "bulk_update_category_prices",
        "adjust_stock",
        "create_product_draft",
    }
    if action_type not in allowed_actions:
        raise HTTPException(
            status_code=400, detail=f"Action '{action_type}' is not permitted."
        )

    result = execute_confirmed_action(db, action_type, params, admin_user_id)
    if not result.get("success"):
        raise HTTPException(
            status_code=422, detail=result.get("message", "Action failed.")
        )
    return result


@app.get("/api/v1/ai/admin/analytics", tags=["AI"])
async def get_ai_analytics(
    days: int = 30,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_admin),
):
    """Get AI usage + cost analytics (admin only)."""
    from service.ai_service import get_ai_analytics

    return get_ai_analytics(db, days)


# ── pgvector embedding management endpoints ───────────────────────────────────


@app.post("/api/v1/ai/embeddings/refresh", tags=["AI Embeddings"])
async def refresh_all_embeddings(
    batch_size: int = 50,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_admin),
):
    """
    DEPRECATED: Embedding generation is not supported by Groq and other OpenAI-compatible providers.
    
    This endpoint is kept for backward compatibility but will return a notice.
    Consider using an external embedding service (e.g., OpenAI, Cohere) if embeddings are needed.
    
    batch_size controls how many products to process per call (max 200).
    """
    from service.ai_service import generate_product_embeddings_batch

    size = min(batch_size, 200)
    result = generate_product_embeddings_batch(db, batch_size=size)
    if not result.get("success"):
        raise HTTPException(
            status_code=503, detail=result.get("error", "Embedding generation failed")
        )
    return result


@app.post("/api/v1/ai/embeddings/refresh-all", tags=["AI Embeddings"])
async def refresh_all_embeddings_force(
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_admin),
):
    """
    Force-regenerate embeddings for ALL products (overwrite existing).
    Use this after a bulk product description update.
    """
    from service.ai_service import generate_product_embeddings_batch
    from sqlalchemy import text as sa_text

    db.execute(sa_text("UPDATE products SET embedding = NULL"))
    db.commit()
    result = generate_product_embeddings_batch(db, batch_size=500)
    return result


@app.post("/api/v1/ai/embeddings/product/{product_id}", tags=["AI Embeddings"])
async def refresh_single_embedding(
    product_id: str,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_admin),
):
    """Generate or refresh embedding for a single product by ID."""
    from service.ai_service import generate_single_product_embedding

    result = generate_single_product_embedding(db, product_id)
    if not result.get("success"):
        raise HTTPException(
            status_code=503, detail=result.get("error", "Embedding generation failed")
        )
    return result


@app.get("/api/v1/ai/embeddings/status", tags=["AI Embeddings"])
async def get_embedding_status(
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_admin),
):
    """Get current embedding coverage stats for all products."""
    from sqlalchemy import text as sa_text

    total = db.execute(sa_text("SELECT COUNT(*) FROM products")).scalar() or 0
    with_emb = (
        db.execute(
            sa_text("SELECT COUNT(*) FROM products WHERE embedding IS NOT NULL")
        ).scalar()
        or 0
    )
    without_emb = total - with_emb
    coverage_pct = round((with_emb / total * 100) if total > 0 else 0, 1)
    return {
        "total_products": total,
        "with_embeddings": with_emb,
        "without_embeddings": without_emb,
        "coverage_percent": coverage_pct,
        "ready_for_semantic_search": with_emb > 0,
        "message": f"{with_emb}/{total} products indexed ({coverage_pct}%)",
    }


@app.get("/api/v1/super/ai-monitoring", tags=["AI"])
async def get_ai_monitoring(
    days: int = 30,
    role: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_super_admin),
):
    """Full AI monitoring: daily chart, per-model, per-user, recent sessions."""
    role_clause = "AND s.role = :role" if role else ""
    params: dict = {"days": days}
    if role:
        params["role"] = role

    # Daily cost/token breakdown for chart
    daily = db.execute(
        text(f"""
        SELECT DATE(m.created_at) as day,
               SUM(m.tokens_in) as tokens_in,
               SUM(m.tokens_out) as tokens_out,
               SUM(m.cost) as cost,
               COUNT(*) as messages,
               m.model_used
        FROM ai_messages m
        JOIN ai_sessions s ON s.session_id = m.session_id
        WHERE m.created_at >= NOW() - INTERVAL '{days} days'
          AND m.role = 'assistant'
          {role_clause}
        GROUP BY DATE(m.created_at), m.model_used
        ORDER BY day ASC
    """),
        params,
    ).fetchall()

    # Per-model summary
    by_model = db.execute(
        text(f"""
        SELECT m.model_used,
               COUNT(DISTINCT s.session_id) as sessions,
               COUNT(*) FILTER (WHERE m.role = 'assistant') as responses,
               SUM(m.tokens_in) as tokens_in,
               SUM(m.tokens_out) as tokens_out,
               SUM(m.cost) as cost
        FROM ai_messages m
        JOIN ai_sessions s ON s.session_id = m.session_id
        WHERE m.created_at >= NOW() - INTERVAL '{days} days'
          {role_clause}
        GROUP BY m.model_used
    """),
        params,
    ).fetchall()

    # Top 10 users by cost
    top_users = db.execute(
        text(f"""
        SELECT u.email, u.username, s.role,
               COUNT(DISTINCT s.session_id) as sessions,
               SUM(s.total_cost) as cost
        FROM ai_sessions s
        LEFT JOIN users u ON u.id = s.user_id
        WHERE s.created_at >= NOW() - INTERVAL '{days} days'
          {role_clause}
        GROUP BY u.email, u.username, s.role
        ORDER BY cost DESC LIMIT 10
    """),
        params,
    ).fetchall()

    # Summary totals
    totals = db.execute(
        text(f"""
        SELECT COUNT(DISTINCT session_id) as sessions,
               SUM(total_tokens_in) as tokens_in,
               SUM(total_tokens_out) as tokens_out,
               SUM(total_cost) as cost,
               SUM(message_count) as messages
        FROM ai_sessions
        WHERE created_at >= NOW() - INTERVAL '{days} days'
          {role_clause}
    """),
        params,
    ).fetchone()

    # Today's cost
    today_cost = (
        db.execute(
            text(f"""
        SELECT COALESCE(SUM(cost), 0) FROM ai_messages
        WHERE created_at >= CURRENT_DATE
        AND role = 'assistant'
    """)
        ).scalar()
        or 0
    )

    return {
        "period_days": days,
        "totals": {
            "sessions": totals[0] or 0,
            "tokens_in": totals[1] or 0,
            "tokens_out": totals[2] or 0,
            "cost_usd": float(totals[3] or 0),
            "messages": totals[4] or 0,
        },
        "today_cost_usd": float(today_cost),
        "daily": [
            {
                "date": str(r[0]),
                "tokens_in": r[1] or 0,
                "tokens_out": r[2] or 0,
                "cost_usd": float(r[3] or 0),
                "messages": r[4],
                "model": r[5],
            }
            for r in daily
        ],
        "by_model": [
            {
                "model": r[0],
                "sessions": r[1],
                "responses": r[2],
                "tokens_in": r[3] or 0,
                "tokens_out": r[4] or 0,
                "cost_usd": float(r[5] or 0),
            }
            for r in by_model
        ],
        "top_users": [
            {
                "email": r[0],
                "username": r[1],
                "role": r[2],
                "sessions": r[3],
                "cost_usd": float(r[4] or 0),
            }
            for r in top_users
        ],
    }


@app.get("/api/v1/ai/admin/sessions/{session_id}/messages", tags=["AI"])
async def get_session_messages(
    session_id: str,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_admin),
):
    """Get all messages for a specific AI session."""
    session = db.execute(
        text("""
        SELECT s.session_id, s.role, s.user_id, u.email, u.username,
               s.message_count, s.total_tokens_in, s.total_tokens_out,
               s.total_cost, s.created_at, s.last_activity
        FROM ai_sessions s LEFT JOIN users u ON u.id = s.user_id
        WHERE s.session_id = :sid
    """),
        {"sid": session_id},
    ).fetchone()
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")

    messages = db.execute(
        text("""
        SELECT id, role, content, tokens_in, tokens_out, cost,
               model_used, tool_calls, tool_results, image_urls, created_at
        FROM ai_messages WHERE session_id = :sid ORDER BY created_at ASC
    """),
        {"sid": session_id},
    ).fetchall()

    return {
        "session": {
            "session_id": session[0],
            "role": session[1],
            "user_id": session[2],
            "email": session[3],
            "username": session[4],
            "message_count": session[5],
            "tokens_in": session[6],
            "tokens_out": session[7],
            "cost_usd": float(session[8] or 0),
            "created_at": str(session[9]),
            "last_activity": str(session[10]),
        },
        "messages": [
            {
                "id": r[0],
                "role": r[1],
                "content": r[2],
                "tokens_in": r[3],
                "tokens_out": r[4],
                "cost_usd": float(r[5] or 0),
                "model": r[6],
                "tool_calls": r[7],
                "tool_results": r[8],
                "has_images": bool(r[9]),
                "created_at": str(r[10]),
            }
            for r in messages
        ],
    }


@app.get("/api/v1/ai/admin/export/csv", tags=["AI"])
async def export_ai_sessions_csv(
    days: int = 30,
    role: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_admin),
):
    """Export AI sessions as CSV download."""
    import csv, io

    role_clause = (
        f"AND s.role = '{role}'" if role and role in ("customer", "admin") else ""
    )
    rows = db.execute(
        text(f"""
        SELECT s.session_id, s.role, u.email, u.username,
               s.message_count, s.total_tokens_in, s.total_tokens_out,
               s.total_cost, s.created_at, s.last_activity
        FROM ai_sessions s LEFT JOIN users u ON u.id = s.user_id
        WHERE s.created_at >= NOW() - INTERVAL '{days} days'
          {role_clause}
        ORDER BY s.last_activity DESC
    """)
    ).fetchall()

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(
        [
            "session_id",
            "role",
            "email",
            "username",
            "messages",
            "tokens_in",
            "tokens_out",
            "cost_usd",
            "created_at",
            "last_activity",
        ]
    )
    for r in rows:
        writer.writerow(
            [
                r[0],
                r[1],
                r[2] or "",
                r[3] or "",
                r[4],
                r[5],
                r[6],
                f"{float(r[7] or 0):.8f}",
                str(r[8]),
                str(r[9]),
            ]
        )

    content = buf.getvalue().encode("utf-8")
    return Response(
        content=content,
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=ai_sessions_{days}d.csv"
        },
    )


# ── AI Settings (CRUD) ────────────────────────────────────────────────────────

from utils.encryption import encrypt_api_key, decrypt_api_key


@app.get("/api/v1/super/ai-settings", tags=["AI Settings"])
async def get_ai_settings(
    db: Session = Depends(get_db), current_user: dict = Depends(require_super_admin)
):
    """Get all AI settings (secrets masked)."""
    rows = db.execute(
        text("""
        SELECT key, value, description, is_secret, category, updated_at
        FROM ai_settings ORDER BY category, key
    """)
    ).fetchall()
    return {
        "settings": [
            {
                "key": r[0],
                "value": "••••••••" if r[3] and r[1] else r[1],
                "raw_set": bool(r[1]),
                "description": r[2],
                "is_secret": r[3],
                "category": r[4],
                "updated_at": str(r[5]),
            }
            for r in rows
        ]
    }


@app.get("/api/v1/super/ai-providers/status", tags=["AI Providers"])
async def get_ai_providers_status(
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_super_admin),
):
    """Get real-time status of all AI providers with rate limits and usage."""
    try:
        from core.ai_key_rotation import get_provider_status
        status = get_provider_status(db)
        return {
            "providers": status,
            "total_providers": len(status),
            "enabled_providers": sum(1 for p in status if p["enabled"]),
        }
    except Exception as e:
        logger.error(f"Failed to get AI provider status: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to get provider status: {str(e)}")


@app.post("/api/v1/super/ai-settings/test-key", tags=["AI Settings"])
async def test_api_key(
    data: dict,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_super_admin),
):
    """Test an API key for any supported AI provider."""
    key = data.get("api_key", "").strip()
    provider = data.get("provider", "groq")

    if not key or key == "__current__":
        # Get from environment
        key = os.environ.get(f"{provider.upper()}_API_KEY", "")

    if not key:
        raise HTTPException(status_code=400, detail="No API key configured")

    try:
        # Test Groq (PRIMARY provider)
        if provider == "groq":
            from groq import Groq
            client = Groq(api_key=key)
            resp = client.chat.completions.create(
                model="llama-3.3-70b-versatile",
                messages=[{"role": "user", "content": "Say 'ok' in one word"}],
                max_tokens=5
            )
            return {"valid": True, "response": resp.choices[0].message.content.strip(), "provider": provider}

        # Test OpenRouter
        elif provider == "openrouter":
            from openai import OpenAI
            client = OpenAI(api_key=key, base_url="https://openrouter.ai/api/v1")
            resp = client.chat.completions.create(
                model="meta-llama/llama-3.3-70b-instruct:free",
                messages=[{"role": "user", "content": "Say 'ok' in one word"}],
                max_tokens=5
            )
            return {"valid": True, "response": resp.choices[0].message.content.strip(), "provider": provider}

        # Test GLM
        elif provider == "glm":
            from openai import OpenAI
            client = OpenAI(api_key=key, base_url="https://open.bigmodel.cn/api/paas/v4")
            resp = client.chat.completions.create(
                model="glm-4-flash",
                messages=[{"role": "user", "content": "Say 'ok' in one word"}],
                max_tokens=5
            )
            return {"valid": True, "response": resp.choices[0].message.content.strip(), "provider": provider}

        # Test NVIDIA
        elif provider == "nvidia":
            from openai import OpenAI
            client = OpenAI(api_key=key, base_url="https://integrate.api.nvidia.com/v1")
            resp = client.chat.completions.create(
                model="meta/llama3-70b-instruct",
                messages=[{"role": "user", "content": "Say 'ok' in one word"}],
                max_tokens=5
            )
            return {"valid": True, "response": resp.choices[0].message.content.strip(), "provider": provider}

        else:
            raise HTTPException(status_code=400, detail=f"Unsupported provider: {provider}. Supported: groq, openrouter, glm, nvidia")

    except Exception as e:
        return {"valid": False, "error": str(e), "provider": provider}


@app.put("/api/v1/super/ai-settings/bulk", tags=["AI Settings"])
async def bulk_update_ai_settings(
    data: dict,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_super_admin),
):
    """Bulk update multiple AI settings at once. API keys are automatically encrypted."""
    settings_data: dict = data.get("settings", {})
    user_id = current_user.get("user_id")
    now = datetime.now(timezone.utc)
    
    for k, value in settings_data.items():
        # Encrypt API keys
        is_api_key = k.endswith("_API_KEY")
        value_to_store = encrypt_api_key(value) if is_api_key and value else value
        
        existing = db.execute(
            text("SELECT id FROM ai_settings WHERE key = :k"), {"k": k}
        ).fetchone()
        if existing:
            db.execute(
                text("""
                UPDATE ai_settings SET value = :v, updated_at = :now, updated_by = :uid WHERE key = :k
            """),
                {"v": value_to_store, "now": now, "uid": user_id, "k": k},
            )
        else:
            db.execute(
                text("""
                INSERT INTO ai_settings (key, value, updated_by, updated_at) VALUES (:k, :v, :uid, :now)
            """),
                {"k": k, "v": value_to_store, "uid": user_id, "now": now},
            )
        
        # Update environment variable (decrypted for runtime)
        if is_api_key and value:
            os.environ[k] = value
    
    db.commit()
    return {"success": True, "updated": list(settings_data.keys()), "encrypted_keys": [k for k in settings_data.keys() if k.endswith("_API_KEY")]}


@app.put("/api/v1/super/ai-settings/{key}", tags=["AI Settings"])
async def update_ai_setting(
    key: str,
    data: dict,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_super_admin),
):
    """Update a single AI setting value. API keys are automatically encrypted."""
    value = data.get("value", "")
    user_id = current_user.get("user_id")
    
    # Check if this is an API key that should be encrypted
    is_api_key = key.endswith("_API_KEY")
    value_to_store = encrypt_api_key(value) if is_api_key and value else value

    existing = db.execute(
        text("SELECT id FROM ai_settings WHERE key = :k"), {"k": key}
    ).fetchone()
    if existing:
        db.execute(
            text("""
            UPDATE ai_settings SET value = :v, updated_at = :now, updated_by = :uid
            WHERE key = :k
        """),
            {"v": value_to_store, "now": datetime.now(timezone.utc), "uid": user_id, "k": key},
        )
    else:
        db.execute(
            text("""
            INSERT INTO ai_settings (key, value, updated_by, updated_at)
            VALUES (:k, :v, :uid, :now)
        """),
            {"k": key, "v": value_to_store, "uid": user_id, "now": datetime.now(timezone.utc)},
        )
    db.commit()

    # Also update environment variable (decrypt if needed for runtime use)
    if is_api_key and value:
        os.environ[key] = value  # Store decrypted in env for runtime use

    return {"success": True, "key": key, "encrypted": is_api_key and value}


# ==================== Landing Config ====================


@app.get("/api/v1/admin/landing/config", tags=["Landing Config"])
async def get_landing_config(
    db: Session = Depends(get_db), user: dict = Depends(require_admin)
):
    rows = db.execute(text("SELECT * FROM landing_config ORDER BY section")).fetchall()
    return {"sections": [dict(r._mapping) for r in rows]}


@app.put("/api/v1/admin/landing/config/{section}", tags=["Landing Config"])
async def update_landing_config(
    section: str,
    data: LandingConfigUpdate,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    existing = db.execute(
        text("SELECT id FROM landing_config WHERE section = :s"), {"s": section}
    ).fetchone()
    config_json = json.dumps(data.config)
    user_id = user.get("user_id")

    if existing:
        updates = {
            "config": config_json,
            "updated_by": user_id,
            "now": datetime.now(timezone.utc),
            "s": section,
        }
        if data.is_active is not None:
            db.execute(
                text(
                    "UPDATE landing_config SET config = :config::jsonb, is_active = :active, updated_by = :updated_by, updated_at = :now WHERE section = :s"
                ),
                {**updates, "active": data.is_active},
            )
        else:
            db.execute(
                text(
                    "UPDATE landing_config SET config = :config::jsonb, updated_by = :updated_by, updated_at = :now WHERE section = :s"
                ),
                updates,
            )
    else:
        db.execute(
            text(
                "INSERT INTO landing_config (section, config, updated_by) VALUES (:s, :config::jsonb, :by)"
            ),
            {"s": section, "config": config_json, "by": user_id},
        )
    db.commit()
    redis_client.invalidate_pattern("public:landing:*")
    return {"message": f"Landing config for '{section}' updated"}


@app.get("/api/v1/admin/landing/images", tags=["Landing Config"])
async def get_landing_images(
    section: Optional[str] = None,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    if section:
        rows = db.execute(
            text(
                "SELECT * FROM landing_images WHERE section = :s ORDER BY display_order"
            ),
            {"s": section},
        ).fetchall()
    else:
        rows = db.execute(
            text("SELECT * FROM landing_images ORDER BY section, display_order")
        ).fetchall()
    return {"images": [dict(r._mapping) for r in rows]}


@app.post("/api/v1/admin/landing/images", tags=["Landing Config"])
async def add_landing_image(
    data: LandingImageCreate,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    """Add a landing image using a pre-existing URL (e.g. from presigned upload)."""
    result = db.execute(
        text(
            "INSERT INTO landing_images (section, image_url, title, subtitle, link_url, display_order, device_variant) "
            "VALUES (:s, :url, :title, :sub, :link, :order, :variant) RETURNING id"
        ),
        {
            "s": data.section,
            "url": data.image_url,
            "title": data.title,
            "sub": data.subtitle,
            "link": data.link_url,
            "order": data.display_order,
            "variant": data.device_variant,
        },
    )
    db.commit()
    redis_client.invalidate_pattern(
        "public:landing:*"
    )  # Invalidate cache so changes reflect immediately
    return {"message": "Image added", "image_id": result.scalar()}


@app.post("/api/v1/admin/landing/images/upload", tags=["Landing Config"])
async def upload_landing_image(
    section: str,
    file: UploadFile = File(...),
    title: Optional[str] = None,
    subtitle: Optional[str] = None,
    link_url: Optional[str] = None,
    display_order: int = 0,
    device_variant: Optional[str] = None,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    """Upload a landing page image directly to Cloudflare R2 and save metadata."""
    image_url = await r2_service.upload_image(file, folder="landing")

    result = db.execute(
        text(
            "INSERT INTO landing_images (section, image_url, title, subtitle, link_url, display_order, device_variant) "
            "VALUES (:s, :url, :title, :sub, :link, :order, :variant) RETURNING id"
        ),
        {
            "s": section,
            "url": image_url,
            "title": title,
            "sub": subtitle,
            "link": link_url,
            "order": display_order,
            "variant": device_variant,
        },
    )
    db.commit()

    # Invalidate cache so changes show immediately
    redis_client.invalidate_pattern("public:landing:*")

    return {
        "message": "Image uploaded and saved",
        "image_id": result.scalar(),
        "image_url": image_url,
    }


@app.patch("/api/v1/admin/landing/images/{image_id}", tags=["Landing Config"])
async def update_landing_image(
    image_id: int,
    title: Optional[str] = None,
    subtitle: Optional[str] = None,
    link_url: Optional[str] = None,
    display_order: Optional[int] = None,
    is_active: Optional[bool] = None,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    """Update landing image metadata."""
    sets, params = [], {"id": image_id}
    if title is not None:
        sets.append("title = :title")
        params["title"] = title
    if subtitle is not None:
        sets.append("subtitle = :subtitle")
        params["subtitle"] = subtitle
    if link_url is not None:
        sets.append("link_url = :link_url")
        params["link_url"] = link_url
    if display_order is not None:
        sets.append("display_order = :display_order")
        params["display_order"] = display_order
    if is_active is not None:
        sets.append("is_active = :is_active")
        params["is_active"] = is_active
    if not sets:
        return {"message": "Nothing to update"}
    db.execute(
        text(f"UPDATE landing_images SET {', '.join(sets)} WHERE id = :id"), params
    )
    db.commit()
    redis_client.invalidate_pattern(
        "public:landing:*"
    )  # Invalidate cache so changes reflect immediately
    return {"message": "Image updated"}


@app.post("/api/v1/admin/landing/images/reorder", tags=["Landing Config"])
async def reorder_landing_images(
    section: str,
    ordered_ids: List[int],
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    """Reorder images in a section by providing ordered list of image IDs."""
    for idx, image_id in enumerate(ordered_ids):
        db.execute(
            text(
                "UPDATE landing_images SET display_order = :order WHERE id = :id AND section = :section"
            ),
            {"order": idx, "id": image_id, "section": section},
        )
    db.commit()
    redis_client.invalidate_pattern("public:landing:*")
    return {"message": f"Reordered {len(ordered_ids)} images in section '{section}'"}


@app.delete("/api/v1/admin/landing/images/{image_id}", tags=["Landing Config"])
async def delete_landing_image(
    image_id: int, db: Session = Depends(get_db), user: dict = Depends(require_admin)
):
    """Delete a landing image from both the database and Cloudflare R2."""
    row = db.execute(
        text("SELECT image_url FROM landing_images WHERE id = :id"), {"id": image_id}
    ).fetchone()
    if row and row[0]:
        await r2_service.delete_image(row[0])
    db.execute(text("DELETE FROM landing_images WHERE id = :id"), {"id": image_id})
    db.commit()
    redis_client.invalidate_pattern("public:landing:*")
    return {"message": "Image deleted"}


# ==================== Landing Products Management ====================


@app.get("/api/v1/admin/landing/products", tags=["Landing Config"])
async def get_landing_products(
    section: Optional[str] = None,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    """Get admin-selected products for landing sections with full product details."""
    query = """
        SELECT lp.id, lp.section, lp.product_id, lp.display_order, lp.is_active,
               p.name, p.slug, p.base_price as price, p.mrp, p.short_description,
               c.name as collection_name,
               pi.image_url as primary_image
        FROM landing_products lp
        JOIN products p ON lp.product_id = p.id
        LEFT JOIN collections c ON p.category_id = c.id
        LEFT JOIN product_images pi ON p.id = pi.product_id AND pi.is_primary = true
    """
    params = {}
    if section:
        query += " WHERE lp.section = :section"
        params["section"] = section
    query += " ORDER BY lp.section, lp.display_order"
    rows = db.execute(text(query), params).fetchall()
    products = []
    for r in rows:
        products.append(
            {
                "id": r[0],
                "section": r[1],
                "product_id": r[2],
                "display_order": r[3],
                "is_active": r[4],
                "name": r[5],
                "slug": r[6],
                "price": float(r[7]) if r[7] else 0,
                "mrp": float(r[8]) if r[8] else None,
                "short_description": r[9],
                "collection_name": r[10],
                "primary_image": _get_r2_public_url(r[11]) if r[11] else "",
            }
        )
    return {"products": products}


@app.post("/api/v1/admin/landing/products", tags=["Landing Config"])
async def add_landing_product(
    data: LandingProductCreate,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    """Add a product to a landing section."""
    product_exists = db.execute(
        text("SELECT id FROM products WHERE id = :pid AND is_active = true"),
        {"pid": data.product_id},
    ).fetchone()
    if not product_exists:
        raise HTTPException(status_code=404, detail="Product not found or inactive")
    result = db.execute(
        text(
            "INSERT INTO landing_products (section, product_id, display_order, is_active) "
            "VALUES (:s, :pid, :order, :active) "
            "ON CONFLICT (section, product_id) DO UPDATE SET is_active = :active, display_order = :order "
            "RETURNING id"
        ),
        {
            "s": data.section,
            "pid": data.product_id,
            "order": data.display_order,
            "active": data.is_active,
        },
    )
    db.commit()
    redis_client.invalidate_pattern("public:landing:*")
    return {
        "message": "Product added to landing section",
        "landing_product_id": result.scalar(),
    }


@app.patch(
    "/api/v1/admin/landing/products/{landing_product_id}", tags=["Landing Config"]
)
async def update_landing_product(
    landing_product_id: str,
    data: LandingProductUpdate,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    """Update display order or active status of a landing product."""
    sets, params = [], {"id": landing_product_id}
    if data.display_order is not None:
        sets.append("display_order = :order")
        params["order"] = data.display_order
    if data.is_active is not None:
        sets.append("is_active = :active")
        params["active"] = data.is_active
    if not sets:
        return {"message": "Nothing to update"}
    db.execute(
        text(f"UPDATE landing_products SET {', '.join(sets)} WHERE id = :id"), params
    )
    db.commit()
    redis_client.invalidate_pattern("public:landing:*")
    return {"message": "Landing product updated"}


@app.delete(
    "/api/v1/admin/landing/products/{landing_product_id}", tags=["Landing Config"]
)
async def delete_landing_product(
    landing_product_id: str,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    """Remove a product from a landing section."""
    db.execute(
        text("DELETE FROM landing_products WHERE id = :id"), {"id": landing_product_id}
    )
    db.commit()
    redis_client.invalidate_pattern("public:landing:*")
    return {"message": "Product removed from landing section"}


@app.post("/api/v1/admin/landing/products/reorder", tags=["Landing Config"])
async def reorder_landing_products(
    section: str,
    ordered_ids: List[int],
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    """Reorder products in a landing section by providing ordered list of landing_product IDs."""
    for idx, lp_id in enumerate(ordered_ids):
        db.execute(
            text(
                "UPDATE landing_products SET display_order = :order WHERE id = :id AND section = :section"
            ),
            {"order": idx, "id": lp_id, "section": section},
        )
    db.commit()
    redis_client.invalidate_pattern("public:landing:*")
    return {"message": f"Reordered {len(ordered_ids)} products in section '{section}'"}


# ==================== Public Landing API (No Auth Required) ====================


@app.get("/api/v1/landing/config", tags=["Public Landing"])
async def get_public_landing_config(db: Session = Depends(get_db)):
    """Get landing page configuration for public display. No authentication required.

    Returns only active sections with their configuration.
    """
    cache_key = "public:landing:config"
    cached = redis_client.get_cache(cache_key)
    if cached:
        return cached

    rows = db.execute(
        text(
            "SELECT section, config, is_active FROM landing_config WHERE is_active = true ORDER BY section"
        )
    ).fetchall()

    sections = {}
    for r in rows:
        # Safely parse config - use json.loads instead of eval
        raw_config = r[1]
        if isinstance(raw_config, dict):
            config = raw_config
        elif raw_config is None:
            config = {}
        elif isinstance(raw_config, str):
            try:
                config = json.loads(raw_config)
            except json.JSONDecodeError:
                config = {}
        else:
            config = {}
        sections[r[0]] = {"config": config, "is_active": r[2]}

    result = {"sections": sections}
    redis_client.set_cache(cache_key, result, ttl=30)  # Cache for 30 seconds
    return result


@app.get("/api/v1/landing/images", tags=["Public Landing"])
async def get_public_landing_images(
    section: Optional[str] = None, db: Session = Depends(get_db)
):
    """Get landing page images for public display. No authentication required.

    Returns images for active sections only.
    """
    cache_key = f"public:landing:images:{section or 'all'}"
    cached = redis_client.get_cache(cache_key)
    if cached:
        return cached

    if section:
        # Check if section is active
        section_active = db.execute(
            text("SELECT is_active FROM landing_config WHERE section = :s"),
            {"s": section},
        ).fetchone()

        if not section_active or not section_active[0]:
            return {"images": []}

        rows = db.execute(
            text(
                "SELECT id, section, image_url, title, subtitle, link_url, display_order, device_variant "
                "FROM landing_images WHERE section = :s ORDER BY display_order"
            ),
            {"s": section},
        ).fetchall()
    else:
        # Get all images for active sections
        rows = db.execute(
            text("""
            SELECT li.id, li.section, li.image_url, li.title, li.subtitle, li.link_url, li.display_order, li.device_variant
            FROM landing_images li
            JOIN landing_config lc ON lc.section = li.section
            WHERE lc.is_active = true
            ORDER BY li.section, li.display_order
        """)
        ).fetchall()

    images = [
        {
            "id": r[0],
            "section": r[1],
            "image_url": r[2],
            "title": r[3],
            "subtitle": r[4],
            "link_url": r[5],
            "display_order": r[6],
            "device_variant": r[7],
        }
        for r in rows
    ]

    result = {"images": images}
    redis_client.set_cache(cache_key, result, ttl=30)  # Cache for 30 seconds
    return result


def _get_r2_public_url(image_url: str) -> str:
    """Convert image URL to full R2 public URL if needed.

    This ensures all image URLs returned to frontend are complete, ready-to-use URLs.
    Backend handles all R2 URL construction - frontend doesn't need to know about R2.
    """
    if not image_url:
        return ""

    # Already a full URL
    if image_url.startswith("http://") or image_url.startswith("https://"):
        return image_url

    # Construct R2 public URL from shared settings
    r2_base = settings.R2_PUBLIC_URL.rstrip("/")
    return f"{r2_base}/{image_url.lstrip('/')}"


def _get_default_landing_data() -> dict:
    """Get default landing page data when database is empty.

    Returns empty data - no fallbacks. Admin must configure everything.
    """
    return {
        "hero": {"tagline": "", "slides": [], "buttons": []},
        "newArrivals": {"title": "", "subtitle": "", "products": []},
        "collections": {"title": "", "categories": []},
        "about": {"title": "", "story": "", "stats": [], "images": []},
    }


@app.get("/api/v1/landing/all", tags=["Public Landing"])
async def get_public_landing_all(db: Session = Depends(get_db)):
    """Get all landing page data (config + images) in a single request. No authentication required.

    This is the main endpoint for the public landing page to fetch all data at once.
    Returns fully formatted, ready-to-use data - no frontend transformation needed.

    ARCHITECTURE: Database is the PRIMARY source of truth.
    - All data comes from database tables (landing_config, landing_images)
    - Hardcoded defaults are ONLY used as fallback when database is empty
    - All R2 URLs are constructed here
    - Frontend just displays the data as-is
    """
    cache_key = "public:landing:all"
    cached = redis_client.get_cache(cache_key)
    if cached:
        return cached

    # Get active sections config from DATABASE (PRIMARY SOURCE)
    config_rows = db.execute(
        text(
            "SELECT section, config, is_active FROM landing_config WHERE is_active = true ORDER BY section"
        )
    ).fetchall()

    sections = {}
    import json

    for r in config_rows:
        raw = r[1]
        if isinstance(raw, dict):
            config = raw
        elif raw is None:
            config = {}
        elif isinstance(raw, str):
            try:
                config = json.loads(raw)
            except Exception:
                config = {}
        else:
            config = raw
        sections[r[0]] = {"config": config, "is_active": r[2]}

    # Get images for active sections from DATABASE (PRIMARY SOURCE)
    image_rows = db.execute(
        text("""
        SELECT li.id, li.section, li.image_url, li.title, li.subtitle, li.link_url, li.display_order, li.device_variant
        FROM landing_images li
        JOIN landing_config lc ON lc.section = li.section
        WHERE lc.is_active = true
        ORDER BY li.section, li.display_order
    """)
    ).fetchall()

    # Check if database has any data
    has_database_data = bool(sections or image_rows)

    # DEBUG: Log what we found
    logger.info(
        f"Landing API - sections: {len(sections)}, image_rows: {len(image_rows)}, has_data: {has_database_data}"
    )
    if sections:
        logger.info(f"Available sections: {list(sections.keys())}")

    # ONLY use hardcoded defaults if database is COMPLETELY empty
    # This ensures database is always the PRIMARY source
    if not has_database_data:
        result = _get_default_landing_data()
        redis_client.set_cache(cache_key, result, ttl=30)
        return result

    # Build result from DATABASE (PRIMARY SOURCE)
    # If no data in database, return empty structure
    result = {
        "hero": {"tagline": "", "slides": [], "buttons": []},
        "newArrivals": {"title": "", "subtitle": "", "products": []},
        "collections": {"title": "", "categories": []},
        "about": {"title": "", "story": "", "stats": [], "images": []},
    }

    # Build hero section from database
    if "hero" in sections and sections["hero"].get("config"):
        hero_config = sections["hero"]["config"]
        result["hero"]["tagline"] = hero_config.get("tagline", "")

    # Default buttons are already set above

    # Build hero slides from database images
    # Group by display_order so desktop + mobile variants form ONE slide
    hero_images = [r for r in image_rows if r[1] == "hero"]
    if hero_images:
        hero_groups = {}
        for r in sorted(hero_images, key=lambda x: x[6]):
            order = r[6]  # display_order
            variant = (r[7] or "desktop").lower()  # device_variant column
            if order not in hero_groups:
                hero_groups[order] = {
                    "title": r[3],
                    "subtitle": r[4],
                    "link": r[5],
                }
            if variant in ("mobile", "phone"):
                hero_groups[order]["imageMobile"] = _get_r2_public_url(r[2])
            else:
                hero_groups[order]["image"] = _get_r2_public_url(r[2])
        result["hero"]["slides"] = list(hero_groups.values())

    # Pull buttons from hero config (admin-configurable)
    if "hero" in sections and sections["hero"].get("config"):
        hero_config = sections["hero"]["config"]
        # Handle both button array format and button1/button2 format
        if hero_config.get("buttons"):
            result["hero"]["buttons"] = hero_config["buttons"]
        elif hero_config.get("button1_text"):
            # Convert button1/button2 format to buttons array
            buttons = []
            if hero_config.get("button1_text") and hero_config.get("button1_link"):
                buttons.append(
                    {
                        "text": hero_config["button1_text"],
                        "link": hero_config["button1_link"],
                    }
                )
            if hero_config.get("button2_text") and hero_config.get("button2_link"):
                buttons.append(
                    {
                        "text": hero_config["button2_text"],
                        "link": hero_config["button2_link"],
                    }
                )
            result["hero"]["buttons"] = buttons

    # Build newArrivals section from database config + fetch admin-selected products
    if "newArrivals" in sections and sections["newArrivals"].get("config"):
        na_config = sections["newArrivals"]["config"]
        result["newArrivals"]["title"] = na_config.get("title", "New Arrivals")
        result["newArrivals"]["subtitle"] = na_config.get("subtitle", "")

    # Fetch new arrival products from landing_products table (admin-curated)
    try:
        na_limit = (sections.get("newArrivals", {}).get("config") or {}).get(
            "max_display", 8
        )
        product_rows = db.execute(
            text("""
            SELECT p.id, p.name, p.slug, p.base_price, p.mrp, p.short_description,
                   p.is_new_arrival, p.is_featured, p.is_active,
                   c.name AS collection_name, c.slug AS collection_slug,
                   pi.image_url AS primary_image
            FROM landing_products lp
            JOIN products p ON p.id = lp.product_id
            LEFT JOIN collections c ON p.category_id = c.id
            LEFT JOIN product_images pi ON p.id = pi.product_id AND pi.is_primary = true
            WHERE lp.section = 'newArrivals' 
              AND lp.is_active = true 
              AND p.is_active = true
            ORDER BY lp.display_order
            LIMIT :limit
        """),
            {"limit": na_limit},
        ).fetchall()
        result["newArrivals"]["products"] = [
            {
                "id": r[0],
                "name": r[1],
                "slug": r[2],
                "price": float(r[3]) if r[3] else 0,
                "mrp": float(r[4]) if r[4] else None,
                "short_description": r[5],
                "is_new_arrival": r[6],
                "is_featured": r[7],
                "collection_name": r[9],
                "collection_slug": r[10],
                "image_url": _get_r2_public_url(r[11]) if r[11] else "",
                "primary_image": _get_r2_public_url(r[11]) if r[11] else "",
            }
            for r in product_rows
        ]
    except Exception as e:
        logger.warning(f"Failed to fetch newArrivals products: {e}")
        result["newArrivals"]["products"] = []

    # Build collections section from database config + fetch featured categories
    if "collections" in sections and sections["collections"].get("config"):
        coll_config = sections["collections"]["config"]
        result["collections"]["title"] = coll_config.get("title", "Collections")

    # Fetch featured categories for collections section
    try:
        coll_limit = (sections.get("collections", {}).get("config") or {}).get(
            "max_display", 6
        )
        cat_rows = db.execute(
            text("""
            SELECT id, name, slug, description, image_url, is_active, is_featured
            FROM collections
            WHERE is_active = true AND is_featured = true
            ORDER BY display_order NULLS LAST, name
            LIMIT :limit
        """),
            {"limit": coll_limit},
        ).fetchall()
        result["collections"]["categories"] = [
            {
                "id": r[0],
                "name": r[1],
                "slug": r[2],
                "description": r[3],
                "image": _get_r2_public_url(r[4]) if r[4] else "",
                "image_url": _get_r2_public_url(r[4]) if r[4] else "",
                "link": f"/collections/{r[2]}",
                "is_active": r[5],
                "is_featured": r[6],
            }
            for r in cat_rows
        ]
    except Exception as e:
        logger.warning(f"Failed to fetch collections categories: {e}")
        result["collections"]["categories"] = []

    # Build about section from database
    if "about" in sections and sections["about"].get("config"):
        about_config = sections["about"]["config"]
        result["about"]["title"] = about_config.get("title", "")
        result["about"]["story"] = about_config.get(
            "description", about_config.get("story", "")
        )

        # Build stats from database if available
        if about_config.get("stats"):
            result["about"]["stats"] = about_config.get("stats")

    # Build about images from database
    about_images = [r for r in image_rows if r[1] == "about"]
    if about_images:
        result["about"]["images"] = [
            _get_r2_public_url(r[2]) for r in sorted(about_images, key=lambda x: x[6])
        ]

    redis_client.set_cache(cache_key, result, ttl=30)  # Cache for 30 seconds
    return result


@app.get("/api/v1/site/config", tags=["Public Landing"])
async def get_site_config(db: Session = Depends(get_db)):
    """Get site-wide configuration from database."""
    cache_key = "public:site:config"
    cached = redis_client.get_cache(cache_key)
    if cached:
        return cached

    rows = db.execute(text("SELECT key, value FROM site_config")).fetchall()
    config = {r[0]: r[1] for r in rows}

    r2_base = (
        settings.R2_PUBLIC_URL or "https://pub-7846c786f7154610b57735df47899fa0.r2.dev"
    )

    # Map database keys to expected response structure
    result = {
        "logo": config.get("logo_url") or f"{r2_base}/logo.png",
        "video": {
            "intro": config.get("intro_video_url")
            or f"{r2_base}/Create_a_video_202602141450_ub9p5.mp4"
        },
        "noise": f"{r2_base}/noise.png",
        "r2BaseUrl": r2_base,
        "site_name": config.get("site_name", "Aarya Clothing"),
        "contact_email": config.get("contact_email", ""),
        "contact_phone": config.get("contact_phone", ""),
        "free_shipping_threshold": float(config.get("free_shipping_threshold", 1000)),
        "intro_video_enabled": config.get("intro_video_enabled") == "true",
    }

    redis_client.set_cache(cache_key, result, ttl=3600)
    return result


# ==================== Presigned URL (Frontend Direct Upload) ====================


@app.post("/api/v1/admin/upload/presigned-url", tags=["Admin Upload"])
async def get_presigned_upload_url(
    filename: str,
    folder: str = Query(
        "landing", regex="^(landing|banners|categories|products|inventory|videos)$"
    ),
    content_type: str = Query("image/jpeg"),
    user: dict = Depends(require_admin),
):
    """Generate a presigned Cloudflare R2 URL for frontend direct upload.

    The frontend can PUT the file directly to the returned upload_url,
    then use the final_url when saving metadata.
    """
    return r2_service.generate_presigned_url(
        filename=filename, folder=folder, content_type=content_type
    )


@app.post("/api/v1/admin/upload/image", tags=["Admin Upload"])
async def upload_admin_image(
    file: UploadFile = File(...),
    folder: str = Query(
        "landing", regex="^(landing|banners|categories|products|inventory|videos)$"
    ),
    user: dict = Depends(require_admin),
):
    """Generic image upload endpoint for any admin-managed asset.

    Uploads the file to Cloudflare R2 and returns the public URL.
    Use the returned URL when creating/updating records.
    """
    image_url = await r2_service.upload_image(file, folder=folder)
    return {"image_url": image_url, "folder": folder}


@app.delete("/api/v1/admin/upload/image", tags=["Admin Upload"])
async def delete_admin_image(
    image_url: str,
    user: dict = Depends(require_admin),
):
    """Delete an image from Cloudflare R2 by its URL."""
    deleted = await r2_service.delete_image(image_url)
    return {"deleted": deleted, "image_url": image_url}


# ==================== Video Upload Endpoints ====================


@app.post("/api/v1/admin/upload/video", tags=["Admin Upload"])
async def upload_admin_video(
    file: UploadFile = File(...),
    folder: str = Query(
        "videos", regex="^(landing|banners|categories|products|inventory|videos)$"
    ),
    user: dict = Depends(require_admin),
):
    """Upload a video to Cloudflare R2.

    Supports MP4, WebM, and MOV formats. Max file size: 50MB.
    Uploads the file to Cloudflare R2 and returns the public URL.
    Use the returned URL when creating/updating records.
    """
    video_url = await r2_service.upload_video(file, folder=folder)
    return {"video_url": video_url, "folder": folder}


@app.delete("/api/v1/admin/upload/video", tags=["Admin Upload"])
async def delete_admin_video(
    video_url: str,
    user: dict = Depends(require_admin),
):
    """Delete a video from Cloudflare R2 by its URL."""
    deleted = await r2_service.delete_video(video_url)
    return {"deleted": deleted, "video_url": video_url}


@app.post("/api/v1/admin/landing/videos/upload", tags=["Admin Upload"])
async def upload_landing_video(
    file: UploadFile = File(...),
    device_variant: str = Query("desktop", regex="^(desktop|mobile)$"),
    user: dict = Depends(require_admin),
):
    """Upload intro video for landing page with device variant metadata.

    Supports separate desktop (16:9) and mobile (9:16) videos.
    Returns the video URL and device variant for saving to site config.
    """
    video_url = await r2_service.upload_video(file, folder="videos")
    return {
        "video_url": video_url,
        "device_variant": device_variant,
        "folder": "videos"
    }


# ==================== Quick Actions ====================


@app.get("/api/v1/staff/quick-actions", tags=["Staff"])
async def get_quick_actions():
    return {
        "actions": [
            {
                "name": "Add Stock",
                "endpoint": "/api/v1/staff/inventory/add-stock",
                "icon": "plus-box",
            },
            {
                "name": "Process Orders",
                "endpoint": "/api/v1/staff/orders/pending",
                "icon": "package",
            },
            {
                "name": "Low Stock Alert",
                "endpoint": "/api/v1/staff/inventory/low-stock",
                "icon": "alert",
            },
            {
                "name": "Stock Movement",
                "endpoint": "/api/v1/staff/inventory/movements",
                "icon": "chart-line",
            },
        ]
    }


# ==================== Product Performance Analytics ====================


@app.get("/api/v1/admin/analytics/products/performance", tags=["Analytics"])
async def get_product_performance(
    period: str = Query("30d", regex="^(7d|30d|90d|1y)$"),
    limit: int = Query(20, le=100),
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    """Detailed product performance: views, orders, revenue, conversion."""
    days = {"7d": 7, "30d": 30, "90d": 90, "1y": 365}[period]
    since = datetime.now(timezone.utc) - timedelta(days=days)
    rows = db.execute(
        text("""
        SELECT p.id, p.name, 
               MAX(i.sku) as sku, 
               p.base_price as price, 
               SUM(COALESCE(i.quantity, 0)) as total_stock,
               COALESCE(s.total_sold, 0) as total_sold,
               COALESCE(s.total_revenue, 0) as total_revenue,
               COALESCE(s.order_count, 0) as order_count,
               COALESCE(r.avg_rating, 0) as avg_rating,
               COALESCE(r.review_count, 0) as review_count
        FROM products p
        LEFT JOIN inventory i ON i.product_id = p.id
        LEFT JOIN (
            SELECT p.id as product_id,
                   SUM(oi.quantity) as total_sold,
                   SUM(oi.unit_price * oi.quantity) as total_revenue,
                   COUNT(DISTINCT oi.order_id) as order_count
            FROM order_items oi
            JOIN inventory inv ON inv.id = oi.inventory_id
            JOIN products p ON p.id = inv.product_id
            JOIN orders o ON o.id = oi.order_id
            WHERE o.created_at >= :since AND o.status != 'cancelled'
            GROUP BY p.id
        ) s ON s.product_id = p.id
        LEFT JOIN (
            SELECT product_id, AVG(rating) as avg_rating, COUNT(*) as review_count
            FROM reviews WHERE is_approved = true GROUP BY product_id
        ) r ON r.product_id = p.id
        WHERE p.is_active = true
        GROUP BY p.id, p.name, p.base_price, s.total_sold, s.total_revenue, s.order_count, r.avg_rating, r.review_count
        ORDER BY total_revenue DESC LIMIT :lim
    """),
        {"since": since, "lim": limit},
    ).fetchall()

    products = [
        {
            "id": r[0],
            "name": r[1],
            "sku": r[2],
            "price": float(r[3]),
            "stock": r[4],
            "total_sold": r[5],
            "revenue": float(r[6]),
            "order_count": r[7],
            "avg_rating": round(float(r[8]), 1),
            "review_count": r[9],
        }
        for r in rows
    ]

    return {"products": products, "period": period, "total": len(products)}


@app.get("/api/v1/admin/analytics/customers/detailed", tags=["Analytics"])
async def get_detailed_customer_analytics(
    period: str = Query("30d", regex="^(7d|30d|90d|1y)$"),
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    """Detailed customer analytics: LTV, segments, acquisition trends."""
    days = {"7d": 7, "30d": 30, "90d": 90, "1y": 365}[period]
    since = datetime.now(timezone.utc) - timedelta(days=days)

    # Registration trend by day
    reg_trend = db.execute(
        text("""
        SELECT DATE(created_at) as day, COUNT(*) as count
        FROM users WHERE role = 'customer' AND created_at >= :since
        GROUP BY DATE(created_at) ORDER BY day
    """),
        {"since": since},
    ).fetchall()

    # Top customers by spend
    top_customers = db.execute(
        text("""
        SELECT u.id, u.username, u.email, COUNT(o.id) as orders, SUM(o.total_amount) as total_spent
        FROM users u JOIN orders o ON o.user_id = u.id
        WHERE o.status != 'cancelled' AND o.created_at >= :since
        GROUP BY u.id, u.username, u.email
        ORDER BY total_spent DESC LIMIT 10
    """),
        {"since": since},
    ).fetchall()

    # Customer segments
    avg_ltv = (
        db.execute(
            text("""
        SELECT AVG(total) FROM (
            SELECT SUM(total_amount) as total FROM orders
            WHERE status != 'cancelled'
            GROUP BY user_id
        ) sub
    """)
        ).scalar()
        or 0
    )

    high_value = (
        db.execute(
            text("""
        SELECT COUNT(DISTINCT user_id) FROM orders
        WHERE status != 'cancelled'
        GROUP BY user_id HAVING SUM(total_amount) > :threshold
    """),
            {"threshold": float(avg_ltv) * 2},
        ).scalar()
        or 0
    )

    return {
        "registration_trend": [{"date": str(r[0]), "count": r[1]} for r in reg_trend],
        "top_customers": [
            {
                "id": r[0],
                "username": r[1],
                "email": r[2],
                "orders": r[3],
                "total_spent": float(r[4]),
            }
            for r in top_customers
        ],
        "average_ltv": round(float(avg_ltv), 2),
        "high_value_customers": high_value,
        "period": period,
    }


# ==================== Users/Customers Management ====================


@app.get("/api/v1/admin/users", response_model=List[UserListItem], tags=["Admin Users"])
async def list_users(
    role: Optional[str] = Query(None, regex="^(admin|staff|customer)$"),
    is_active: Optional[bool] = Query(None),
    search: Optional[str] = Query(None),
    sort_by: Optional[str] = Query(None),
    sort_order: str = Query("desc", regex="^(asc|desc)$"),
    limit: int = Query(50, ge=1, le=100),
    skip: int = Query(0, ge=0),
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    """List users with optional filtering by role, status, and search."""
    # Build WHERE clause and parameters
    where_conditions = []
    params = {"limit": limit, "skip": skip}

    if role:
        where_conditions.append("u.role = :role")
        params["role"] = role

    if is_active is not None:
        where_conditions.append("u.is_active = :is_active")
        params["is_active"] = is_active

    if search:
        where_conditions.append(
            "(u.email ILIKE :search OR u.username ILIKE :search OR up.full_name ILIKE :search)"
        )
        params["search"] = f"%{search}%"

    where_clause = "WHERE " + " AND ".join(where_conditions) if where_conditions else ""
    sort_columns = {
        "full_name": "COALESCE(up.full_name, '')",
        "email": "u.email",
        "order_count": "COALESCE(order_stats.order_count, 0)",
        "total_spent": "COALESCE(order_stats.total_spent, 0)",
        "is_active": "u.is_active",
        "created_at": "u.created_at",
    }
    order_column = sort_columns.get(sort_by, "u.created_at")
    order_direction = "ASC" if sort_order == "asc" else "DESC"
    order_clause = f"ORDER BY {order_column} {order_direction}, u.id DESC"

    # Main query with order statistics
    query = f"""
        SELECT 
            u.id,
            u.email,
            u.username,
            COALESCE(up.full_name, '') as full_name,
            COALESCE(up.phone, '') as phone,
            u.role,
            u.is_active,
            u.created_at,
            COALESCE(order_stats.order_count, 0) as order_count,
            COALESCE(order_stats.total_spent, 0) as total_spent
        FROM users u
        LEFT JOIN user_profiles up ON u.id = up.user_id
        LEFT JOIN (
            SELECT 
                user_id,
                COUNT(*) as order_count,
                COALESCE(SUM(total_amount), 0) as total_spent
            FROM orders 
            WHERE status != 'cancelled'
            GROUP BY user_id
        ) order_stats ON u.id = order_stats.user_id
        {where_clause}
        {order_clause}
        LIMIT :limit OFFSET :skip
    """

    rows = db.execute(text(query), params).fetchall()

    users = []
    for row in rows:
        users.append(
            {
                "id": row[0],
                "email": row[1],
                "username": row[2],
                "full_name": row[3],
                "phone": row[4],
                "role": row[5],
                "is_active": row[6],
                "created_at": row[7],
                "order_count": row[8],
                "total_spent": float(row[9]),
            }
        )

    return users


@app.get("/api/v1/admin/users/count", tags=["Admin Users"])
async def count_users(
    role: Optional[str] = Query(None, regex="^(admin|staff|customer)$"),
    is_active: Optional[bool] = Query(None),
    search: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    """Get total count of users with optional filtering."""
    where_conditions = []
    params = {}

    if role:
        where_conditions.append("u.role = :role")
        params["role"] = role

    if is_active is not None:
        where_conditions.append("u.is_active = :is_active")
        params["is_active"] = is_active

    if search:
        where_conditions.append(
            "(u.email ILIKE :search OR u.username ILIKE :search OR up.full_name ILIKE :search)"
        )
        params["search"] = f"%{search}%"

    where_clause = "WHERE " + " AND ".join(where_conditions) if where_conditions else ""

    query = f"""
        SELECT COUNT(*)
        FROM users u
        LEFT JOIN user_profiles up ON u.id = up.user_id
        {where_clause}
    """

    result = db.execute(text(query), params).scalar()
    return {"count": result}


@app.get(
    "/api/v1/admin/users/{user_id}", response_model=UserListItem, tags=["Admin Users"]
)
async def get_user(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_admin),
):
    """Get detailed information about a specific user."""
    query = """
        SELECT
            u.id,
            u.email,
            u.username,
            COALESCE(up.full_name, '') as full_name,
            COALESCE(up.phone, '') as phone,
            u.role,
            u.is_active,
            u.created_at,
            u.updated_at,
            COALESCE(order_stats.order_count, 0) as order_count,
            COALESCE(order_stats.total_spent, 0) as total_spent,
            order_stats.last_order_date
        FROM users u
        LEFT JOIN user_profiles up ON u.id = up.user_id
        LEFT JOIN (
            SELECT
                user_id,
                COUNT(*) as order_count,
                COALESCE(SUM(total_amount), 0) as total_spent,
                MAX(created_at) as last_order_date
            FROM orders
            WHERE status != 'cancelled'
            GROUP BY user_id
        ) order_stats ON u.id = order_stats.user_id
        WHERE u.id = :user_id
    """

    row = db.execute(text(query), {"user_id": user_id}).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="User not found")

    return {
        "id": row[0],
        "email": row[1],
        "username": row[2],
        "full_name": row[3],
        "phone": row[4],
        "role": row[5],
        "is_active": row[6],
        "created_at": row[7],
        "updated_at": str(row[8]) if row[8] else None,
        "order_count": row[9],
        "total_spent": float(row[10]),
        "last_order_date": str(row[11]) if row[11] else None,
    }


@app.put("/api/v1/admin/users/{user_id}/status", tags=["Admin Users"])
async def update_user_status(
    user_id: int,
    data: UserStatusUpdate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_admin),
):
    """Update user active status."""
    # Check if user exists
    user_exists = db.execute(
        text("SELECT id FROM users WHERE id = :user_id"), {"user_id": user_id}
    ).fetchone()
    if not user_exists:
        raise HTTPException(status_code=404, detail="User not found")

    # Update status
    db.execute(
        text(
            "UPDATE users SET is_active = :is_active, updated_at = :now WHERE id = :user_id"
        ),
        {
            "is_active": data.is_active,
            "now": datetime.now(timezone.utc),
            "user_id": user_id,
        },
    )
    db.commit()

    # Invalidate relevant cache
    redis_client.invalidate_pattern("admin:dashboard:*")
    redis_client.invalidate_pattern("admin:analytics:*")

    return {
        "message": f"User status updated to {'active' if data.is_active else 'inactive'}"
    }


@app.patch("/api/v1/admin/users/bulk-status", tags=["Admin Users"])
async def bulk_update_user_status(
    data: BulkUserStatusUpdate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_admin),
):
    """Bulk update user active status."""
    if not data.user_ids:
        raise HTTPException(status_code=400, detail="No user IDs provided")

    # Update all users in batch
    db.execute(
        text("""
            UPDATE users 
            SET is_active = :is_active, updated_at = :now 
            WHERE id = ANY(:user_ids)
        """),
        {
            "is_active": data.is_active,
            "now": datetime.now(timezone.utc),
            "user_ids": data.user_ids,
        },
    )
    db.commit()

    # Invalidate relevant cache
    redis_client.invalidate_pattern("admin:dashboard:*")
    redis_client.invalidate_pattern("admin:analytics:*")

    return {
        "message": f"Updated {len(data.user_ids)} users to {'active' if data.is_active else 'inactive'}"
    }


# ==================== Discount / Promotion Management ====================


@app.get("/api/v1/admin/discounts", tags=["Admin Discounts"])
async def list_discounts(
    is_active: Optional[bool] = None,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    """List all promotions/discounts."""
    where = ""
    if is_active is not None:
        where = f"WHERE is_active = {'true' if is_active else 'false'}"
    rows = db.execute(
        text(f"SELECT * FROM promotions {where} ORDER BY created_at DESC")
    ).fetchall()
    return {"discounts": [dict(r._mapping) for r in rows]}


@app.post("/api/v1/admin/discounts", tags=["Admin Discounts"])
async def create_discount(
    code: str,
    discount_type: str,
    discount_value: float,
    min_order_value: float = 0,
    max_discount_amount: Optional[float] = None,
    usage_limit: Optional[int] = None,
    valid_from: Optional[str] = None,
    valid_until: Optional[str] = None,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    """Create a new promotion/discount code."""
    result = db.execute(
        text("""
        INSERT INTO promotions (code, discount_type, discount_value, min_order_value,
            max_discount_amount, usage_limit, valid_from, valid_until, is_active)
        VALUES (:code, :dt, :dv, :mov, :mda, :ul, :vf, :vu, true) RETURNING id
    """),
        {
            "code": code.upper(),
            "dt": discount_type,
            "dv": discount_value,
            "mov": min_order_value,
            "mda": max_discount_amount,
            "ul": usage_limit,
            "vf": valid_from,
            "vu": valid_until,
        },
    )
    db.commit()
    return {
        "message": "Discount created",
        "discount_id": result.scalar(),
        "code": code.upper(),
    }


@app.put("/api/v1/admin/discounts/{discount_id}", tags=["Admin Discounts"])
async def update_discount(
    discount_id: int,
    is_active: Optional[bool] = None,
    usage_limit: Optional[int] = None,
    valid_until: Optional[str] = None,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    """Update a promotion/discount."""
    sets, params = (
        ["updated_at = :now"],
        {"id": discount_id, "now": datetime.now(timezone.utc)},
    )
    if is_active is not None:
        sets.append("is_active = :active")
        params["active"] = is_active
    if usage_limit is not None:
        sets.append("usage_limit = :ul")
        params["ul"] = usage_limit
    if valid_until is not None:
        sets.append("valid_until = :vu")
        params["vu"] = valid_until
    db.execute(text(f"UPDATE promotions SET {', '.join(sets)} WHERE id = :id"), params)
    db.commit()
    return {"message": "Discount updated"}


@app.delete("/api/v1/admin/discounts/{discount_id}", tags=["Admin Discounts"])
async def delete_discount(
    discount_id: int, db: Session = Depends(get_db), user: dict = Depends(require_admin)
):
    """Delete a discount code."""
    db.execute(text("DELETE FROM promotions WHERE id = :id"), {"id": discount_id})
    db.commit()
    return {"message": "Discount deleted"}


# ==================== Product Management (Admin) ====================


@app.get("/api/v1/admin/products", tags=["Admin Products"])
async def admin_list_products(
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=200),
    collection_id: Optional[int] = None,
    active_only: Optional[bool] = None,
    new_arrivals: Optional[bool] = None,
    featured: Optional[bool] = None,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    """List all products for admin (includes inactive)."""
    where_parts = []
    params = {"limit": limit, "skip": skip}
    if active_only is True:
        where_parts.append("p.is_active = true")
    elif active_only is False:
        where_parts.append("p.is_active = false")
    if collection_id is not None:
        where_parts.append("p.category_id = :cid")
        params["cid"] = collection_id
    if new_arrivals is True:
        where_parts.append("p.is_new_arrival = true")
    if featured is True:
        where_parts.append("p.is_featured = true")
    where_clause = ("WHERE " + " AND ".join(where_parts)) if where_parts else ""
    rows = db.execute(
        text(f"""
        SELECT p.id, p.name, p.slug, p.base_price, p.mrp, p.short_description,
               p.description, p.is_active, p.is_featured, p.is_new_arrival,
               p.category_id, c.name as collection_name,
               p.meta_title, p.meta_description, p.created_at, p.updated_at,
               COALESCE(SUM(i.quantity), 0) as total_stock,
               pi.image_url as primary_image
        FROM products p
        LEFT JOIN collections c ON p.category_id = c.id
        LEFT JOIN inventory i ON i.product_id = p.id
        LEFT JOIN product_images pi ON p.id = pi.product_id AND pi.is_primary = true
        {where_clause}
        GROUP BY p.id, p.name, p.slug, p.base_price, p.mrp, p.short_description,
                 p.description, p.is_active, p.is_featured, p.is_new_arrival,
                 p.category_id, c.name, p.meta_title, p.meta_description,
                 p.created_at, p.updated_at, pi.image_url
        ORDER BY p.created_at DESC
        LIMIT :limit OFFSET :skip
    """),
        params,
    ).fetchall()
    products = []
    for r in rows:
        img = r[17]
        products.append(
            {
                "id": r[0],
                "name": r[1],
                "slug": r[2],
                "price": float(r[3]) if r[3] else 0,
                "mrp": float(r[4]) if r[4] else None,
                "short_description": r[5],
                "description": r[6],
                "is_active": r[7],
                "is_featured": r[8],
                "is_new_arrival": r[9],
                "category_id": r[10],
                "collection_id": r[10],
                "collection_name": r[11],
                "meta_title": r[12],
                "meta_description": r[13],
                "created_at": r[14],
                "updated_at": r[15],
                "total_stock": int(r[16]),
                "image_url": _get_r2_public_url(img) if img else None,
                "primary_image": _get_r2_public_url(img) if img else None,
            }
        )
    return products


@app.post("/api/v1/admin/products", status_code=201, tags=["Admin Products"])
async def admin_create_product(
    data: ProductCreate,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    """Create a new product (admin only)."""
    # Resolve collection_id -> category_id
    category_id = data.collection_id or data.category_id
    # Auto-generate slug if not provided
    slug = data.slug
    if not slug:
        slug = re.sub(r"[^a-z0-9]+", "-", data.name.lower()).strip("-")
    # Ensure slug uniqueness
    existing = db.execute(
        text("SELECT id FROM products WHERE slug = :slug"), {"slug": slug}
    ).fetchone()
    if existing:
        slug = f"{slug}-{int(datetime.now(timezone.utc).timestamp())}"
    result = db.execute(
        text("""
        INSERT INTO products (name, slug, description, short_description, base_price, mrp,
            category_id, brand, is_active, is_featured, is_new_arrival,
            meta_title, meta_description, created_at, updated_at)
        VALUES (:name, :slug, :desc, :short_desc, :price, :mrp,
            :cat_id, :brand, :active, :featured, :new_arrival,
            :meta_title, :meta_desc, :now, :now)
        RETURNING id
    """),
        {
            "name": data.name,
            "slug": slug,
            "desc": data.description,
            "short_desc": data.short_description,
            "price": data.base_price,
            "mrp": data.mrp,
            "cat_id": category_id,
            "brand": getattr(data, "brand", None),
            "active": data.is_active,
            "featured": data.is_featured,
            "new_arrival": data.is_new_arrival,
            "meta_title": data.meta_title,
            "meta_desc": data.meta_description,
            "now": datetime.now(timezone.utc),
        },
    )
    product_id = result.scalar()
    db.commit()
    # NOTE: No default inventory is created here. Variants must be added separately
    # via POST /api/v1/admin/products/{product_id}/variants. A product without
    # inventory variants will not be visible to customers.
    redis_client.invalidate_pattern("products:*")
    redis_client.invalidate_pattern("public:landing:*")
    return {
        "id": product_id,
        "name": data.name,
        "slug": slug,
        "message": "Product created. Add at least one variant (size/color) to make it visible to customers.",
    }


@app.patch("/api/v1/admin/products/{product_id}", tags=["Admin Products"])
async def admin_update_product(
    product_id: str,
    data: ProductUpdate,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    """Update a product (admin only)."""
    # Resolve slug or ID
    resolved = db.execute(
        text("SELECT id FROM products WHERE id::text = :pid OR slug = :pid"),
        {"pid": product_id},
    ).fetchone()
    if not resolved:
        raise HTTPException(status_code=404, detail="Product not found")
    pid = resolved[0]
    sets, params = (
        ["updated_at = :now"],
        {"id": pid, "now": datetime.now(timezone.utc)},
    )
    update_map = {
        "name": "name",
        "slug": "slug",
        "description": "description",
        "short_description": "short_description",
        "is_active": "is_active",
        "is_featured": "is_featured",
        "is_new_arrival": "is_new_arrival",
        "meta_title": "meta_title",
        "meta_description": "meta_description",
    }
    for field, col in update_map.items():
        val = getattr(data, field, None)
        if val is not None:
            sets.append(f"{col} = :{field}")
            params[field] = val
    if data.base_price is not None:
        sets.append("base_price = :base_price")
        params["base_price"] = data.base_price
    if data.mrp is not None:
        sets.append("mrp = :mrp")
        params["mrp"] = data.mrp
    cat_id = getattr(data, "collection_id", None) or getattr(data, "category_id", None)
    if cat_id is not None:
        sets.append("category_id = :cat_id")
        params["cat_id"] = cat_id
    result = db.execute(
        text(f"UPDATE products SET {', '.join(sets)} WHERE id = :id RETURNING id"),
        params,
    )
    if not result.fetchone():
        raise HTTPException(status_code=404, detail="Product not found")
    db.commit()
    redis_client.invalidate_pattern("products:*")
    redis_client.invalidate_pattern("public:landing:*")
    return {"message": "Product updated", "id": product_id}


@app.delete(
    "/api/v1/admin/products/{product_id}", status_code=204, tags=["Admin Products"]
)
async def admin_delete_product(
    product_id: str, db: Session = Depends(get_db), user: dict = Depends(require_admin)
):
    """Delete a product and its R2 images (admin only)."""
    # Resolve slug or ID
    resolved = db.execute(
        text("SELECT id FROM products WHERE id::text = :pid OR slug = :pid"),
        {"pid": product_id},
    ).fetchone()
    if not resolved:
        raise HTTPException(status_code=404, detail="Product not found")
    pid = resolved[0]
    # Delete R2 images first
    images = db.execute(
        text("SELECT image_url FROM product_images WHERE product_id = :pid"),
        {"pid": pid},
    ).fetchall()
    for img in images:
        if img[0]:
            await r2_service.delete_image(img[0])
    db.execute(
        text("DELETE FROM product_images WHERE product_id = :pid"), {"pid": pid}
    )
    result = db.execute(
        text("DELETE FROM products WHERE id = :id RETURNING id"), {"id": pid}
    )
    if not result.fetchone():
        raise HTTPException(status_code=404, detail="Product not found")
    db.commit()
    redis_client.invalidate_pattern("products:*")
    redis_client.invalidate_pattern("public:landing:*")


# ==================== Product Bulk Operations (Admin) ====================


@app.post("/api/v1/admin/products/bulk/price", tags=["Admin Products"])
async def admin_bulk_price_update(
    data: BulkPriceUpdate,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    """Bulk price update for products."""
    from decimal import Decimal

    products = db.execute(
        text("SELECT id, base_price, mrp FROM products WHERE id = ANY(:ids)"),
        {"ids": data.product_ids},
    ).fetchall()
    updated = 0
    for p in products:
        sets, params = (
            ["updated_at = :now"],
            {"id": p[0], "now": datetime.now(timezone.utc)},
        )
        price = Decimal(str(p[1])) if p[1] else Decimal("0")
        mrp = Decimal(str(p[2])) if p[2] else None
        if data.price is not None:
            sets.append("base_price = :price")
            params["price"] = data.price
        if data.mrp is not None:
            sets.append("mrp = :mrp")
            params["mrp"] = data.mrp
        if data.price_adjustment is not None:
            new_price = max(
                Decimal("0.01"), price + Decimal(str(data.price_adjustment))
            )
            sets.append("base_price = :price")
            params["price"] = float(new_price)
        if data.price_percentage is not None:
            factor = Decimal(str(1 + data.price_percentage / 100))
            new_price = max(Decimal("0.01"), price * factor)
            sets.append("base_price = :price")
            params["price"] = float(new_price)
        if data.mrp_adjustment is not None and mrp is not None:
            new_mrp = max(Decimal("0"), mrp + Decimal(str(data.mrp_adjustment)))
            sets.append("mrp = :mrp")
            params["mrp"] = float(new_mrp)
        if data.mrp_percentage is not None and mrp is not None:
            factor = Decimal(str(1 + data.mrp_percentage / 100))
            new_mrp = max(Decimal("0"), mrp * factor)
            sets.append("mrp = :mrp")
            params["mrp"] = float(new_mrp)
        if sets:
            db.execute(
                text(f"UPDATE products SET {', '.join(sets)} WHERE id = :id"), params
            )
            updated += 1
    db.commit()
    redis_client.invalidate_pattern("products:*")
    return {"updated": updated}


@app.post("/api/v1/admin/products/bulk/status", tags=["Admin Products"])
async def admin_bulk_status_update(
    data: BulkStatusUpdate,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    """Bulk activate/deactivate/feature products."""
    sets, params = (
        ["updated_at = :now"],
        {"ids": data.product_ids, "now": datetime.now(timezone.utc)},
    )
    if data.is_active is not None:
        sets.append("is_active = :active")
        params["active"] = data.is_active
    if data.is_featured is not None:
        sets.append("is_featured = :featured")
        params["featured"] = data.is_featured
    if data.is_new_arrival is not None:
        sets.append("is_new_arrival = :new_arrival")
        params["new_arrival"] = data.is_new_arrival
    db.execute(
        text(f"UPDATE products SET {', '.join(sets)} WHERE id = ANY(:ids)"), params
    )
    db.commit()
    redis_client.invalidate_pattern("products:*")
    redis_client.invalidate_pattern("public:landing:*")
    return {"updated": len(data.product_ids)}


@app.post("/api/v1/admin/products/bulk/collection", tags=["Admin Products"])
async def admin_bulk_assign_collection(
    data: BulkCollectionAssign,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    """Bulk assign products to a collection."""
    coll = db.execute(
        text("SELECT id FROM collections WHERE id = :id"), {"id": data.collection_id}
    ).fetchone()
    if not coll:
        raise HTTPException(status_code=404, detail="Collection not found")
    db.execute(
        text(
            "UPDATE products SET category_id = :cid, updated_at = :now WHERE id = ANY(:ids)"
        ),
        {
            "cid": data.collection_id,
            "ids": data.product_ids,
            "now": datetime.now(timezone.utc),
        },
    )
    db.commit()
    redis_client.invalidate_pattern("products:*")
    return {"updated": len(data.product_ids), "collection_id": data.collection_id}


@app.post("/api/v1/admin/products/bulk/inventory", tags=["Admin Products"])
async def admin_bulk_inventory_update(
    data: dict, db: Session = Depends(get_db), user: dict = Depends(require_admin)
):
    """Bulk update inventory by SKU."""
    updates = data.get("updates", [])
    updated, errors = 0, []
    for item in updates:
        sku = item.get("sku")
        qty = item.get("quantity")
        if not sku or qty is None:
            errors.append({"sku": sku, "error": "Missing sku or quantity"})
            continue
        result = db.execute(
            text(
                "UPDATE inventory SET quantity = :qty, updated_at = :now WHERE sku = :sku"
            ),
            {"qty": qty, "sku": sku, "now": datetime.now(timezone.utc)},
        )
        if result.rowcount:
            updated += 1
        else:
            errors.append({"sku": sku, "error": "SKU not found"})
    db.commit()
    redis_client.invalidate_pattern("products:*")
    return {"updated": updated, "errors": errors}


@app.post("/api/v1/admin/products/bulk/delete", tags=["Admin Products"])
async def admin_bulk_delete_products(
    data: dict, db: Session = Depends(get_db), user: dict = Depends(require_admin)
):
    """Bulk delete products."""
    product_ids = data.get("product_ids", [])
    deleted = 0
    for pid in product_ids:
        images = db.execute(
            text("SELECT image_url FROM product_images WHERE product_id = :pid"),
            {"pid": pid},
        ).fetchall()
        for img in images:
            if img[0]:
                await r2_service.delete_image(img[0])
        db.execute(
            text("DELETE FROM product_images WHERE product_id = :pid"), {"pid": pid}
        )
        result = db.execute(
            text("DELETE FROM products WHERE id = :id RETURNING id"), {"id": pid}
        )
        if result.fetchone():
            deleted += 1
    db.commit()
    redis_client.invalidate_pattern("products:*")
    redis_client.invalidate_pattern("public:landing:*")
    return {"deleted": deleted}


# ==================== Product Image Management (Admin) ====================


@app.post(
    "/api/v1/admin/products/{product_id}/images",
    status_code=201,
    tags=["Admin Products"],
)
async def admin_upload_product_image(
    product_id: str,
    file: UploadFile = File(...),
    alt_text: Optional[str] = None,
    is_primary: bool = False,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    """Upload a product image to R2 (admin only)."""
    product = db.execute(
        text("SELECT id, name FROM products WHERE id::text = :pid OR slug = :pid"),
        {"pid": product_id},
    ).fetchone()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    resolved_product_id = int(product[0])
    full_url = await r2_service.upload_image(file, folder="products")
    r2_base = settings.R2_PUBLIC_URL.rstrip("/") if settings.R2_PUBLIC_URL else ""
    relative_path = (
        full_url.replace(r2_base + "/", "")
        if r2_base and full_url.startswith(r2_base)
        else full_url
    )
    if is_primary:
        db.execute(
            text(
                "UPDATE product_images SET is_primary = false WHERE product_id = :pid"
            ),
            {"pid": resolved_product_id},
        )
    img_count = (
        db.execute(
            text("SELECT COUNT(*) FROM product_images WHERE product_id = :pid"),
            {"pid": resolved_product_id},
        ).scalar()
        or 0
    )
    result = db.execute(
        text("""
        INSERT INTO product_images (product_id, image_url, alt_text, is_primary, display_order, created_at)
        VALUES (:pid, :url, :alt, :primary, :order, :now) RETURNING id
    """),
        {
            "pid": resolved_product_id,
            "url": relative_path,
            "alt": alt_text or f"{product[1]} - Image {img_count + 1}",
            "primary": is_primary,
            "order": img_count,
            "now": datetime.now(timezone.utc),
        },
    )
    image_id = result.scalar()
    db.commit()
    redis_client.invalidate_pattern("products:*")
    redis_client.invalidate_pattern("public:landing:*")
    return {
        "id": image_id,
        "product_id": resolved_product_id,
        "image_url": _get_r2_public_url(relative_path),
        "alt_text": alt_text or f"{product[1]} - Image {img_count + 1}",
        "is_primary": is_primary,
        "display_order": img_count,
    }


@app.delete(
    "/api/v1/admin/products/{product_id}/images/{image_id}",
    status_code=204,
    tags=["Admin Products"],
)
async def admin_delete_product_image(
    product_id: str,
    image_id: int,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    """Delete a product image from R2 and DB (admin only)."""
    product = db.execute(
        text("SELECT id FROM products WHERE id::text = :pid OR slug = :pid"),
        {"pid": product_id},
    ).fetchone()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    resolved_product_id = int(product[0])

    img = db.execute(
        text(
            "SELECT image_url FROM product_images WHERE id = :id AND product_id = :pid"
        ),
        {"id": image_id, "pid": resolved_product_id},
    ).fetchone()
    if not img:
        raise HTTPException(status_code=404, detail="Image not found")
    if img[0]:
        await r2_service.delete_image(img[0])
    db.execute(text("DELETE FROM product_images WHERE id = :id"), {"id": image_id})
    db.commit()
    redis_client.invalidate_pattern("products:*")
    redis_client.invalidate_pattern("public:landing:*")


@app.patch(
    "/api/v1/admin/products/{product_id}/images/reorder", tags=["Admin Products"]
)
async def admin_reorder_product_images(
    product_id: str,
    image_ids: List[int],
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    """Reorder product images by providing new display order (list of image IDs in desired order)."""
    product = db.execute(
        text("SELECT id FROM products WHERE id::text = :pid OR slug = :pid"),
        {"pid": product_id},
    ).fetchone()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    resolved_product_id = int(product[0])
    for idx, img_id in enumerate(image_ids):
        db.execute(
            text(
                "UPDATE product_images SET display_order = :order WHERE id = :id AND product_id = :pid"
            ),
            {"order": idx, "id": img_id, "pid": resolved_product_id},
        )
    db.commit()
    redis_client.invalidate_pattern("products:*")
    redis_client.invalidate_pattern("public:landing:*")
    return {"message": "Images reordered", "order": image_ids}


@app.patch(
    "/api/v1/admin/products/{product_id}/images/{image_id}/primary",
    tags=["Admin Products"],
)
async def admin_set_primary_image(
    product_id: str,
    image_id: int,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    """Set a specific image as the primary image for a product."""
    product = db.execute(
        text("SELECT id FROM products WHERE id::text = :pid OR slug = :pid"),
        {"pid": product_id},
    ).fetchone()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    resolved_product_id = int(product[0])

    img = db.execute(
        text("SELECT id FROM product_images WHERE id = :id AND product_id = :pid"),
        {"id": image_id, "pid": resolved_product_id},
    ).fetchone()
    if not img:
        raise HTTPException(status_code=404, detail="Image not found")
    db.execute(
        text("UPDATE product_images SET is_primary = false WHERE product_id = :pid"),
        {"pid": resolved_product_id},
    )
    db.execute(
        text("UPDATE product_images SET is_primary = true WHERE id = :id"),
        {"id": image_id},
    )
    db.commit()
    redis_client.invalidate_pattern("products:*")
    redis_client.invalidate_pattern("public:landing:*")
    return {"message": "Primary image updated", "image_id": image_id}


@app.get("/api/v1/admin/products/{product_id}", tags=["Admin Products"])
async def admin_get_product(
    product_id: str, db: Session = Depends(get_db), user: dict = Depends(require_admin)
):
    """Get a single product by ID or slug with images and inventory (admin only)."""
    logger = logging.getLogger(__name__)

    # Sanitize user input to prevent log injection attacks
    import re
    safe_sub = re.sub(r'[\n\r\t]', '', str(user.get('sub', 'unknown')))
    logger.info(f"[AdminProduct] Fetching product ID/slug={product_id} for user={safe_sub}")

    # Resolve slug or ID
    resolved = db.execute(
        text("SELECT id FROM products WHERE id::text = :pid OR slug = :pid"),
        {"pid": product_id},
    ).fetchone()
    if not resolved:
        raise HTTPException(status_code=404, detail="Product not found")
    pid = resolved[0]

    row = db.execute(
        text("""
        SELECT p.id, p.name, p.slug, p.base_price, p.mrp, p.short_description,
               p.description, p.is_active, p.is_featured, p.is_new_arrival,
               p.category_id, c.name as collection_name, p.brand,
               p.meta_title, p.meta_description, p.created_at, p.updated_at,
               COALESCE(SUM(i.quantity), 0) as total_stock
        FROM products p
        LEFT JOIN collections c ON p.category_id = c.id
        LEFT JOIN inventory i ON i.product_id = p.id
        WHERE p.id = :id
        GROUP BY p.id, p.name, p.slug, p.base_price, p.mrp, p.short_description,
                 p.description, p.is_active, p.is_featured, p.is_new_arrival,
                 p.category_id, c.name, p.brand, p.meta_title, p.meta_description,
                 p.created_at, p.updated_at
        """),
        {"id": pid},
    ).fetchone()
    
    if not row:
        logger.warning(f"[AdminProduct] Product ID={product_id} not found in database")
        raise HTTPException(status_code=404, detail="Product not found")
    
    logger.info(f"[AdminProduct] Product found: {row[1]} (ID={row[0]})")
    
    images = db.execute(
        text("SELECT id, image_url, alt_text, is_primary, display_order FROM product_images WHERE product_id = :pid ORDER BY is_primary DESC, display_order"),
        {"pid": product_id},
    ).fetchall()
    inventory = db.execute(
        text("SELECT * FROM inventory WHERE product_id = :pid ORDER BY size, color"),
        {"pid": product_id},
    ).fetchall()
    r2_base = settings.R2_PUBLIC_URL.rstrip("/") if settings.R2_PUBLIC_URL else ""

    def full_url(path):
        if not path:
            return None
        if path.startswith("http://") or path.startswith("https://"):
            return path
        return f"{r2_base}/{path.lstrip('/')}" if r2_base else path

    primary_img = next((full_url(img[1]) for img in images if img[3]), None) or (full_url(images[0][1]) if images else None)
    logger.info(f"[AdminProduct] Returning product with {len(images)} images and {len(inventory)} variants")
    
    return {
        "id": row[0],
        "name": row[1],
        "slug": row[2],
        "price": float(row[3]) if row[3] else 0,
        "base_price": float(row[3]) if row[3] else 0,
        "mrp": float(row[4]) if row[4] else None,
        "short_description": row[5],
        "description": row[6],
        "is_active": row[7],
        "is_featured": row[8],
        "is_new_arrival": row[9],
        "category_id": row[10],
        "collection_id": row[10],
        "collection_name": row[11],
        "brand": row[12],
        "meta_title": row[13],
        "meta_description": row[14],
        "created_at": row[15],
        "updated_at": row[16],
        "total_stock": int(row[17]),
        "image_url": primary_img,
        "primary_image": primary_img,
        "images": [
            {"id": img[0], "image_url": full_url(img[1]), "alt_text": img[2], "is_primary": img[3], "display_order": img[4]}
            for img in images
        ],
        "inventory": [dict(inv._mapping) for inv in inventory],
    }


@app.get("/api/v1/admin/products/{product_id}/variants", tags=["Admin Products"])
async def admin_get_product_variants(
    product_id: str, db: Session = Depends(get_db), user: dict = Depends(require_staff)
):
    """Get all inventory variants for a product."""
    resolved = db.execute(
        text("SELECT id FROM products WHERE id::text = :pid OR slug = :pid"),
        {"pid": product_id},
    ).fetchone()
    if not resolved:
        raise HTTPException(status_code=404, detail="Product not found")
    pid = resolved[0]
    rows = db.execute(
        text("SELECT * FROM inventory WHERE product_id = :pid ORDER BY size, color"),
        {"pid": pid},
    ).fetchall()
    return {"variants": [dict(r._mapping) for r in rows]}


@app.post(
    "/api/v1/admin/products/{product_id}/variants",
    status_code=201,
    tags=["Admin Products"],
)
async def admin_create_product_variant(
    product_id: str,
    data: VariantCreate,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    """Create an inventory variant for a product."""
    # Resolve slug or ID
    product = db.execute(
        text("SELECT id FROM products WHERE id::text = :pid OR slug = :pid"),
        {"pid": product_id},
    ).fetchone()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    pid = product[0]
    existing = db.execute(
        text("SELECT id FROM inventory WHERE product_id = :pid AND sku = :sku"),
        {"pid": pid, "sku": data.sku},
    ).fetchone()
    if existing:
        raise HTTPException(
            status_code=400, detail=f"SKU '{data.sku}' already exists for this product"
        )
    result = db.execute(
        text("""
        INSERT INTO inventory (product_id, sku, size, color, color_hex, quantity, low_stock_threshold, created_at, updated_at)
        VALUES (:pid, :sku, :size, :color, :color_hex, :qty, 5, :now, :now) RETURNING id
    """),
        {
            "pid": pid,
            "sku": data.sku,
            "size": data.size,
            "color": data.color,
            "color_hex": data.color_hex,
            "qty": data.quantity,
            "now": datetime.now(timezone.utc),
        },
    )
    inv_id = result.scalar()
    db.commit()
    redis_client.invalidate_pattern("products:*")
    return {"id": inv_id, "message": "Variant created", "sku": data.sku}


@app.patch(
    "/api/v1/admin/products/{product_id}/variants/{variant_id}", tags=["Admin Products"]
)
async def admin_update_product_variant(
    product_id: str,
    variant_id: int,
    data: VariantUpdate,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    """Update a product inventory variant."""
    # Resolve slug or ID
    resolved = db.execute(
        text("SELECT id FROM products WHERE id::text = :pid OR slug = :pid"),
        {"pid": product_id},
    ).fetchone()
    if not resolved:
        raise HTTPException(status_code=404, detail="Product not found")
    pid = resolved[0]
    inv = db.execute(
        text("SELECT id FROM inventory WHERE id = :id AND product_id = :pid"),
        {"id": variant_id, "pid": pid},
    ).fetchone()
    if not inv:
        raise HTTPException(status_code=404, detail="Variant not found")
    sets, params = (
        ["updated_at = :now"],
        {"id": variant_id, "now": datetime.now(timezone.utc)},
    )
    if data.size is not None:
        sets.append("size = :size")
        params["size"] = data.size
    if data.color is not None:
        sets.append("color = :color")
        params["color"] = data.color
    if data.color_hex is not None:
        sets.append("color_hex = :color_hex")
        params["color_hex"] = data.color_hex
    if data.quantity is not None:
        sets.append("quantity = :qty")
        params["qty"] = data.quantity
    db.execute(text(f"UPDATE inventory SET {', '.join(sets)} WHERE id = :id"), params)
    db.commit()
    redis_client.invalidate_pattern("products:*")
    return {"message": "Variant updated"}


@app.delete(
    "/api/v1/admin/products/{product_id}/variants/{variant_id}",
    status_code=204,
    tags=["Admin Products"],
)
async def admin_delete_product_variant(
    product_id: str,
    variant_id: int,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    """Delete a product inventory variant."""
    # Resolve slug or ID
    resolved = db.execute(
        text("SELECT id FROM products WHERE id::text = :pid OR slug = :pid"),
        {"pid": product_id},
    ).fetchone()
    if not resolved:
        raise HTTPException(status_code=404, detail="Product not found")
    pid = resolved[0]
    result = db.execute(
        text("DELETE FROM inventory WHERE id = :id AND product_id = :pid RETURNING id"),
        {"id": variant_id, "pid": pid},
    )
    if not result.fetchone():
        raise HTTPException(status_code=404, detail="Variant not found")
    db.commit()
    redis_client.invalidate_pattern("products:*")


@app.post(
    "/api/v1/admin/products/{product_id}/variants/{variant_id}/adjust-stock",
    tags=["Admin Products"],
)
async def admin_adjust_variant_stock(
    product_id: str,
    variant_id: int,
    data: InventoryAdjustRequest,
    db: Session = Depends(get_db),
    user: dict = Depends(require_staff),
):
    """Adjust stock for a specific inventory variant."""
    inv = db.execute(
        text("SELECT id, quantity FROM inventory WHERE id = :id AND product_id = :pid"),
        {"id": variant_id, "pid": product_id},
    ).fetchone()
    if not inv:
        raise HTTPException(status_code=404, detail="Variant not found")
    new_qty = max(0, inv[1] + data.adjustment)
    db.execute(
        text("UPDATE inventory SET quantity = :qty, updated_at = :now WHERE id = :id"),
        {"qty": new_qty, "id": inv[0], "now": datetime.now(timezone.utc)},
    )
    db.commit()
    redis_client.invalidate_pattern("products:*")
    return {"id": inv[0], "new_quantity": new_qty, "adjustment": data.adjustment}


# ==================== Collection / Category Management (Admin) ====================


@app.get("/api/v1/admin/collections", tags=["Admin Collections"])
@app.get("/api/v1/admin/categories", tags=["Admin Collections"])
async def admin_list_collections(
    featured_only: bool = False,
    active_only: bool = False,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    """List all collections for admin (includes inactive)."""
    where_parts = []
    if active_only:
        where_parts.append("is_active = true")
    if featured_only:
        where_parts.append("is_featured = true")
    where_clause = ("WHERE " + " AND ".join(where_parts)) if where_parts else ""
    rows = db.execute(
        text(f"""
        SELECT c.id, c.name, c.slug, c.description, c.image_url,
               c.display_order, c.is_active, c.is_featured, c.created_at, c.updated_at,
               COUNT(p.id) as product_count
        FROM collections c
        LEFT JOIN products p ON p.category_id = c.id
        {where_clause}
        GROUP BY c.id, c.name, c.slug, c.description, c.image_url,
                 c.display_order, c.is_active, c.is_featured, c.created_at, c.updated_at
        ORDER BY c.display_order NULLS LAST, c.name
    """)
    ).fetchall()
    collections = [
        {
            "id": r[0],
            "name": r[1],
            "slug": r[2],
            "description": r[3],
            "image_url": _get_r2_public_url(r[4]) if r[4] else None,
            "display_order": r[5],
            "is_active": r[6],
            "is_featured": r[7],
            "created_at": r[8],
            "updated_at": r[9],
            "product_count": r[10],
        }
        for r in rows
    ]
    return collections


@app.post("/api/v1/admin/collections", status_code=201, tags=["Admin Collections"])
@app.post("/api/v1/admin/categories", status_code=201, tags=["Admin Collections"])
async def admin_create_collection(
    data: CategoryCreate,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    """Create a new collection (admin only)."""
    import re  # noqa: PLC0415

    slug = data.slug
    if not slug:
        slug = re.sub(r"[^a-z0-9]+", "-", data.name.lower()).strip("-")
    existing = db.execute(
        text("SELECT id FROM collections WHERE slug = :slug"), {"slug": slug}
    ).fetchone()
    if existing:
        slug = f"{slug}-{int(datetime.now(timezone.utc).timestamp())}"
    result = db.execute(
        text("""
        INSERT INTO collections (name, slug, description, image_url, display_order,
            is_active, is_featured, created_at, updated_at)
        VALUES (:name, :slug, :desc, :img, :order, :active, :featured, :now, :now)
        RETURNING id
    """),
        {
            "name": data.name,
            "slug": slug,
            "desc": data.description,
            "img": data.image_url,
            "order": data.display_order,
            "active": data.is_active,
            "featured": data.is_featured,
            "now": datetime.now(timezone.utc),
        },
    )
    coll_id = result.scalar()
    db.commit()
    redis_client.invalidate_pattern("collections:*")
    redis_client.invalidate_pattern("public:landing:*")
    return {
        "id": coll_id,
        "name": data.name,
        "slug": slug,
        "message": "Collection created",
    }


@app.patch("/api/v1/admin/collections/{collection_id}", tags=["Admin Collections"])
@app.patch("/api/v1/admin/categories/{collection_id}", tags=["Admin Collections"])
async def admin_update_collection(
    collection_id: int,
    data: CategoryUpdate,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    """Update a collection (admin only)."""
    sets, params = (
        ["updated_at = :now"],
        {"id": collection_id, "now": datetime.now(timezone.utc)},
    )
    if data.name is not None:
        sets.append("name = :name")
        params["name"] = data.name
    if data.slug is not None:
        sets.append("slug = :slug")
        params["slug"] = data.slug
    if data.description is not None:
        sets.append("description = :desc")
        params["desc"] = data.description
    if data.image_url is not None:
        sets.append("image_url = :img")
        params["img"] = data.image_url
    if data.display_order is not None:
        sets.append("display_order = :order")
        params["order"] = data.display_order
    if data.is_active is not None:
        sets.append("is_active = :active")
        params["active"] = data.is_active
    if data.is_featured is not None:
        sets.append("is_featured = :featured")
        params["featured"] = data.is_featured
    result = db.execute(
        text(f"UPDATE collections SET {', '.join(sets)} WHERE id = :id RETURNING id"),
        params,
    )
    if not result.fetchone():
        raise HTTPException(status_code=404, detail="Collection not found")
    db.commit()
    redis_client.invalidate_pattern("collections:*")
    redis_client.invalidate_pattern("public:landing:*")
    return {"message": "Collection updated", "id": collection_id}


@app.delete(
    "/api/v1/admin/collections/{collection_id}",
    status_code=204,
    tags=["Admin Collections"],
)
@app.delete(
    "/api/v1/admin/categories/{collection_id}",
    status_code=204,
    tags=["Admin Collections"],
)
async def admin_delete_collection(
    collection_id: int,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    """Delete a collection (admin only). Fails if active products exist."""
    product_count = (
        db.execute(
            text("SELECT COUNT(*) FROM products WHERE category_id = :id"),
            {"id": collection_id},
        ).scalar()
        or 0
    )
    if product_count > 0:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot delete collection with {product_count} products. Reassign or delete products first.",
        )
    result = db.execute(
        text("DELETE FROM collections WHERE id = :id RETURNING id"),
        {"id": collection_id},
    )
    if not result.fetchone():
        raise HTTPException(status_code=404, detail="Collection not found")
    db.commit()
    redis_client.invalidate_pattern("collections:*")
    redis_client.invalidate_pattern("public:landing:*")


@app.post("/api/v1/admin/collections/bulk/status", tags=["Admin Collections"])
async def admin_bulk_collection_status(
    data: BulkCollectionStatusUpdate,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    """Bulk activate/deactivate collections."""
    db.execute(
        text(
            "UPDATE collections SET is_active = :active, updated_at = :now WHERE id = ANY(:ids)"
        ),
        {"active": data.is_active, "ids": data.ids, "now": datetime.now(timezone.utc)},
    )
    db.commit()
    redis_client.invalidate_pattern("collections:*")
    redis_client.invalidate_pattern("public:landing:*")
    return {"updated": len(data.ids), "is_active": data.is_active}


@app.post("/api/v1/admin/collections/bulk/reorder", tags=["Admin Collections"])
async def admin_bulk_collection_reorder(
    data: BulkCollectionReorder,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    """Bulk reorder collections."""
    for item in data.items:
        db.execute(
            text(
                "UPDATE collections SET display_order = :order, updated_at = :now WHERE id = :id"
            ),
            {
                "order": item["display_order"],
                "id": item["id"],
                "now": datetime.now(timezone.utc),
            },
        )
    db.commit()
    redis_client.invalidate_pattern("collections:*")
    redis_client.invalidate_pattern("public:landing:*")
    return {"reordered": len(data.items)}


@app.post("/api/v1/admin/categories/{category_id}/image", tags=["Admin Collections"])
async def admin_upload_collection_image(
    category_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    """Upload an image for a collection/category to R2."""
    coll = db.execute(
        text("SELECT id, image_url FROM collections WHERE id = :id"),
        {"id": category_id},
    ).fetchone()
    if not coll:
        raise HTTPException(status_code=404, detail="Collection not found")
    if coll[1]:
        await r2_service.delete_image(coll[1])
    full_url = await r2_service.upload_image(file, folder="collections")
    r2_base = settings.R2_PUBLIC_URL.rstrip("/") if settings.R2_PUBLIC_URL else ""
    relative_path = (
        full_url.replace(r2_base + "/", "")
        if r2_base and full_url.startswith(r2_base)
        else full_url
    )
    db.execute(
        text(
            "UPDATE collections SET image_url = :img, updated_at = :now WHERE id = :id"
        ),
        {"img": relative_path, "id": category_id, "now": datetime.now(timezone.utc)},
    )
    db.commit()
    redis_client.invalidate_pattern("collections:*")
    redis_client.invalidate_pattern("public:landing:*")
    return {
        "message": "Collection image updated",
        "image_url": _get_r2_public_url(relative_path),
    }


@app.delete(
    "/api/v1/admin/categories/{category_id}/image",
    status_code=204,
    tags=["Admin Collections"],
)
async def admin_delete_collection_image(
    category_id: int, db: Session = Depends(get_db), user: dict = Depends(require_admin)
):
    """Delete a collection image from R2 and clear the record."""
    coll = db.execute(
        text("SELECT id, image_url FROM collections WHERE id = :id"),
        {"id": category_id},
    ).fetchone()
    if not coll:
        raise HTTPException(status_code=404, detail="Collection not found")
    if coll[1]:
        await r2_service.delete_image(coll[1])
        db.execute(
            text(
                "UPDATE collections SET image_url = NULL, updated_at = :now WHERE id = :id"
            ),
            {"id": category_id, "now": datetime.now(timezone.utc)},
        )
        db.commit()
    redis_client.invalidate_pattern("collections:*")


# ==================== Inventory Full CRUD (Admin) ====================


@app.post("/api/v1/admin/inventory", status_code=201, tags=["Admin Inventory"])
async def admin_create_inventory(
    data: dict, db: Session = Depends(get_db), user: dict = Depends(require_staff)
):
    """Create an inventory record (admin/staff only)."""
    required = ["product_id", "sku", "quantity"]
    for f in required:
        if f not in data:
            raise HTTPException(status_code=400, detail=f"Missing required field: {f}")
    result = db.execute(
        text("""
        INSERT INTO inventory (product_id, sku, size, color, quantity, low_stock_threshold,
            cost_price, created_at, updated_at)
        VALUES (:pid, :sku, :size, :color, :qty, :threshold, :cost, :now, :now)
        RETURNING id
    """),
        {
            "pid": data["product_id"],
            "sku": data["sku"],
            "size": data.get("size"),
            "color": data.get("color"),
            "qty": data["quantity"],
            "threshold": data.get("low_stock_threshold", 5),
            "cost": data.get("cost_price"),
            "now": datetime.now(timezone.utc),
        },
    )
    inv_id = result.scalar()
    db.commit()
    redis_client.invalidate_pattern("products:*")
    return {"id": inv_id, "message": "Inventory record created"}


@app.patch("/api/v1/admin/inventory/{inventory_id}", tags=["Admin Inventory"])
async def admin_update_inventory(
    inventory_id: int,
    data: dict,
    db: Session = Depends(get_db),
    user: dict = Depends(require_staff),
):
    """Update an inventory record (admin/staff only)."""
    inv = db.execute(
        text("SELECT id FROM inventory WHERE id = :id"), {"id": inventory_id}
    ).fetchone()
    if not inv:
        raise HTTPException(status_code=404, detail="Inventory item not found")
    ALLOWED = {"quantity", "low_stock_threshold", "cost_price", "size", "color"}
    sets, params = (
        ["updated_at = :now"],
        {"id": inventory_id, "now": datetime.now(timezone.utc)},
    )
    for field in ALLOWED:
        if field in data:
            sets.append(f"{field} = :{field}")
            params[field] = data[field]
    if len(sets) == 1:
        return {"message": "Nothing to update"}
    db.execute(text(f"UPDATE inventory SET {', '.join(sets)} WHERE id = :id"), params)
    db.commit()
    redis_client.invalidate_pattern("products:*")
    return {"message": "Inventory updated", "id": inventory_id}


@app.get("/api/v1/admin/inventory/movements", tags=["Admin Inventory"])
async def admin_inventory_movements(
    product_id: Optional[int] = None,
    page: int = Query(1, ge=1),
    limit: int = Query(20, le=100),
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    """Get inventory movement history (admin only)."""
    offset = (page - 1) * limit
    where, params = "", {"lim": limit, "off": offset}
    if product_id:
        where = "WHERE im.product_id = :pid"
        params["pid"] = product_id
    rows = db.execute(
        text(f"""
        SELECT im.*, p.name as product_name
        FROM inventory_movements im
        JOIN products p ON p.id = im.product_id
        {where}
        ORDER BY im.created_at DESC LIMIT :lim OFFSET :off
    """),
        params,
    ).fetchall()
    return {"movements": [dict(r._mapping) for r in rows]}


@app.get("/api/v1/admin/inventory/out-of-stock", tags=["Admin Inventory"])
async def admin_out_of_stock(
    db: Session = Depends(get_db), user: dict = Depends(require_admin)
):
    """Get out-of-stock inventory items."""
    rows = db.execute(
        text(
            "SELECT i.*, p.name as product_name FROM inventory i "
            "JOIN products p ON p.id = i.product_id WHERE i.quantity = 0 ORDER BY p.name"
        )
    ).fetchall()
    return {"items": [dict(r._mapping) for r in rows]}


# ==================== Returns Management (Admin) ====================


@app.get("/api/v1/admin/returns", tags=["Admin Returns"])
async def admin_list_returns(
    status_filter: Optional[str] = None,
    skip: int = 0,
    limit: int = 50,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    """List all return requests (admin only)."""
    where, params = "", {"limit": limit, "skip": skip}
    if status_filter:
        where = "WHERE r.status = :status"
        params["status"] = status_filter
    rows = db.execute(
        text(f"""
        SELECT r.id, r.order_id, r.user_id, r.reason, r.description,
               r.status, r.refund_amount, r.requested_at, r.updated_at,
               u.email as customer_email, u.username as customer_username
        FROM return_requests r
        LEFT JOIN users u ON r.user_id = u.id
        {where}
        ORDER BY r.requested_at DESC
        LIMIT :limit OFFSET :skip
    """),
        params,
    ).fetchall()
    returns = [
        {
            "id": r[0],
            "order_id": r[1],
            "user_id": r[2],
            "reason": r[3],
            "description": r[4],
            "status": r[5],
            "refund_amount": float(r[6]) if r[6] else None,
            "requested_at": r[7],
            "created_at": r[7],
            "updated_at": r[8],
            "customer_email": r[9],
            "customer_username": r[10],
        }
        for r in rows
    ]
    return {"returns": returns, "total": len(returns)}


@app.get("/api/v1/admin/returns/{return_id}", tags=["Admin Returns"])
async def admin_get_return(
    return_id: int, db: Session = Depends(get_db), user: dict = Depends(require_admin)
):
    """Get return request details (admin only)."""
    row = db.execute(
        text("""
        SELECT r.id, r.order_id, r.user_id, r.reason, r.description,
               r.status, r.refund_amount, r.requested_at, r.updated_at,
               r.approved_at, r.received_at, r.refunded_at, r.refund_transaction_id,
               r.rejection_reason, r.return_tracking_number,
               o.shipping_address, o.total_amount,
               u.email as customer_email, u.username as customer_username
        FROM return_requests r
        LEFT JOIN orders o ON o.id = r.order_id
        LEFT JOIN users u ON r.user_id = u.id
        WHERE r.id = :id
    """),
        {"id": return_id},
    ).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Return request not found")

    items = db.execute(
        text("""
        SELECT oi.product_id, oi.product_name, oi.sku, oi.size, oi.color, oi.quantity, oi.unit_price
        FROM order_items oi
        WHERE oi.order_id = :order_id
        ORDER BY oi.id
    """),
        {"order_id": row[1]},
    ).fetchall()

    timeline = [
        {
            "status": "requested",
            "date": row[7],
            "note": "Return request created",
            "actor": "Customer",
        }
    ]
    if row[9]:
        timeline.append(
            {
                "status": "approved",
                "date": row[9],
                "note": "Return approved",
                "actor": "Admin",
            }
        )
    if row[10]:
        timeline.append(
            {
                "status": "received",
                "date": row[10],
                "note": "Returned item marked as received",
                "actor": "Admin",
            }
        )
    if row[11]:
        timeline.append(
            {
                "status": "refunded",
                "date": row[11],
                "note": "Refund processed",
                "actor": "Admin",
            }
        )
    if row[13] and row[5] == "rejected":
        timeline.append(
            {
                "status": "rejected",
                "date": row[8],
                "note": row[13],
                "actor": "Admin",
            }
        )

    return {
        "id": row[0],
        "order_id": row[1],
        "user_id": row[2],
        "reason": row[3],
        "description": row[4],
        "status": row[5],
        "refund_amount": float(row[6]) if row[6] else None,
        "requested_at": row[7],
        "created_at": row[7],
        "updated_at": row[8],
        "approved_at": row[9],
        "received_at": row[10],
        "refunded_at": row[11],
        "refund_transaction_id": row[12],
        "rejection_reason": row[13],
        "return_tracking_number": row[14],
        "return_number": f"RET-{str(row[0]).zfill(6)}",
        "type": "return",
        "order_number": f"#{row[1]}",
        "total_amount": float(row[16]) if row[16] else 0,
        "customer_email": row[17],
        "customer_username": row[18],
        "customer": {
            "name": row[18] or row[17] or "Customer",
            "email": row[17],
            "total_orders": None,
            "total_spent": float(row[16]) if row[16] else 0,
        },
        "shipping_address": {
            "name": None,
            "address": row[15],
            "city": None,
            "state": None,
            "pincode": None,
            "phone": None,
        },
        "items": [
            {
                "product_id": item[0],
                "name": item[1],
                "sku": item[2],
                "size": item[3],
                "color": item[4],
                "quantity": item[5],
                "price": float(item[6]) if item[6] else 0,
            }
            for item in items
        ],
        "timeline": timeline,
        "refund": {
            "amount": float(row[6]) if row[6] else 0,
            "method": row[12] or "manual",
            "status": row[5],
        }
        if row[5] == "refunded"
        else None,
    }


@app.post("/api/v1/admin/returns/{return_id}/approve", tags=["Admin Returns"])
async def admin_approve_return(
    return_id: int,
    data: dict = None,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    """Approve a return request (admin only)."""
    if data is None:
        data = {}
    ret = db.execute(
        text("SELECT id, status FROM return_requests WHERE id = :id"), {"id": return_id}
    ).fetchone()
    if not ret:
        raise HTTPException(status_code=404, detail="Return request not found")
    refund_amount = data.get("refund_amount")
    sets = ["status = 'approved'", "updated_at = :now"]
    params = {"id": return_id, "now": datetime.now(timezone.utc)}
    if refund_amount is not None:
        sets.append("refund_amount = :refund")
        params["refund"] = refund_amount
    db.execute(
        text(f"UPDATE return_requests SET {', '.join(sets)} WHERE id = :id"), params
    )
    db.commit()
    return {"message": "Return approved", "return_id": return_id}


@app.post("/api/v1/admin/returns/{return_id}/reject", tags=["Admin Returns"])
async def admin_reject_return(
    return_id: int,
    data: dict,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    """Reject a return request (admin only)."""
    ret = db.execute(
        text("SELECT id FROM return_requests WHERE id = :id"), {"id": return_id}
    ).fetchone()
    if not ret:
        raise HTTPException(status_code=404, detail="Return request not found")
    reason = data.get("reason", "")
    db.execute(
        text("""
        UPDATE return_requests
        SET status = 'rejected', rejection_reason = :reason, updated_at = :now
        WHERE id = :id
    """),
        {"reason": reason, "now": datetime.now(timezone.utc), "id": return_id},
    )
    db.commit()
    return {"message": "Return rejected", "return_id": return_id}


@app.post("/api/v1/admin/returns/{return_id}/receive", tags=["Admin Returns"])
async def admin_receive_return(
    return_id: int,
    data: dict = None,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    """Mark a return as received (admin only)."""
    if data is None:
        data = {}
    ret = db.execute(
        text("SELECT id FROM return_requests WHERE id = :id"), {"id": return_id}
    ).fetchone()
    if not ret:
        raise HTTPException(status_code=404, detail="Return request not found")
    sets = ["status = 'received'", "updated_at = :now"]
    params = {"id": return_id, "now": datetime.now(timezone.utc)}
    tracking = data.get("tracking_number")
    if tracking:
        sets.append("return_tracking_number = :tn")
        params["tn"] = tracking
    db.execute(
        text(f"UPDATE return_requests SET {', '.join(sets)} WHERE id = :id"), params
    )
    db.commit()
    return {"message": "Return marked as received", "return_id": return_id}


@app.post("/api/v1/admin/returns/{return_id}/refund", tags=["Admin Returns"])
async def admin_process_return_refund(
    return_id: int,
    data: dict,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    """Process refund for a return (admin only)."""
    ret = db.execute(
        text("SELECT id FROM return_requests WHERE id = :id"), {"id": return_id}
    ).fetchone()
    if not ret:
        raise HTTPException(status_code=404, detail="Return request not found")
    txn_id = data.get("refund_transaction_id", "")
    db.execute(
        text("""
        UPDATE return_requests
        SET status = 'refunded', refund_transaction_id = :txn, updated_at = :now
        WHERE id = :id
    """),
        {"txn": txn_id, "now": datetime.now(timezone.utc), "id": return_id},
    )
    db.commit()
    return {"message": "Refund processed", "return_id": return_id}


# ==================== Reviews Management (Admin) ====================


@app.get("/api/v1/admin/reviews", tags=["Admin Reviews"])
async def admin_list_reviews(
    product_id: Optional[int] = None,
    is_approved: Optional[bool] = None,
    skip: int = 0,
    limit: int = 50,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin),
):
    """List all reviews (admin only)."""
    where_parts = []
    params = {"limit": limit, "skip": skip}
    if product_id:
        where_parts.append("r.product_id = :product_id")
        params["product_id"] = product_id
    if is_approved is not None:
        where_parts.append("r.is_approved = :approved")
        params["approved"] = is_approved
    where_clause = ("WHERE " + " AND ".join(where_parts)) if where_parts else ""
    rows = db.execute(
        text(f"""
        SELECT r.id, r.product_id, r.user_id, r.rating, r.comment,
               r.is_approved, r.created_at,
               p.name as product_name,
               u.username as customer_username, u.email as customer_email
        FROM reviews r
        LEFT JOIN products p ON r.product_id = p.id
        LEFT JOIN users u ON r.user_id = u.id
        {where_clause}
        ORDER BY r.created_at DESC
        LIMIT :limit OFFSET :skip
    """),
        params,
    ).fetchall()
    reviews = [
        {
            "id": r[0],
            "product_id": r[1],
            "user_id": r[2],
            "rating": r[3],
            "comment": r[4],
            "is_approved": r[5],
            "created_at": r[6],
            "product_name": r[7],
            "customer_username": r[8],
            "customer_email": r[9],
        }
        for r in rows
    ]
    return {"reviews": reviews, "total": len(reviews)}


@app.patch("/api/v1/admin/reviews/{review_id}/approve", tags=["Admin Reviews"])
async def admin_approve_review(
    review_id: int, db: Session = Depends(get_db), user: dict = Depends(require_admin)
):
    """Approve a review (admin only)."""
    result = db.execute(
        text(
            "UPDATE reviews SET is_approved = true, updated_at = :now WHERE id = :id RETURNING id"
        ),
        {"id": review_id, "now": datetime.now(timezone.utc)},
    )
    if not result.fetchone():
        raise HTTPException(status_code=404, detail="Review not found")
    db.commit()
    redis_client.invalidate_pattern("products:*")
    return {"message": "Review approved", "review_id": review_id}


@app.patch("/api/v1/admin/reviews/{review_id}/reject", tags=["Admin Reviews"])
async def admin_reject_review(
    review_id: int, db: Session = Depends(get_db), user: dict = Depends(require_admin)
):
    """Reject (unapprove) a review (admin only)."""
    result = db.execute(
        text(
            "UPDATE reviews SET is_approved = false, updated_at = :now WHERE id = :id RETURNING id"
        ),
        {"id": review_id, "now": datetime.now(timezone.utc)},
    )
    if not result.fetchone():
        raise HTTPException(status_code=404, detail="Review not found")
    db.commit()
    redis_client.invalidate_pattern("products:*")
    return {"message": "Review rejected", "review_id": review_id}


@app.delete(
    "/api/v1/admin/reviews/{review_id}", status_code=204, tags=["Admin Reviews"]
)
async def admin_delete_review(
    review_id: int, db: Session = Depends(get_db), user: dict = Depends(require_admin)
):
    """Delete a review (admin only)."""
    result = db.execute(
        text("DELETE FROM reviews WHERE id = :id RETURNING id"), {"id": review_id}
    )
    if not result.fetchone():
        raise HTTPException(status_code=404, detail="Review not found")
    db.commit()
    redis_client.invalidate_pattern("products:*")


# ==================== Super Admin Backup System ====================

import subprocess
import os
import glob

BACKUP_DIR = os.environ.get("BACKUP_DIR", "/backups")


@app.post("/api/v1/admin/backup/create", tags=["Super Admin Backup"])
async def create_backup(user: dict = Depends(require_super_admin)):
    """Trigger a manual PostgreSQL database backup (super admin only)."""
    os.makedirs(BACKUP_DIR, exist_ok=True)
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"aarya_backup_{timestamp}.sql"
    filepath = os.path.join(BACKUP_DIR, filename)

    db_url = os.environ.get("DATABASE_URL", "")
    # Parse connection details from DATABASE_URL
    # Format: postgresql://user:password@host:port/dbname
    try:
        from urllib.parse import urlparse
        parsed = urlparse(db_url)
        env = {
            **os.environ,
            "PGPASSWORD": parsed.password or "",
        }
        cmd = [
            "pg_dump",
            "-h", parsed.hostname or "postgres",
            "-p", str(parsed.port or 5432),
            "-U", parsed.username or "postgres",
            "-d", parsed.path.lstrip("/") or "aarya_clothing",
            "-f", filepath,
            "--no-owner",
            "--no-acl",
        ]
        result = subprocess.run(cmd, env=env, capture_output=True, text=True, timeout=120)
        if result.returncode != 0:
            raise Exception(result.stderr)

        file_size = os.path.getsize(filepath)
        logger.info(f"Backup created: {filename} ({file_size} bytes) by user {user.get('user_id')}")
        return {
            "success": True,
            "filename": filename,
            "size_bytes": file_size,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "message": f"Backup created successfully: {filename}",
        }
    except subprocess.TimeoutExpired:
        raise HTTPException(status_code=504, detail="Backup timed out")
    except Exception as e:
        logger.error(f"Backup failed: {e}")
        raise HTTPException(status_code=500, detail=f"Backup failed: {str(e)}")


@app.get("/api/v1/admin/backup/list", tags=["Super Admin Backup"])
async def list_backups(user: dict = Depends(require_super_admin)):
    """List all available database backups (super admin only)."""
    os.makedirs(BACKUP_DIR, exist_ok=True)
    files = sorted(glob.glob(os.path.join(BACKUP_DIR, "*.sql")), reverse=True)
    backups = []
    for f in files[:50]:  # Return max 50 most recent
        stat = os.stat(f)
        backups.append({
            "filename": os.path.basename(f),
            "size_bytes": stat.st_size,
            "size_mb": round(stat.st_size / 1024 / 1024, 2),
            "created_at": datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).isoformat(),
        })
    return {"backups": backups, "total": len(backups), "backup_dir": BACKUP_DIR}


@app.delete("/api/v1/admin/backup/{filename}", tags=["Super Admin Backup"])
async def delete_backup(filename: str, user: dict = Depends(require_super_admin)):
    """Delete a specific backup file (super admin only)."""
    # Security: only allow .sql files and no path traversal
    if not filename.endswith(".sql") or "/" in filename or ".." in filename:
        raise HTTPException(status_code=400, detail="Invalid filename")
    filepath = os.path.join(BACKUP_DIR, filename)
    if not os.path.exists(filepath):
        raise HTTPException(status_code=404, detail="Backup not found")
    os.remove(filepath)
    logger.info(f"Backup deleted: {filename} by user {user.get('user_id')}")
    return {"success": True, "message": f"Backup {filename} deleted"}
