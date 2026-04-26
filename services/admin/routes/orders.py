"""Admin order operations.

Owns the admin/staff side of orders: listing with filters/search, detail
read with line items + tracking history, status transitions (single + bulk),
tracking-history fetch, and Excel export plus the POD upload + template
flow that lets ops mark a batch of orders as shipped from a spreadsheet.

Customer-facing order routes live in commerce/routes/customer_orders.py.
"""

from __future__ import annotations

import io
import logging
from collections import defaultdict
from datetime import datetime
from shared.time_utils import now_ist
from typing import Optional

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import text
from sqlalchemy.orm import Session

from core.exception_handler import log_business_event
from core.redis_client import redis_client
from database.database import get_db
from schemas.admin import BulkOrderUpdate, OrderStatusUpdate
from shared.auth_middleware import require_admin, require_staff

logger = logging.getLogger(__name__)

router = APIRouter(tags=["Admin"])


# ==================== Payment recovery (Razorpay vs DB) ====================


@router.get("/api/v1/admin/orders/payment-recovery", tags=["Admin"])
async def get_payment_recovery_report(
    from_timestamp: Optional[int] = Query(
        None, description="Unix timestamp; default last 48h"
    ),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_admin),
):
    """Cross-reference captured Razorpay payments with ``orders.transaction_id``."""
    from service.order_payment_recovery import payment_recovery_report

    return payment_recovery_report(db, from_timestamp)


@router.post("/api/v1/admin/orders/force-create", tags=["Admin"])
async def admin_force_create_order_from_payment(
    payload: dict,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_admin),
):
    """Create a confirmed order row from a captured Razorpay ``payment_id`` (ops recovery)."""
    from service.order_payment_recovery import force_create_order_from_payment

    pid = str(payload.get("payment_id", "")).strip()
    return force_create_order_from_payment(db, pid, current_user)


# ==================== Admin Orders (Direct DB) ====================


@router.get("/api/v1/admin/orders", tags=["Admin"])
async def list_all_orders(
    status: Optional[str] = None,
    search: Optional[str] = None,
    user_id: Optional[int] = None,
    skip: int = 0,
    limit: int = 20,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_staff),
):
    """List all orders (admin/staff) - direct DB query instead of httpx proxy."""
    where_clauses = []
    params = {"lim": limit, "off": skip}
    if status:
        where_clauses.append("o.status = :status")
        params["status"] = status
    if user_id is not None:
        where_clauses.append("o.user_id = :user_id")
        params["user_id"] = user_id
    if search:
        where_clauses.append(
            "(CAST(o.id AS TEXT) ILIKE :search OR u.email ILIKE :search OR u.username ILIKE :search OR COALESCE(o.tracking_number, '') ILIKE :search)"
        )
        params["search"] = f"%{search.strip()}%"

    where_clause = f"WHERE {' AND '.join(where_clauses)}" if where_clauses else ""

    rows = db.execute(
        text(f"""
        SELECT o.id, o.user_id, o.subtotal, o.shipping_cost,
               o.total_amount, o.payment_method, o.status, o.tracking_number,
               o.order_notes, o.created_at, o.updated_at,
               u.email as customer_email, COALESCE(up.full_name, u.username) as customer_name,
               COALESCE(up.phone, '') as customer_phone,
               o.invoice_number, o.shipping_address, o.razorpay_payment_id
        FROM orders o
        LEFT JOIN users u ON u.id = o.user_id
        LEFT JOIN user_profiles up ON up.user_id = o.user_id
        {where_clause}
        ORDER BY o.created_at DESC
        LIMIT :lim OFFSET :off
    """),
        params,
    ).fetchall()

    count_params = {k: v for k, v in params.items() if k not in ("lim", "off")}
    total = (
        db.execute(
            text(
                f"SELECT COUNT(*) FROM orders o LEFT JOIN users u ON u.id = o.user_id {where_clause}"
            ),
            count_params,
        ).scalar()
        or 0
    )

    # Collect order IDs for batch fetching order items
    order_ids = [r[0] for r in rows]

    # Batch fetch order items for all orders in a single query (avoids N+1)
    items_by_order = {}
    if order_ids:
        placeholders = ", ".join(f":oid_{i}" for i in range(len(order_ids)))
        items_params = {f"oid_{i}": oid for i, oid in enumerate(order_ids)}
        items_rows = db.execute(
            text(f"""
                SELECT oi.order_id, oi.id as item_id, oi.product_id, oi.product_name,
                       oi.size, oi.color, oi.quantity, oi.unit_price, oi.line_total,
                       p.name as product_name_from_catalog,
                       pi.image_url as image_url
                FROM order_items oi
                LEFT JOIN products p ON p.id = oi.product_id
                LEFT JOIN product_images pi ON p.id = pi.product_id AND pi.is_primary = true
                WHERE oi.order_id IN ({placeholders})
                ORDER BY oi.order_id, oi.id
            """),
            items_params,
        ).fetchall()
        for item_row in items_rows:
            oid = item_row[0]
            if oid not in items_by_order:
                items_by_order[oid] = []
            items_by_order[oid].append(
                {
                    "id": item_row[1],
                    "product_id": item_row[2],
                    "product_name": item_row[3] or item_row[9] or "Unknown Product",
                    "size": item_row[4],
                    "color": item_row[5],
                    "quantity": item_row[6],
                    "unit_price": float(item_row[7] or 0),
                    "total_price": float(item_row[8] or 0),
                    "image_url": item_row[10],
                }
            )

    import json
    orders = []
    for r in rows:
        order_id = r[0]
        items = items_by_order.get(order_id, [])

        # Enhanced customer info resolution
        email = r[11]
        name = r[12]
        phone = r[13]

        # If user is missing (guest), try to parse from shipping address
        if not email or not name:
            try:
                addr_json = r[15]
                if addr_json:
                    addr = json.loads(addr_json) if isinstance(addr_json, str) else addr_json
                    name = name or addr.get("name") or addr.get("full_name")
                    email = email or addr.get("email")
                    phone = phone or addr.get("phone")
            except Exception:
                pass

        orders.append({
            "id": order_id,
            "user_id": r[1],
            "subtotal": float(r[2] or 0),
            "shipping_cost": float(r[3] or 0),
            "total_amount": float(r[4] or 0),
            "payment_method": r[5],
            "status": r[6],
            "tracking_number": r[7],
            "order_notes": r[8],
            "created_at": str(r[9]),
            "updated_at": str(r[10]),
            "customer_email": email or "Guest",
            "customer_name": name or "Guest Customer",
            "customer_phone": phone or "-",
            "invoice_number": r[14],
            "shipping_address": r[15],
            "razorpay_payment_id": r[16],
            "order_number": f"ORD-{r[0]:06d}",
            "items": items,
            "item_count": sum(it["quantity"] for it in items),
        })

    # Get accurate status counts for the current filter
    status_counts_rows = db.execute(
        text(f"""
            SELECT o.status, COUNT(*) 
            FROM orders o 
            LEFT JOIN users u ON u.id = o.user_id 
            {where_clause} 
            GROUP BY o.status
        """),
        count_params
    ).fetchall()
    status_counts = {row[0]: row[1] for row in status_counts_rows}
    
    # Get total revenue for the current filter
    total_revenue = db.execute(
        text(f"""
            SELECT SUM(o.total_amount) 
            FROM orders o 
            LEFT JOIN users u ON u.id = o.user_id 
            {where_clause}
        """),
        count_params
    ).scalar() or 0

    return {
        "orders": orders, 
        "total": total, 
        "skip": skip, 
        "limit": limit,
        "status_counts": status_counts,
        "total_revenue": float(total_revenue)
    }


@router.get("/api/v1/admin/orders/{order_id}", tags=["Admin"])
async def get_order(
    order_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_staff),
):
    """Get order details - direct DB query instead of httpx proxy."""
    order = db.execute(
        text("""
        SELECT o.*, u.email as customer_email, u.username as customer_name, up.phone as customer_phone
        FROM orders o
        LEFT JOIN users u ON u.id = o.user_id
        LEFT JOIN user_profiles up ON up.user_id = u.id
        WHERE o.id = :id
    """),
        {"id": order_id},
    ).fetchone()

    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    order_dict = dict(order._mapping)

    # Fetch order items
    items = db.execute(
        text("""
        SELECT oi.id, oi.order_id, oi.inventory_id, oi.product_id, oi.product_name,
               oi.sku, oi.size, oi.color, oi.hsn_code, oi.gst_rate,
               oi.quantity, oi.unit_price, oi.price, oi.created_at,
               p.name as product_name,
               COALESCE(i.image_url, pi.image_url) as image_url
        FROM order_items oi
        LEFT JOIN products p ON p.id = oi.product_id
        LEFT JOIN inventory i ON i.id = oi.inventory_id
        LEFT JOIN LATERAL (
            SELECT image_url FROM product_images
            WHERE product_id = oi.product_id AND is_primary = true
            LIMIT 1
        ) pi ON true
        WHERE oi.order_id = :oid
    """),
        {"oid": order_id},
    ).fetchall()
    order_dict["items"] = [dict(i._mapping) for i in items]

    # Fetch tracking
    tracking = db.execute(
        text(
            "SELECT * FROM order_tracking WHERE order_id = :oid ORDER BY created_at DESC"
        ),
        {"oid": order_id},
    ).fetchall()
    order_dict["tracking"] = [dict(t._mapping) for t in tracking]

    return {
        "order": order_dict,
        "items": order_dict["items"],
        "tracking": order_dict["tracking"],
        "customer": {
            "id": order_dict.get("user_id"),
            "full_name": order_dict.get("customer_name") or "Customer",
            "email": order_dict.get("customer_email"),
            "phone": order_dict.get("customer_phone"),
        },
    }


@router.patch("/api/v1/admin/orders/{order_id}/status", tags=["Admin"])
async def update_order_status(
    order_id: int,
    data: OrderStatusUpdate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_staff),
):
    """Update order status (admin/staff only)."""
    new_status = data.status

    # Validate status
    valid_statuses = ["confirmed", "shipped", "delivered", "cancelled"]
    if new_status not in valid_statuses:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid status. Must be one of: {', '.join(valid_statuses)}",
        )

    # POD number is required when shipping
    pod_number = data.get_pod()
    if new_status == "shipped" and not pod_number:
        raise HTTPException(
            status_code=400,
            detail="POD number is required when marking order as shipped",
        )

    now = now_ist()
    extra_sets = ""
    params = {"status": new_status, "now": now, "id": order_id}

    if new_status == "shipped" and pod_number:
        # Extract courier_name from the request if provided
        courier_name = getattr(data, "courier", None) or getattr(
            data, "courier_name", None
        )
        extra_sets = (
            ", tracking_number = :pod, courier_name = :courier_name, shipped_at = :now"
        )
        params["pod"] = pod_number
        params["courier_name"] = courier_name
    elif new_status == "delivered":
        extra_sets = ", delivered_at = :now"
    elif new_status == "cancelled":
        extra_sets = ", cancelled_at = :now"

    result = db.execute(
        text(
            f"UPDATE orders SET status = :status, updated_at = :now{extra_sets} WHERE id = :id RETURNING id"
        ),
        params,
    )
    db.commit()

    if not result.fetchone():
        raise HTTPException(status_code=404, detail="Order not found")

    default_notes = {
        "shipped": f"Order shipped — POD number: {pod_number}"
        if pod_number
        else "Order shipped",
        "delivered": "Order delivered to customer",
        "cancelled": "Order cancelled by admin",
        "confirmed": "Order confirmed",
    }

    # Extract courier_name for tracking entry
    courier_name = getattr(data, "courier", None) or getattr(data, "courier_name", None)

    db.execute(
        text("""
        INSERT INTO order_tracking (order_id, status, notes, updated_by, created_at, courier_name)
        VALUES (:order_id, :status, :notes, :updated_by, :created_at, :courier_name)
    """),
        {
            "order_id": order_id,
            "status": new_status,
            "notes": data.notes or default_notes.get(new_status),
            "updated_by": current_user.get("user_id"),
            "created_at": now,
            "courier_name": courier_name,
        },
    )
    db.commit()

    # Invalidate cache
    redis_client.delete_cache(f"order:{order_id}")
    redis_client.delete_cache("admin:orders:*")

    log_business_event(
        "order_status_updated",
        {
            "order_id": order_id,
            "new_status": new_status,
            "updated_by": current_user.get("user_id"),
        },
    )

    return {"success": True, "order_id": order_id, "status": new_status}


@router.patch("/api/v1/admin/orders/bulk-status", tags=["Admin"])
async def bulk_update_order_status(
    data: BulkOrderUpdate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_staff),
):
    """Bulk update order statuses (admin/staff only)."""
    if not data.order_ids:
        raise HTTPException(status_code=400, detail="order_ids is required")

    valid_statuses = ["confirmed", "shipped", "delivered", "cancelled"]
    if data.status not in valid_statuses:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid status. Must be one of: {', '.join(valid_statuses)}",
        )

    now = now_ist()
    updated = db.execute(
        text(
            """
        UPDATE orders
        SET status = :status, updated_at = :now
            """
            + (
                """, tracking_number = :pod, courier_name = :courier, shipped_at = :now"""
                if data.status == "shipped" and data.pod_number
                else ""
            )
            + """
        WHERE id = ANY(:order_ids)
        RETURNING id
    """
        ),
        {
            "status": data.status,
            "now": now,
            "order_ids": data.order_ids,
            "pod": data.pod_number if data.status == "shipped" else None,
            "courier": getattr(data, "courier_name", None)
            if data.status == "shipped"
            else None,
        },
    ).fetchall()

    for row in updated:
        db.execute(
            text(
                """
            INSERT INTO order_tracking (order_id, status, notes, updated_by, created_at"""
                + (
                    ", tracking_number, courier_name"
                    if data.status == "shipped" and data.pod_number
                    else ""
                )
                + """)
            VALUES (:order_id, :status, :notes, :updated_by, :created_at"""
                + (
                    ", :pod, :courier"
                    if data.status == "shipped" and data.pod_number
                    else ""
                )
                + """)
        """
            ),
            {
                "order_id": row[0],
                "status": data.status,
                "notes": data.notes,
                "updated_by": current_user.get("user_id"),
                "created_at": now,
                "pod": data.pod_number if data.status == "shipped" else None,
                "courier": getattr(data, "courier_name", None)
                if data.status == "shipped"
                else None,
            },
        )

    db.commit()
    redis_client.delete_cache("admin:orders:*")
    return {"updated": len(updated), "status": data.status}


@router.get("/api/v1/admin/orders/{order_id}/tracking", tags=["Admin"])
async def get_order_tracking(
    order_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_staff),
):
    """Get order tracking history (admin/staff only)."""
    tracking = db.execute(
        text(
            "SELECT * FROM order_tracking WHERE order_id = :oid ORDER BY created_at DESC"
        ),
        {"oid": order_id},
    ).fetchall()
    return {"tracking": [dict(t._mapping) for t in tracking]}


@router.get("/api/v1/admin/orders/export/excel", tags=["Admin"])
async def export_orders_excel(
    status: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_staff),
):
    """Export orders as Excel file with separate sheets per day. Filters: status, from_date (YYYY-MM-DD), to_date (YYYY-MM-DD)."""
    try:
        import io
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from fastapi.responses import StreamingResponse
    except ImportError:
        raise HTTPException(
            status_code=500, detail="openpyxl not installed. Run: pip install openpyxl"
        )

    where_clauses = []
    params: dict = {}
    if status and status in ["confirmed", "shipped", "delivered", "cancelled"]:
        where_clauses.append("o.status = :status")
        params["status"] = status
    if from_date:
        where_clauses.append("DATE(o.created_at) >= :from_date")
        params["from_date"] = from_date
    if to_date:
        where_clauses.append("DATE(o.created_at) <= :to_date")
        params["to_date"] = to_date

    where_clause = f"WHERE {' AND '.join(where_clauses)}" if where_clauses else ""

    rows = db.execute(
        text(f"""
        SELECT o.id, u.email,
               COALESCE(u.full_name, u.username) AS customer_name,
               u.phone AS customer_phone,
               o.total_amount, o.payment_method, o.status,
               o.tracking_number, o.courier_name,
               o.shipping_address, o.created_at
        FROM orders o
        LEFT JOIN users u ON u.id = o.user_id
        {where_clause}
        ORDER BY o.created_at DESC
        LIMIT 5000
    """),
        params,
    ).fetchall()

    # Group orders by date
    from collections import defaultdict

    orders_by_date = defaultdict(list)
    for r in rows:
        order_date = r[9]
        if order_date:
            date_key = (
                order_date.strftime("%Y-%m-%d")
                if hasattr(order_date, "strftime")
                else str(order_date)[:10]
            )
            orders_by_date[date_key].append(r)

    wb = openpyxl.Workbook()

    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    headers = [
        "Order ID",
        "Order #",
        "Customer Email",
        "Customer Name",
        "Phone",
        "Total (₹)",
        "Payment Method",
        "POD / Tracking No.",
        "Courier Service",
        "Shipping Address",
        "Order Date",
    ]

    header_fill = PatternFill(
        start_color="180F14", end_color="180F14", fill_type="solid"
    )
    header_font = Font(bold=True, color="F2C29A")

    # Create a sheet for each date, sorted by date
    for date_key in sorted(orders_by_date.keys()):
        date_orders = orders_by_date[date_key]

        # Create sheet with date as name (truncate if too long)
        sheet_name = date_key[:10]  # YYYY-MM-DD
        # Handle sheet name length limit (31 chars max)
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]
        ws = wb.create_sheet(title=sheet_name)

        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Write data
        # SQL column indices: 0=id, 1=email, 2=customer_name, 3=customer_phone,
        # 4=total_amount, 5=payment_method, 6=status, 7=tracking_number, 8=courier_name,
        # 9=shipping_address, 10=created_at
        for row_idx, r in enumerate(date_orders, 2):
            order_id = r[0]
            ws.cell(row=row_idx, column=1, value=order_id)
            ws.cell(row=row_idx, column=2, value=f"ORD-{order_id:06d}")
            ws.cell(row=row_idx, column=3, value=r[1])  # email
            ws.cell(row=row_idx, column=4, value=r[2])  # customer_name
            ws.cell(row=row_idx, column=5, value=r[3] or "")  # customer_phone
            ws.cell(row=row_idx, column=6, value=float(r[4] or 0))  # total_amount
            ws.cell(row=row_idx, column=7, value=r[5])  # payment_method
            ws.cell(row=row_idx, column=8, value=r[7])  # tracking_number / POD
            ws.cell(row=row_idx, column=9, value=r[8] or "")  # courier_name
            ws.cell(row=row_idx, column=10, value=r[9])  # shipping_address
            ws.cell(
                row=row_idx, column=11, value=str(r[10])[:19] if r[10] else ""
            )  # created_at

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    # If no orders, create empty sheet
    if not orders_by_date:
        ws = wb.create_sheet(title="No Orders")
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"orders_{status or 'all'}_{now_ist().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ==================== POD Excel Upload ====================


@router.post("/api/v1/admin/orders/upload-pod-excel", tags=["Admin"])
async def upload_pod_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_staff),
):
    """
    Upload Excel file with columns: Order ID, POD / Tracking No.
    Batch-updates tracking_number + status=shipped for confirmed orders.
    Returns: updated_count, skipped, errors list.
    """
    try:
        import io, openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx files are accepted")

    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    ws = wb.active

    # Detect header row — look for "Order ID", "POD", and optional "Courier" columns
    header_map = {}
    for col_idx, cell in enumerate(ws[1], 1):
        val = str(cell.value or "").strip().lower()
        if "order id" in val or val == "order #" or val == "id":
            header_map["order_id_col"] = col_idx
        if "pod" in val or "tracking" in val:
            header_map["pod_col"] = col_idx
        if "courier" in val:
            header_map["courier_col"] = col_idx

    if "order_id_col" not in header_map or "pod_col" not in header_map:
        raise HTTPException(
            status_code=400,
            detail="Excel must have columns: 'Order ID' and 'POD / Tracking No.'",
        )

    updated, skipped, errors = 0, 0, []
    staff_id = current_user.get("user_id")

    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            raw_id = row[header_map["order_id_col"] - 1]
            raw_pod = row[header_map["pod_col"] - 1]
            # Extract courier_name if column exists
            courier_name = None
            if "courier_col" in header_map:
                courier_name = (
                    str(row[header_map["courier_col"] - 1] or "").strip() or None
                )

            if raw_id is None:
                continue
            # Handle "ORD-000001" format
            order_id_str = str(raw_id).strip().replace("ORD-", "").lstrip("0") or "0"
            order_id = int(order_id_str)
            pod = str(raw_pod or "").strip()
            if not pod:
                skipped += 1
                continue

            order = db.execute(
                text("SELECT id, status FROM orders WHERE id = :oid"), {"oid": order_id}
            ).fetchone()

            if not order:
                errors.append(f"Order #{order_id}: not found")
                continue
            if str(order[1]) not in ("confirmed",):
                errors.append(
                    f"Order #{order_id}: status is '{order[1]}', can only ship confirmed orders"
                )
                skipped += 1
                continue

            db.execute(
                text("""
                UPDATE orders
                SET status = 'shipped', tracking_number = :pod, courier_name = :courier_name,
                    shipped_at = :now, updated_at = :now
                WHERE id = :oid
            """),
                {
                    "pod": pod,
                    "courier_name": courier_name,
                    "now": now_ist(),
                    "oid": order_id,
                },
            )

            db.execute(
                text("""
                INSERT INTO order_tracking (order_id, status, notes, updated_by, created_at, courier_name)
                VALUES (:oid, 'shipped', :notes, :by, :now, :courier_name)
            """),
                {
                    "oid": order_id,
                    "notes": f"Shipped via Excel upload. POD: {pod}"
                    + (f", Courier: {courier_name}" if courier_name else ""),
                    "by": staff_id,
                    "now": now_ist(),
                    "courier_name": courier_name,
                },
            )
            updated += 1
        except (ValueError, IndexError) as e:
            errors.append(f"Row error: {e}")

    db.commit()
    return {
        "updated": updated,
        "skipped": skipped,
        "errors": errors[:20],
        "message": f"Successfully shipped {updated} order(s). {skipped} skipped.",
    }


# ==================== Order Deletion ====================


@router.delete("/api/v1/admin/orders/{order_id}", tags=["Admin"])
async def delete_order(
    order_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_admin),
):
    """Delete an order and restore inventory quantities."""
    # Check if order exists
    order = db.execute(
        text("SELECT id, status FROM orders WHERE id = :oid"), {"oid": order_id}
    ).first()

    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    # Prevent deletion of shipped/delivered orders
    if order[1] in ["shipped", "delivered"]:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot delete order with status '{order[1]}'. Cancel it first.",
        )

    try:
        # Restore inventory quantities before deleting
        items = db.execute(
            text("""
                SELECT oi.product_id, oi.size, oi.color, oi.quantity, oi.variant_id
                FROM order_items oi
                WHERE oi.order_id = :oid
            """),
            {"oid": order_id},
        ).fetchall()

        # Restore inventory for each item
        for item in items:
            db.execute(
                text("""
                    UPDATE inventory
                    SET quantity = quantity + :qty
                    WHERE id = :vid
                """),
                {
                    "qty": item[3],  # quantity
                    "vid": item[4],  # variant_id
                },
            )

        # Delete related records
        db.execute(
            text("DELETE FROM order_items WHERE order_id = :oid"), {"oid": order_id}
        )
        db.execute(
            text("DELETE FROM payment_transactions WHERE order_id = :oid"),
            {"oid": order_id},
        )
        db.execute(
            text("DELETE FROM order_tracking WHERE order_id = :oid"), {"oid": order_id}
        )

        # Finally delete the order
        db.execute(text("DELETE FROM orders WHERE id = :oid"), {"oid": order_id})

        db.commit()

        # Clear cache
        redis_client.delete_cache(f"order:{order_id}")
        redis_client.delete_cache("admin:orders:*")

        log_business_event(
            action="order_deleted",
            user_id=current_user.get("user_id"),
            details={"order_id": order_id, "restored_inventory": True},
        )

        return {"message": "Order deleted successfully", "order_id": order_id}
    except Exception as e:
        db.rollback()
        logger.error(f"Error deleting order {order_id}: {e}")
        raise HTTPException(status_code=500, detail="Failed to delete order")


@router.post("/api/v1/admin/orders/bulk-delete", tags=["Admin"])
async def bulk_delete_orders(
    payload: dict,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_admin),
):
    """Bulk delete multiple orders."""
    order_ids = payload.get("order_ids", [])
    if not order_ids:
        raise HTTPException(status_code=400, detail="No order IDs provided")

    deleted = 0
    failed = []

    for oid in order_ids:
        try:
            # Check status
            order = db.execute(
                text("SELECT id, status FROM orders WHERE id = :oid"), {"oid": oid}
            ).first()

            if not order:
                failed.append({"id": oid, "reason": "Not found"})
                continue

            if order[1] in ["shipped", "delivered"]:
                failed.append({"id": oid, "reason": f"Cannot delete {order[1]} order"})
                continue

            # Restore inventory
            items = db.execute(
                text(
                    "SELECT variant_id, quantity FROM order_items WHERE order_id = :oid"
                ),
                {"oid": oid},
            ).fetchall()

            for item in items:
                db.execute(
                    text(
                        "UPDATE inventory SET quantity = quantity + :qty WHERE id = :vid"
                    ),
                    {"qty": item[1], "vid": item[0]},
                )

            # Delete related records
            db.execute(
                text("DELETE FROM order_items WHERE order_id = :oid"), {"oid": oid}
            )
            db.execute(
                text("DELETE FROM payment_transactions WHERE order_id = :oid"),
                {"oid": oid},
            )
            db.execute(
                text("DELETE FROM order_tracking WHERE order_id = :oid"), {"oid": oid}
            )
            db.execute(text("DELETE FROM orders WHERE id = :oid"), {"oid": oid})

            deleted += 1
        except Exception as e:
            failed.append({"id": oid, "reason": str(e)})
            logger.error(f"Bulk delete failed for order {oid}: {e}")

    db.commit()
    redis_client.delete_cache("admin:orders:*")

    return {
        "deleted": deleted,
        "failed": failed,
        "message": f"Deleted {deleted} order(s)",
    }


@router.get("/api/v1/admin/orders/pod-template", tags=["Admin"])
async def download_pod_template(
    db: Session = Depends(get_db), current_user: dict = Depends(require_staff)
):
    """Download an Excel template pre-filled with confirmed orders for POD entry."""
    try:
        import io, openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from fastapi.responses import StreamingResponse
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    rows = db.execute(
        text("""
        SELECT o.id, u.email, u.username, o.total_amount, o.created_at,
               o.shipping_address
        FROM orders o
        LEFT JOIN users u ON u.id = o.user_id
        WHERE o.status = 'confirmed'
        ORDER BY o.created_at DESC
        LIMIT 500
    """)
    ).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "POD Upload"

    headers = [
        "Order ID",
        "Order #",
        "Customer Email",
        "Customer Name",
        "Total (₹)",
        "Order Date",
        "Shipping Address",
        "POD / Tracking No.",
        "Courier Name",
    ]
    hfill = PatternFill(start_color="180F14", end_color="180F14", fill_type="solid")
    hfont = Font(bold=True, color="F2C29A")

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center")

    # POD and Courier column highlights
    pod_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")

    for ri, r in enumerate(rows, 2):
        ws.cell(ri, 1, value=r[0])
        ws.cell(ri, 2, value=f"ORD-{r[0]:06d}")
        ws.cell(ri, 3, value=r[1])
        ws.cell(ri, 4, value=r[2])
        ws.cell(ri, 5, value=float(r[3] or 0))
        ws.cell(ri, 6, value=str(r[4])[:10] if r[4] else "")
        ws.cell(ri, 7, value=str(r[5] or "")[:80])
        pod_cell = ws.cell(ri, 8, value="")
        pod_cell.fill = pod_fill
        courier_cell = ws.cell(ri, 9, value="")
        courier_cell.fill = pod_fill  # Same highlight for both input fields

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    fname = f"pod_template_{now_ist().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )
